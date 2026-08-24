"""Importer for the five-column 2026 Ground Floor tile catalog workbooks.

Each source follows ``SR. / NAME / IMAGE / SIZE / RATE PER PCS``.  The final
heading is not wholly reliable: some suppliers provide a ``PER SQFT`` rate in
that column, so the exact supplied rate and its unit are retained in ``specs``.
"""
from __future__ import annotations

import io
import re

from ..base import MISSING, BrandAdapter, ExtractionReport, ProductRow, dedupe_iter
from ..image_extractor import ExtractedImage, extract_images_from_xlsx_ex

CATEGORY = "Tiles"
RATE_RE = re.compile(r"^\s*([\d,]+(?:\.\d+)?)\s*PER\s*(PCS|SQFT)\s*$", re.I)


def _slug(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "-", value.lower()).strip("-")


def _compact(value: str) -> str:
    return re.sub(r"[^A-Z0-9]+", "", value.upper())


class TileCatalog2026Adapter(BrandAdapter):
    """Parse a named supplier workbook without fabricating product details."""

    def __init__(self, brand: str) -> None:
        self.brand = brand

    def extract(self, data: bytes, filename: str) -> tuple[list[ProductRow], ExtractionReport]:
        report = ExtractionReport(brand=self.brand, filename=filename, source_type="excel")
        try:
            from openpyxl import load_workbook
            workbook = load_workbook(io.BytesIO(data), data_only=True)
        except Exception as error:  # pragma: no cover - parsing boundary
            report.warnings.append(f"xlsx open failed: {error}")
            return [], report

        rows: list[ProductRow] = []
        sku_occurrences: dict[str, int] = {}
        quality_rank = {"excellent": 4, "good": 3, "acceptable": 2, "poor": 1}
        for worksheet in workbook.worksheets:
            values = [list(row) for row in worksheet.iter_rows(values_only=True)]
            if not values:
                continue
            headers = {str(value or "").strip().lower(): index for index, value in enumerate(values[0])}
            # Panda's workbook uses the same five-column tile layout as this
            # adapter, but labels the fields "PRODUCT NAME" and
            # "PRODUCT SIZE". Keep the source labels intact while resolving
            # those aliases to the canonical parser keys.
            if "name" not in headers and "product name" in headers:
                headers["name"] = headers["product name"]
            if "size" not in headers and "product size" in headers:
                headers["size"] = headers["product size"]
            required = {"name", "size"}
            if not required <= headers.keys():
                report.warnings.append(f"Sheet {worksheet.title!r}: expected NAME and SIZE columns")
                continue
            rate_column = next((index for label, index in headers.items() if "rate" in label), None)
            if rate_column is None:
                report.warnings.append(f"Sheet {worksheet.title!r}: expected a RATE column")
                continue
            image_column = headers.get("image")
            images_by_row: dict[int, ExtractedImage] = {}
            for sheet, row, column, image in extract_images_from_xlsx_ex(data, optimize=False):
                if sheet != worksheet.title or (image_column is not None and column != image_column):
                    continue
                previous = images_by_row.get(row)
                if previous is None or quality_rank.get(image.quality, 0) > quality_rank.get(previous.quality, 0):
                    images_by_row[row] = image
            report.images_found += len(images_by_row)

            for row_number, row in enumerate(values[1:], start=2):
                name = str(row[headers["name"]] or "").strip()
                size = str(row[headers["size"]] or "").strip()
                if not name or not size:
                    continue
                raw_rate = str(row[rate_column] or "").strip()
                match = RATE_RE.match(raw_rate)
                price = float(match.group(1).replace(",", "")) if match else None
                unit = match.group(2).lower() if match else None
                base_sku = f"{_compact(self.brand)}-{_compact(name)[:36] or 'PRODUCT'}-{_compact(size)[:16] or 'SIZE'}"
                occurrence = sku_occurrences.get(base_sku, 0) + 1
                sku_occurrences[base_sku] = occurrence
                sku = base_sku if occurrence == 1 else f"{base_sku}-{occurrence}"
                image = images_by_row.get(row_number)
                if image:
                    report.images_mapped += 1
                product = ProductRow(
                    brand=self.brand, sku=sku, name=f"{name} ({size})", category=CATEGORY,
                    family_key=f"{_slug(self.brand)}:{_slug(name)}", variant=size, size=size,
                    mrp=price if price is not None else 0.0,
                    dealer_price=price if price is not None else 0.0,
                    images=[image.data_url] if image else [], image_meta=[image.to_dict()] if image else [],
                    image_quality=image.quality if image else "missing",
                    specs={"source_file": filename, "source_row": row_number,
                           "source_rate": raw_rate, "price_unit": unit,
                           "needs_pricing": price is None},
                    tags=dedupe_iter(["tiles", _slug(self.brand), unit or "rate-review"]),
                    confidence=0.95 if price is not None else 0.6,
                )
                if price is None:
                    product.issues.append(f"Unrecognized source rate {raw_rate!r}; imported at ₹0 for review")
                if not image:
                    product.issues.append("No image mapped from supplier file")
                if occurrence > 1:
                    product.issues.append(f"Duplicate listing occurrence {occurrence}; SKU suffixed to retain it")
                rows.append(product)
        report.parsed_rows = len(rows)
        return rows, report


class PandaAdapter(TileCatalog2026Adapter):
    """Panda 2026 tile pricelist (PRODUCT NAME / PRODUCT SIZE / RATE)."""

    def __init__(self) -> None:
        super().__init__("Panda")
