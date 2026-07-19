"""OYSTER XLSX adapter (Brook CP Fittings collection, sold under the Oyster brand).

Ships as one workbook per category (Body Jet / Shower / Outlet+HS+Angle Valve /
Basin Mixer), single sheet each, one row per finish variant:

  Row 1  : title ("OYSTER <CATEGORY>")
  Row 2  : column headers (Sr. No., finishes, [Article No. — Shower only],
           Product Discription, Product Image, MRP, QTY, MRP TOTAL, DISCOUNT,
           <blank>, OFFER RATE, TOTAL)
  Row 3+ : one finish variant per row. The embedded product photo floats over
           the "Product Image" cell of that exact row — verified 1:1 row
           anchoring across all 4 files (no side-by-side finishes sharing a
           row, unlike Vitra), so row-only anchor matching (Hansgrohe-style)
           is safe here.

Rules:
* Category comes from the source FILENAME (one file = one category, per
  business rule) — never inferred from row content.
* Brand is always "Oyster"; collection is always "Brook" (the sub-line name
  embedded in every "Product Discription" cell, e.g. "Brook CP Fittings ...").
* "finishes" column is heavily typo'd by the supplier (CROME/CEROME/MAAT
  BALCK/etc.) — normalized via an explicit FINISH_LOOKUP table (not fuzzy
  regex) built from the 18 distinct raw strings actually observed across all
  4 files. Anything new is flagged for manual review, never guessed.
* No real manufacturer article numbers exist anywhere in these files. The
  Shower sheet's "Article No." column actually holds a size string (e.g.
  "1000 x 700"), not a code — recovered into `dimensions` instead of
  discarded. SKU is therefore synthesized, not supplier-provided:
  `OYSTER-{CATEGORY_COMPACT}-{FAMILY_COMPACT}-{FINISH_CODE}`, deterministic
  from category+family+finish so re-running the import always regenerates
  the same SKU (required for idempotency). A handful of rows in the source
  files are exact duplicate listings (same product+finish+price repeated on
  a later row) — the first occurrence keeps the plain SKU and each repeat
  gets a numeric suffix (`-2`, `-3`, ...) so it still imports as its own
  product instead of being dropped. Rows with no MRP anywhere in the source
  import at MRP/price = 0.0 with an explicit `needs_pricing` flag (in both
  `specs` and `tags`) rather than being held back.
* Family key groups sibling finishes of the same product:
  `oyster:{category_slug}:{family_slug}`.
* "OFFER RATE" column is the dealer/selling price (supplier-computed as
  MRP × (1 − discount%); falls back to MRP when no discount was applied).
"""
from __future__ import annotations
import io
import re

from ..base import MISSING, BrandAdapter, ExtractionReport, ProductRow, dedupe_iter
from ..image_extractor import ExtractedImage, extract_images_from_xlsx_ex

BRAND = "Oyster"
COLLECTION_NAME = "Brook"

# Filename -> (category display name, compact SKU code). Order matters only
# in that patterns must stay mutually exclusive for the 4 real filenames.
FILE_TO_CATEGORY: list[tuple[re.Pattern, str, str]] = [
    (re.compile(r"body\s*jet", re.I), "Body Jet", "BODYJET"),
    (re.compile(r"shower", re.I), "Shower", "SHOWER"),
    (re.compile(r"spout|angle|tigger|hs\s*&", re.I), "Outlet / Hand Shower / Angle Valve", "OUTLETHSANGLE"),
    (re.compile(r"besin|basin", re.I), "Basin Mixer", "BASINMIXER"),
]

# Explicit lookup, not regex guessing — built from the 18 distinct raw
# "finishes" cell values actually present across all 4 Oyster files
# (verified by direct inspection before writing this table).
FINISH_LOOKUP: dict[str, tuple[str, str]] = {
    "CROME": ("Chrome", "CR"),
    "CHROME": ("Chrome", "CR"),
    "CEROME": ("Chrome", "CR"),
    "CEOME": ("Chrome", "CR"),
    "MAAT BLACK": ("Matt Black", "MB"),
    "MAAT BALCK": ("Matt Black", "MB"),
    "MATE BLACK": ("Matt Black", "MB"),
    "MATT BLACK": ("Matt Black", "MB"),
    "BRUSHED GOLD": ("Brushed Gold", "BG"),
    "ROSE GOLD": ("Rose Gold", "RG"),
    "BRUSHED ROSE GOLD": ("Brushed Rose Gold", "BRG"),
    "BRUSHE ROSE GLOD": ("Brushed Rose Gold", "BRG"),
    "BRUSHED GUN METAL": ("Brushed Gun Metal", "BGM"),
    "GUN METAL": ("Gun Metal", "GM"),
}


def category_from_filename(filename: str) -> tuple[str, str]:
    for pat, label, code in FILE_TO_CATEGORY:
        if pat.search(filename or ""):
            return label, code
    raise ValueError(f"Oyster adapter: cannot map filename {filename!r} to a known category")


def normalize_finish(raw: str) -> tuple[str | None, str | None, str | None]:
    """Returns (finish_label, finish_code, note). `note` is set either when a
    corrupted cell was repaired (informational) or when the value is
    unrecognized (needs manual review) — callers should surface it either
    way rather than silently swallowing it."""
    s = str(raw or "").replace("\xa0", " ")
    repaired_from = None
    if "+" in s:
        # Repairs the two known corrupted merged-cell artifacts in the
        # source files: "CROME+B3:E16" / "CROME+B3:L44" — everything before
        # the "+" is the real finish name.
        repaired_from = s
        s = s.split("+", 1)[0]
    s = re.sub(r"\s+", " ", s).strip().upper()
    hit = FINISH_LOOKUP.get(s)
    if hit:
        note = f"corrupted finish cell {repaired_from!r} repaired to {s!r}" if repaired_from else None
        return hit[0], hit[1], note
    return None, None, f"unrecognized finish {raw!r} — needs manual review"


def _slug(s: str) -> str:
    return re.sub(r"[^a-z0-9]+", "-", (s or "").lower()).strip("-")


def _compact(s: str) -> str:
    return re.sub(r"[^A-Z0-9]", "", (s or "").upper())


def family_key_for(category_slug: str, family_name: str) -> str:
    return f"oyster:{category_slug}:{_slug(family_name)}"


def sku_for(category_compact: str, family_name: str, finish_code: str) -> str:
    family_compact = _compact(family_name)[:32] or "FAMILY"
    return f"OYSTER-{category_compact}-{family_compact}-{finish_code}"


# "Brook CP FITTINGS ..." / "BROOK UP FITINGS ..." prefix appears (with
# inconsistent spacing/typos) at the start of every "Product Discription"
# cell across all 4 files — strip it to get a clean family display name.
_COLLECTION_PREFIX_RE = re.compile(r"^\s*BROOK\s+(CP|UP)\s+FIT[TI]*INGS?\s*[-:]?\s*", re.I)


def _family_name(raw_description) -> str:
    """The Shower sheet ships a two-line description
    ("<collection line>\\n<product line>") for every row; the other 3 files
    ship one line with the "Brook CP/UP Fit(t)ings" prefix inline. Either
    way, strip the collection prefix to get the family display name."""
    s = str(raw_description or "").strip()
    if "\n" in s:
        _, _, rest = s.partition("\n")
        s = rest.strip() or s
    # Collection prefix stripping applies unconditionally to the final
    # description string, whether it came from a single-line source (Body
    # Jet, Basin Mixer, Outlet/HS/Angle) or the post-newline-split remainder
    # of a multi-line source (Shower) — the "Brook" collection is already
    # captured separately in every row's `collection` field, so repeating it
    # in the family name/SKU/family_key would be redundant.
    s = _COLLECTION_PREFIX_RE.sub("", s).strip()
    s = re.sub(r"\s+", " ", s)
    return s or "Unnamed"


def _find_header_row(all_rows: list[list]) -> int:
    for i, r in enumerate(all_rows[:5]):
        cells = [str(c or "").lower().replace("\n", " ").strip() for c in r]
        if "finishes" in cells:
            return i
    return 1  # fallback: row 2 (0-indexed 1), the observed layout in all 4 files


def _build_col_map(header_row: list) -> dict[str, int]:
    col_map: dict[str, int] = {}
    for c, val in enumerate(header_row):
        key = str(val or "").lower().replace("\n", " ").strip()
        if not key:
            continue
        if key == "finishes":
            col_map["finish"] = c
        elif key.startswith("article"):
            # Shower-only column; its content is actually a size string
            # (e.g. "1000 x 700"), recovered into `dimensions`.
            col_map["article_no"] = c
        elif "discription" in key or "description" in key:
            col_map["description"] = c
        elif "image" in key:
            col_map["image"] = c
        elif key == "mrp":
            col_map["mrp"] = c
        elif key == "offer rate":
            col_map["offer_rate"] = c
    return col_map


class OysterAdapter(BrandAdapter):
    brand = BRAND
    supported_extensions = (".xlsx", ".xls")

    def extract(self, data: bytes, filename: str) -> tuple[list[ProductRow], ExtractionReport]:
        report = ExtractionReport(brand=self.brand, filename=filename, source_type="excel")
        rows: list[ProductRow] = []
        try:
            from openpyxl import load_workbook
        except Exception as e:  # pragma: no cover
            report.warnings.append(f"openpyxl missing: {e}")
            return rows, report

        try:
            category, category_code = category_from_filename(filename)
        except ValueError as e:
            report.warnings.append(str(e))
            return rows, report
        category_slug = _slug(category)

        try:
            wb = load_workbook(io.BytesIO(data), data_only=True)
        except Exception as e:
            report.warnings.append(f"xlsx open failed: {e}")
            return rows, report

        # Image mapping: one photo per data row — verified 1:1 across all 4
        # files (no side-by-side finishes sharing a row, unlike Vitra), so
        # row-only anchoring (Hansgrohe-style) is safe. Highest-quality
        # candidate wins if an anchor somehow has more than one blip.
        sku_counts: dict[str, int] = {}
        by_sheet_row: dict[tuple[str, int], ExtractedImage] = {}
        for sheet, row_idx, _col_idx, img in extract_images_from_xlsx_ex(data):
            key = (sheet, row_idx)
            prev = by_sheet_row.get(key)
            qrank = {"excellent": 4, "good": 3, "acceptable": 2, "poor": 1}
            if prev is None or (qrank.get(img.quality, 0), img.longest_edge) > (qrank.get(prev.quality, 0), prev.longest_edge):
                by_sheet_row[key] = img
        report.images_found = len(by_sheet_row)

        for ws in wb.worksheets:
            all_rows = [list(r) for r in ws.iter_rows(values_only=True)]
            if not all_rows:
                continue
            hdr_idx = _find_header_row(all_rows)
            col_map = _build_col_map(all_rows[hdr_idx])
            if not {"finish", "description", "mrp"} <= col_map.keys():
                report.warnings.append(
                    f"Sheet {ws.title!r} in {filename}: could not locate finish/description/MRP columns"
                )
                continue

            for r_idx, row in enumerate(all_rows[hdr_idx + 1:], start=hdr_idx + 2):
                desc_cell = row[col_map["description"]] if col_map["description"] < len(row) else None
                finish_cell = row[col_map["finish"]] if col_map["finish"] < len(row) else None
                if not desc_cell or not finish_cell:
                    continue  # blank / section / trailer row

                family_name = _family_name(desc_cell)
                finish_label, finish_code, finish_note = normalize_finish(finish_cell)

                mrp_raw = self.to_number(row[col_map["mrp"]] if col_map["mrp"] < len(row) else None)
                needs_pricing = mrp_raw == MISSING
                mrp = 0.0 if needs_pricing else mrp_raw
                offer_rate = self.to_number(
                    row[col_map["offer_rate"]] if "offer_rate" in col_map and col_map["offer_rate"] < len(row) else None
                )
                dealer_price = offer_rate if offer_rate not in (MISSING, None, 0) else mrp

                dimensions = MISSING
                if "article_no" in col_map:
                    art_cell = row[col_map["article_no"]] if col_map["article_no"] < len(row) else None
                    if art_cell:
                        dimensions = str(art_cell).strip()

                family_key = family_key_for(category_slug, family_name)
                is_duplicate_listing = False
                occurrence = 1
                if finish_code:
                    base_sku = sku_for(category_code, family_name, finish_code)
                    occurrence = sku_counts.get(base_sku, 0) + 1
                    sku_counts[base_sku] = occurrence
                    sku = base_sku if occurrence == 1 else f"{base_sku}-{occurrence}"
                    is_duplicate_listing = occurrence > 1
                else:
                    sku = MISSING

                img = by_sheet_row.get((ws.title, r_idx))
                image_urls = [img.data_url] if img else []
                image_meta = [img.to_dict()] if img else []
                image_quality = img.quality if img else "missing"
                if img:
                    report.images_mapped += 1

                pr = ProductRow(
                    brand=self.brand,
                    sku=sku,
                    name=f"{family_name} - {finish_label}" if finish_label else family_name,
                    category=category,
                    subcategory=MISSING,
                    series=MISSING,
                    family_key=family_key,
                    variant=finish_label or MISSING,
                    finish=finish_label or MISSING,
                    finish_code=finish_code or MISSING,
                    colour=finish_label or MISSING,
                    material=MISSING,
                    dimensions=dimensions,
                    description=str(desc_cell).strip(),
                    mrp=mrp,
                    dealer_price=dealer_price,
                    warranty=MISSING,
                    collection=COLLECTION_NAME,
                    images=image_urls,
                    image_meta=image_meta,
                    image_quality=image_quality,
                    image_page=None,
                    specs={
                        "collection": COLLECTION_NAME,
                        "source_file": filename,
                        **({"needs_pricing": True} if needs_pricing else {}),
                        **({"duplicate_listing": True} if is_duplicate_listing else {}),
                    },
                    tags=dedupe_iter([
                        category.lower(), self.brand.lower(), COLLECTION_NAME.lower(), (finish_label or "").lower(),
                        *(["needs-pricing"] if needs_pricing else []),
                    ]),
                    confidence=0.94 if finish_label else 0.5,
                )
                if not finish_label:
                    pr.issues.append(finish_note or f"Unrecognized finish {finish_cell!r} — needs manual review")
                elif finish_note:
                    pr.issues.append(finish_note)
                if needs_pricing:
                    pr.issues.append("Missing MRP in source — imported at ₹0, needs manual pricing")
                if is_duplicate_listing:
                    pr.issues.append(f"Duplicate listing in source file (occurrence {occurrence}) — SKU suffixed to keep as a separate product")
                if not img:
                    pr.issues.append("No image mapped from supplier file")
                rows.append(pr)

        report.parsed_rows = len(rows)
        return rows, report
