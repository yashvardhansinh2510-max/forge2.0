"""RENITE XLSX adapter for the Ground Floor tiles catalog.

The 2026 supplier workbook has one row for each tile variant and the columns
``PRODUCT NAME``, ``PRODUCT SIZE``, ``FINISHES``, box quantities, and ``RATE``.
It contains no SKU and no embedded product photographs, so this adapter creates
stable RENITE SKUs and records the missing-image condition for later enrichment.
"""
from __future__ import annotations

import io
import re

from ..base import MISSING, BrandAdapter, ExtractionReport, ProductRow, dedupe_iter
from ..image_extractor import ExtractedImage, extract_images_from_xlsx_ex

BRAND = "Renite"
CATEGORY = "Tiles"
FINISH_LOOKUP = {"MATT": ("Matt", "MT")}
_RATE_RE = re.compile(r"^\s*([\d,]+(?:\.\d+)?)\s*PER\s*SQFT\s*$", re.I)


def _slug(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "-", (value or "").lower()).strip("-")


def _compact(value: str) -> str:
    return re.sub(r"[^A-Z0-9]", "", (value or "").upper())


def family_key_for(name: str) -> str:
    return f"renite:{_slug(name)}"


def sku_for(name: str, size: str, finish_code: str) -> str:
    return f"RENITE-{_compact(name)[:32] or 'PRODUCT'}-{_compact(size)[:16] or 'SIZE'}-{finish_code}"


def normalize_finish(raw: str) -> tuple[str | None, str | None, str | None]:
    normalized = re.sub(r"\s+", " ", str(raw or "").replace("\xa0", " ")).strip().upper()
    match = FINISH_LOOKUP.get(normalized)
    if match:
        return match[0], match[1], None
    return None, None, f"unrecognized finish {raw!r} — needs manual review"


def parse_rate_per_sqft(raw: object) -> tuple[float | None, str | None]:
    match = _RATE_RE.match(str(raw or "").strip())
    if not match:
        return None, f"unrecognized RATE format {raw!r} — expected '<number> PER SQFT'"
    return float(match.group(1).replace(",", "")), None


def _find_header_row(rows: list[list]) -> int:
    return next((i for i, row in enumerate(rows[:5]) if any("product name" in str(cell or "").lower() for cell in row)), 0)


def _column_map(header: list) -> dict[str, int]:
    result: dict[str, int] = {}
    for index, value in enumerate(header):
        key = str(value or "").lower().strip()
        if "product name" in key:
            result["name"] = index
        elif key == "image" or "image" in key:
            result["image"] = index
        elif "size" in key:
            result["size"] = index
        elif "finish" in key:
            result["finish"] = index
        elif "box" in key and ("pc" in key or "pis" in key):
            result["pcs_per_box"] = index
        elif "box" in key and "sqft" in key:
            result["sqft_per_box"] = index
        elif "rate" in key:
            result["rate"] = index
    return result


def _format_pieces(value: object) -> str | None:
    if value in (None, MISSING):
        return None
    try:
        numeric = float(value)
        return str(int(numeric)) if numeric.is_integer() else str(numeric)
    except (TypeError, ValueError):
        return str(value)


class ReniteAdapter(BrandAdapter):
    brand = BRAND
    supported_extensions = (".xlsx", ".xls")

    def extract(self, data: bytes, filename: str) -> tuple[list[ProductRow], ExtractionReport]:
        report = ExtractionReport(brand=self.brand, filename=filename, source_type="excel")
        try:
            from openpyxl import load_workbook
            workbook = load_workbook(io.BytesIO(data), data_only=True)
        except Exception as error:  # pragma: no cover - defensive parsing boundary
            report.warnings.append(f"xlsx open failed: {error}")
            return [], report

        rows: list[ProductRow] = []
        sku_occurrences: dict[str, int] = {}
        quality_rank = {"excellent": 4, "good": 3, "acceptable": 2, "poor": 1}
        for worksheet in workbook.worksheets:
            raw_rows = [list(row) for row in worksheet.iter_rows(values_only=True)]
            if not raw_rows:
                continue
            header_index = _find_header_row(raw_rows)
            columns = _column_map(raw_rows[header_index])
            required = {"name", "size", "finish", "rate"}
            if not required <= columns.keys():
                report.warnings.append(f"Sheet {worksheet.title!r}: required tile columns were not found")
                continue

            image_column = columns.get("image")
            images_by_row: dict[int, ExtractedImage] = {}
            for sheet_name, row_index, column_index, image in extract_images_from_xlsx_ex(data, optimize=False):
                if sheet_name != worksheet.title or (image_column is not None and column_index != image_column):
                    continue
                previous = images_by_row.get(row_index)
                if previous is None or (quality_rank.get(image.quality, 0), image.longest_edge) > (quality_rank.get(previous.quality, 0), previous.longest_edge):
                    images_by_row[row_index] = image
            report.images_found += len(images_by_row)

            for spreadsheet_row, row in enumerate(raw_rows[header_index + 1:], start=header_index + 2):
                def cell(field: str):
                    index = columns.get(field)
                    return row[index] if index is not None and index < len(row) else None

                raw_name, raw_size, raw_finish = cell("name"), cell("size"), cell("finish")
                if not raw_name or not raw_size or not raw_finish:
                    continue
                name, size = str(raw_name).strip(), str(raw_size).strip()
                finish, finish_code, finish_note = normalize_finish(raw_finish)
                price, price_note = parse_rate_per_sqft(cell("rate"))
                base_sku = sku_for(name, size, finish_code or "UNK")
                occurrence = sku_occurrences.get(base_sku, 0) + 1
                sku_occurrences[base_sku] = occurrence
                sku = base_sku if occurrence == 1 else f"{base_sku}-{occurrence}"
                display_finish = finish or str(raw_finish).strip()
                needs_pricing = price is None
                image = images_by_row.get(spreadsheet_row)
                if image:
                    report.images_mapped += 1

                product = ProductRow(
                    brand=self.brand, sku=sku, name=f"{name} - {display_finish} ({size})",
                    category=CATEGORY, family_key=family_key_for(name),
                    variant=f"{size} · {display_finish}", finish=finish or display_finish,
                    finish_code=finish_code or MISSING, size=size,
                    mrp=price if price is not None else 0.0,
                    dealer_price=price if price is not None else 0.0,
                    images=[image.data_url] if image else [],
                    image_meta=[image.to_dict()] if image else [],
                    image_quality=image.quality if image else "missing",
                    specs={
                        "pcs_per_box": _format_pieces(cell("pcs_per_box")),
                        "sqft_per_box": self.to_number(cell("sqft_per_box")),
                        "rate_per_sqft": price,
                        "rate_source_text": str(cell("rate") or "").strip() or None,
                        "source_row": spreadsheet_row,
                        "source_file": filename,
                        **({"needs_pricing": True} if needs_pricing else {}),
                        **({"duplicate_listing": True} if occurrence > 1 else {}),
                    },
                    tags=dedupe_iter([CATEGORY.lower(), self.brand.lower(), (finish or "").lower()]),
                    confidence=0.95 if finish and not needs_pricing else 0.6,
                )
                if not finish:
                    product.issues.append(finish_note or "Unrecognized finish — needs manual review")
                if needs_pricing:
                    product.issues.append(price_note or "Missing price — imported at ₹0, needs manual pricing")
                if not image:
                    product.issues.append("No image mapped from supplier file")
                if occurrence > 1:
                    product.issues.append(f"Duplicate listing in source (occurrence {occurrence}); SKU suffixed")
                rows.append(product)

        report.parsed_rows = len(rows)
        return rows, report
