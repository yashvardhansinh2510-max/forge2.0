"""HANSGROHE XLSX adapter (parent brand: Hansgrohe; AXOR = premium collection under Hansgrohe).

Ships as per-file worksheets (Thermostat.xlsx, WBM.xlsx, 3hole.xlsx, handshower.xlsx,
SHOWERS_HANSGROHE.xlsx). Every sheet follows the same tabular layout:

  Row 1  : column headers (Sr. No., Article No., Product Description, Product Image, MRP, Qty)
  Row 2+ : product rows OR section header rows (e.g. "Master Bathroom Shower Area", "3Hole Basin Mixers")
           — section headers have text in col A only, no article number.

Rules:
* "HG "-prefixed product name → collection = "Hansgrohe"
* "AX "-prefixed product name → collection = "AXOR"
* Brand is ALWAYS "Hansgrohe" (AXOR is folded under Hansgrohe per business rule).
* Article number (Col B) IS the SKU — never modified, kept as string.
* Section headers found in col A while col B is empty become the subcategory
  for all following rows until the next section header.
* Category is derived from the source filename (Thermostat.xlsx → Thermostat, etc.).
* Series is parsed from the product name using an ordered priority list of
  known Hansgrohe / AXOR series names.
* Finish is parsed from the product name suffix (chrome, Matt Black, Brushed
  Nickel, PBB, BBR, PN, etc.). Finish code is the numeric segment of the SKU
  when applicable (the last 3 digits of an 8-digit article number encode
  Hansgrohe's finish family: 000=Chrome, 140=BBR (Brushed Bronze), 340=BN
  (Brushed Nickel), 700=Matt White, 990=Matt Black, 800=Polished Gold, etc.).
* Family key groups sibling finishes: `hansgrohe:{collection_slug}:{series_slug}:{base_name_slug}`.
"""
from __future__ import annotations
import io
import re

from ..base import MISSING, BrandAdapter, ExtractionReport, ProductRow, dedupe_iter
from ..image_extractor import ExtractedImage, extract_images_from_xlsx_ex


# Category is derived DIRECTLY from the source filename (per business rule):
#   BM.xlsx           → "BM"
#   Ceramic.xlsx      → "Ceramic"
#   HFAV.xlsx         → "HFAV"
#   Holder.xlsx       → "Holder"
#   Thermostat.xlsx   → "Thermostat"
#   WBM.xlsx          → "WBM"
#   3hole.xlsx        → "3hole"
#   handshower.xlsx   → "handshower"
#   SHOWERS_HANSGROHE.xlsx → "SHOWERS_HANSGROHE"
# The stem is used verbatim (whitespace collapsed) so that Forge's UI shows the
# supplier's file-level grouping as the category.
FILE_TO_CATEGORY: dict[str, tuple[str, str | None]] = {}  # kept for legacy imports; unused

# Section headers → subcategory
SECTION_TO_SUBCAT = [
    ("master bathroom shower area", None),           # generic ; keep default
    ("3hole basin mixers",          "3-Hole Basin Mixer"),
    ("hg & axor showers",           None),
    ("hand shower",                 "Hand Shower"),
    ("overhead",                    "Overhead Shower"),
    ("rain shower",                 "Rain Shower"),
    ("shower set",                  "Shower Set"),
    ("shower pipe",                 "Shower Pipe"),
    ("shower system",               "Shower System"),
    ("side shower",                 "Side Shower"),
    ("body jet",                    "Body Jet"),
    ("thermostat",                  "Thermostat"),
    ("valve",                       "Valve"),
    ("waste",                       "Waste System"),
    ("accessor",                    "Accessories"),
]

# Ordered series priority. Longest / most specific first.
KNOWN_SERIES = [
    # AXOR premium collections (checked first because AX prefix flags them)
    "AXOR Citterio E", "AXOR Citterio M", "AXOR Citterio C", "AXOR Citterio",
    "AXOR Starck V", "AXOR Starck X", "AXOR Starck Organic", "AXOR Starck",
    "AXOR Massaud", "AXOR Uno", "AXOR Urquiola", "AXOR Montreux",
    "AXOR Edge", "AXOR Universal", "AXOR ShowerCollection", "AXOR ShowerHeaven",
    "AXOR ShowerSelect", "AXOR Carlton", "AXOR MyEdition", "AXOR One",
    # Hansgrohe series
    "Ecostat Comfort", "Ecostat E", "Ecostat S", "Ecostat Fine",
    "Croma Select S", "Croma Select E", "Croma 100", "Croma 220",
    "Raindance Select S", "Raindance Select E", "Raindance E", "Raindance S",
    "Raindance Rainmaker", "Raindance Rainfall",
    "Pulsify Select S", "Pulsify Select E", "Pulsify S", "Pulsify E", "Pulsify",
    "Rainfinity",
    "Metropol Classic", "Metropol",
    "Metris S", "Metris E", "Metris Select", "Metris Classic", "Metris",
    "Talis Select S", "Talis Select E", "Talis S", "Talis E", "Talis M", "Talis",
    "Focus S", "Focus E", "Focus M", "Focus N", "Focus",
    "Logis Classic", "Logis Loop", "Logis",
    "Finoris",
    "Tecturis E", "Tecturis S", "Tecturis",
    "Vernis Shape", "Vernis Blend", "Vernis",
    "PuraVida",
    "ShowerSelect Comfort", "ShowerSelect S", "ShowerSelect E", "ShowerSelect",
    "ShowerTablet Select", "ShowerTablet",
    "iBox",
    "Sportive", "Novus", "MyClub", "MySport",
]

# Ordered finish patterns (most specific first)
FINISH_ALIASES: list[tuple[re.Pattern, str, str]] = [
    (re.compile(r"\b(matt|matte)\s*black\b", re.I),       "Matt Black",       "MB"),
    (re.compile(r"\b(matt|matte)\s*white\b", re.I),       "Matt White",       "MW"),
    (re.compile(r"\bpolished\s*gold\s*optic\b", re.I),    "Polished Gold Optic", "PGO"),
    (re.compile(r"\bpolished\s*gold\b|\bp\.?g\.?o?\.?\b", re.I), "Polished Gold", "PG"),
    (re.compile(r"\bpolished\s*black\s*chrome\b", re.I),  "Polished Black Chrome", "PBC"),
    (re.compile(r"\bbrushed\s*black\s*chrome\b", re.I),   "Brushed Black Chrome", "BBC"),
    (re.compile(r"\bblack\s*chrome\b", re.I),             "Black Chrome",     "BC"),
    (re.compile(r"\bbrushed\s*bronze\b|\bbbr\b", re.I),   "Brushed Bronze",   "BBR"),
    (re.compile(r"\bpolished\s*bronze\b|\bpbb\b", re.I),  "Polished Bronze",  "PBB"),
    (re.compile(r"\bbrushed\s*nickel\b|\bbn\b", re.I),    "Brushed Nickel",   "BN"),
    (re.compile(r"\bpolished\s*nickel\b|\bpn\b", re.I),   "Polished Nickel",  "PN"),
    (re.compile(r"\bstainless\s*steel\s*optic\b|\bsso\b", re.I), "Stainless Steel Optic", "SSO"),
    (re.compile(r"\bstainless\s*steel\b", re.I),          "Stainless Steel",  "SS"),
    (re.compile(r"\bbrushed\s*gold\s*optic\b|\bbgo\b", re.I), "Brushed Gold Optic", "BGO"),
    (re.compile(r"\bbrushed\s*chrome\b", re.I),           "Brushed Chrome",   "BC"),
    (re.compile(r"\bpolished\s*chrome\b", re.I),          "Polished Chrome",  "PC"),
    (re.compile(r"\bchrome\b|\bchr\.?\b|\bzh?\.?chrome\b", re.I), "Chrome",   "CR"),
    (re.compile(r"\bwhite\b", re.I),                       "White",           "WH"),
]

# Finish code from the last 3 digits of Hansgrohe SKUs
SKU_TAIL_FINISH = {
    "000": ("Chrome",         "CR"),
    "800": ("Polished Gold",  "PG"),
    "140": ("Brushed Bronze", "BBR"),
    "340": ("Brushed Nickel", "BN"),
    "700": ("Matt White",     "MW"),
    "670": ("Matt Black",     "MB"),
    "990": ("Matt Black",     "MB"),
    "820": ("Polished Gold Optic", "PGO"),
    "830": ("Brushed Gold Optic",  "BGO"),
    "310": ("Brushed Nickel", "BN"),
    "180": ("Chrome",         "CR"),
    "007": ("Chrome",         "CR"),
    "400": ("Stainless Steel Optic", "SSO"),
    "600": ("Matt Black",     "MB"),
    "090": ("Chrome/Gold-Optic","CR-PGO"),
    "930": ("Bronze",         "BR"),
    "950": ("Bronze",         "BR"),
    "460": ("Brushed Nickel", "BN"),
}

SIZE_RE = re.compile(r"\b(\d{2,4})(?:\s*[×xX]\s*(\d{2,4}))?\s*(?:mm|cm|inch|\")?\b")


def _slug(s: str) -> str:
    return re.sub(r"[^a-z0-9]+", "-", (s or "").lower()).strip("-")


def _category_from_filename(filename: str) -> tuple[str, str | None]:
    """Category = filename stem verbatim (per business rule); subcategory = None."""
    import os
    stem = os.path.splitext(os.path.basename(filename or ""))[0]
    stem = stem.strip().replace("_", " ").strip()
    return (stem or "Uncategorized", None)


def _detect_collection(name: str) -> str:
    """AX prefix or explicit AXOR mention → AXOR; else Hansgrohe."""
    if not name:
        return "Hansgrohe"
    n = name.upper()
    if n.startswith("AX ") or n.startswith("AXOR ") or " AXOR " in f" {n} ":
        return "AXOR"
    return "Hansgrohe"


def _detect_series(name: str) -> str | None:
    if not name:
        return None
    n = name.upper()
    for s in KNOWN_SERIES:
        if s.upper() in n:
            return s
    return None


def _detect_finish(name: str, sku: str) -> tuple[str | None, str | None]:
    for pat, label, code in FINISH_ALIASES:
        if pat.search(name or ""):
            return label, code
    # Fallback from SKU tail
    sku_s = str(sku or "")
    if sku_s.isdigit() and len(sku_s) >= 3:
        tail = sku_s[-3:]
        if tail in SKU_TAIL_FINISH:
            return SKU_TAIL_FINISH[tail]
    return (None, None)


def _base_name(name: str, series: str | None, finish: str | None) -> str:
    """Strip HG / AX prefix, series and finish tokens to derive the family base name."""
    s = (name or "").strip()
    # Strip HG / AX prefixes (also AX ONE etc.)
    s = re.sub(r"^(HG|AX)\b\s*", "", s, flags=re.I).strip()
    if series:
        s = re.sub(re.escape(series), "", s, flags=re.I).strip()
    if finish:
        # Remove the finish word itself (chrome/matt black/etc.)
        s = re.sub(rf"\b{re.escape(finish)}\b", "", s, flags=re.I).strip()
    # Trim trailing punctuation
    s = re.sub(r"[.\-·\s,]+$", "", s).strip()
    return s or "family"


def _detect_section(row0: str) -> str | None:
    """Return the section header text for a row that has a value in col A only."""
    if not row0:
        return None
    r = str(row0).strip()
    if not r or len(r) < 4:
        return None
    # A section header must contain letters and be short-ish
    letters = sum(1 for c in r if c.isalpha())
    if letters < 3 or len(r) > 80:
        return None
    return r


def _subcategory_from_section(section: str | None, default: str | None) -> str | None:
    if not section:
        return default
    low = section.lower()
    for kw, sub in SECTION_TO_SUBCAT:
        if kw in low:
            return sub or default
    return default


def _parse_mrp(v) -> float | str:
    if v in (None, ""):
        return MISSING
    try:
        s = str(v).replace("₹", "").replace(" ", "").strip()
        # Handle Indian format like "21,29,710" (7 digits with commas)
        s = s.replace(",", "")
        f = float(s)
        return f if f > 0 else MISSING
    except (TypeError, ValueError):
        return MISSING


class HansgroheAdapter(BrandAdapter):
    brand = "Hansgrohe"  # AXOR merged under Hansgrohe
    supported_extensions = (".xlsx", ".xls")

    def extract(self, data: bytes, filename: str) -> tuple[list[ProductRow], ExtractionReport]:
        report = ExtractionReport(brand=self.brand, filename=filename, source_type="excel")
        rows: list[ProductRow] = []
        try:
            from openpyxl import load_workbook  # type: ignore
        except Exception as e:  # pragma: no cover
            report.warnings.append(f"openpyxl missing: {e}")
            return rows, report
        try:
            wb = load_workbook(io.BytesIO(data), data_only=True)
        except Exception as e:
            report.warnings.append(f"xlsx open failed: {e}")
            return rows, report

        default_cat, default_subcat = _category_from_filename(filename)

        # ---- Image mapping: (sheet, row) → highest-quality anchor ----
        by_sheet_row: dict[tuple[str, int], ExtractedImage] = {}
        for sheet, row_idx, img in extract_images_from_xlsx_ex(data):
            key = (sheet, row_idx)
            prev = by_sheet_row.get(key)
            qrank = {"excellent": 4, "good": 3, "acceptable": 2, "poor": 1}
            if prev is None or (qrank.get(img.quality, 0), img.longest_edge) > (qrank.get(prev.quality, 0), prev.longest_edge):
                by_sheet_row[key] = img
        report.images_found = len(by_sheet_row)

        anchors_by_sheet: dict[str, list[int]] = {}
        for (sheet, r) in by_sheet_row:
            anchors_by_sheet.setdefault(sheet, []).append(r)
        for s in anchors_by_sheet:
            anchors_by_sheet[s].sort()

        for ws in wb.worksheets:
            all_rows: list[list] = [list(r) for r in ws.iter_rows(values_only=True)]
            if not all_rows:
                continue

            # Find header row (first row containing "Article" in its cells)
            hdr_idx = 0
            for i, r in enumerate(all_rows[:5]):
                if any((str(c or "").lower().replace("\n", " ").strip() in ("article no.", "article no", "article number")) for c in r):
                    hdr_idx = i
                    break
            header_row = all_rows[hdr_idx]

            # Locate columns
            col_map = {}
            for c, val in enumerate(header_row):
                key = str(val or "").lower().replace("\n", " ").strip()
                if not key:
                    continue
                if "article" in key:                col_map["sku"] = c
                elif "discription" in key or "description" in key or "product" == key.strip():
                    if "sku" not in col_map or c != col_map.get("sku"):
                        # First occurrence of a description-like column
                        col_map.setdefault("name", c)
                elif key == "mrp":                  col_map["mrp"] = c
                elif key == "image" or "image" in key: col_map.setdefault("image", c)
                elif key in ("qty", "qty.", "oty", "oty.", "unit"): col_map["qty"] = c

            if "sku" not in col_map or "name" not in col_map:
                report.warnings.append(f"Sheet {ws.title!r} in {filename}: could not locate SKU/Name columns")
                continue

            current_subcat = default_subcat
            current_section: str | None = None

            for r_idx, row in enumerate(all_rows[hdr_idx + 1:], start=hdr_idx + 2):
                # Detect a "section header" row (only col A has text, SKU blank)
                sku_cell = row[col_map["sku"]] if col_map["sku"] < len(row) else None
                if (sku_cell in (None, "")):
                    section = _detect_section(row[0] if row else None)
                    if section:
                        current_section = section
                        current_subcat = _subcategory_from_section(section, default_subcat)
                    continue

                sku_raw = sku_cell
                sku = str(sku_raw).strip()
                if not sku or not re.match(r"^[A-Za-z0-9\-\.\/]+$", sku):
                    continue
                # Hansgrohe SKUs are numeric; strip leading zeros only if pure numeric
                if sku.isdigit():
                    sku = str(int(sku))
                    # Restore expected 8-digit padding when appropriate (leading zeros are informational)
                    # Actually the source file shows 7-8 digit numeric — keep as-is post int() conversion.

                name_cell = row[col_map["name"]] if col_map["name"] < len(row) else None
                name = str(name_cell or "").strip()
                if not name:
                    continue

                mrp_raw = row[col_map["mrp"]] if "mrp" in col_map and col_map["mrp"] < len(row) else None
                mrp = _parse_mrp(mrp_raw)

                collection = _detect_collection(name)
                series = _detect_series(name)
                finish_label, finish_code_from_name = _detect_finish(name, sku)
                if not finish_label:
                    finish_label, finish_code_from_name = _detect_finish(name, sku)
                base = _base_name(name, series, finish_label)

                # Prefer name-derived finish code, else last 3 digits of SKU
                if not finish_code_from_name:
                    if sku.isdigit() and len(sku) >= 3:
                        finish_code_from_name = sku[-3:]
                    else:
                        finish_code_from_name = None

                # Dimensions: look for a size token in the name (mm)
                dim = None
                m = re.search(r"\b(\d{2,4})(?:\s*[×xX]\s*(\d{2,4}))?\s*(?:mm|cm)?\b", name)
                if m and m.group(2):
                    dim = f"{m.group(1)}x{m.group(2)}mm"

                # Family key: collection (Hansgrohe/AXOR) + series + base
                series_slug = _slug(series or "misc")
                base_slug = _slug(base)
                family_key = f"hansgrohe:{_slug(collection)}:{series_slug}:{base_slug}"

                # Image: nearest anchor within ±3 rows
                best: ExtractedImage | None = None
                anchors = anchors_by_sheet.get(ws.title, [])
                best_distance = 99
                for a in anchors:
                    d = abs(a - r_idx)
                    if d <= 3 and d < best_distance:
                        candidate = by_sheet_row.get((ws.title, a))
                        if candidate is not None:
                            best = candidate
                            best_distance = d
                            if d == 0:
                                break
                image_urls = [best.data_url] if best else []
                image_meta = [best.to_dict()] if best else []
                image_quality = best.quality if best else "missing"
                if best:
                    report.images_mapped += 1

                # Colour = same as finish label for Hansgrohe (finish IS the colour)
                colour = finish_label
                subcategory = current_subcat if current_subcat else default_subcat

                f"{series} · {base}".strip(" ·") if series else base

                specs: dict = {}
                if current_section:
                    specs["section"] = current_section
                if dim:
                    specs["size"] = dim
                # Store collection as a spec + as top-level field
                specs["collection"] = collection

                pr = ProductRow(
                    brand=self.brand,
                    sku=sku,
                    name=name,
                    category=default_cat,
                    subcategory=subcategory or MISSING,
                    series=series or MISSING,
                    family_key=family_key,
                    variant=finish_label or MISSING,
                    finish=finish_label or MISSING,
                    finish_code=finish_code_from_name or MISSING,
                    colour=colour or MISSING,
                    material=MISSING,
                    dimensions=dim or MISSING,
                    description=name,
                    mrp=mrp,
                    dealer_price=MISSING,
                    warranty=MISSING,
                    collection=collection,
                    images=image_urls,
                    image_meta=image_meta,
                    image_quality=image_quality,
                    image_page=None,
                    specs=specs,
                    tags=dedupe_iter([
                        default_cat.lower(),
                        self.brand.lower(),
                        collection.lower(),
                        (series or "").lower(),
                        (finish_label or "").lower(),
                    ]),
                    confidence=0.94 if (mrp != MISSING and series and finish_label) else (0.86 if mrp != MISSING else 0.6),
                )
                if mrp == MISSING:
                    pr.issues.append("Missing MRP")
                if not series:
                    pr.issues.append("Series could not be auto-detected — verify manually")
                if not best:
                    pr.issues.append("No image mapped from supplier file")
                elif best.quality == "poor":
                    pr.issues.append(f"Only thumbnail-grade image available ({best.longest_edge}px longest edge)")
                rows.append(pr)

        report.parsed_rows = len(rows)
        return rows, report
