"""QUTONE XLSX adapter (Ground Floor > Tiles).

Ships as a single workbook, single sheet, one row per (series, product name,
size, finish) combination:

  Row 1  : column headers (SR., company NAME, PRODUCT NAME, IMAGE,
           PRODUCT SIZE, SERIES NAME, FINISHES, BOX IN PIS, BOX SQFT, RATE)
  Row 2+ : one row per size/finish variant. The embedded product photo
           floats over the "IMAGE" cell of that exact row — verified 1:1
           row anchoring (zero rows with more than one image), so
           row+column anchor matching is safe here.

Rules:
* Brand is always "Qutone"; category is always "Tiles" (this brand only
  ever ships tile pricelists — a brand adapter encoding its own known
  category is the same pattern every existing adapter already uses, e.g.
  Oyster's `COLLECTION_NAME = "Brook"`).
* No SKU/article code exists anywhere in the source. Verified unique across
  all 452 real rows: (SERIES NAME, PRODUCT NAME, PRODUCT SIZE, FINISHES).
  SKU is therefore synthesized, deterministic:
  `QUTONE-{SERIES_COMPACT}-{NAME_COMPACT}-{SIZE_COMPACT}-{FINISH_CODE}` —
  same inputs always regenerate the same SKU (required for idempotency). A
  handful of rows in a future refreshed file could repeat a listing — the
  first occurrence keeps the plain SKU and each repeat gets a numeric
  suffix (`-2`, `-3`, ...), matching the Oyster adapter's convention, so it
  still imports as its own product instead of being dropped.
* "FINISHES" is normalized via an explicit FINISH_LOOKUP table (not fuzzy
  regex) built from the 6 distinct raw strings actually observed in the
  real file. Anything new is flagged for manual review, never guessed.
* Family key groups sibling size/finish variants of the same product within
  the same series: `qutone:{series_slug}:{name_slug}`.
* "RATE" is the only price the source provides (format "<number> PER
  SQFT") — used for both `mrp` and `dealer_price`/`price`, since there is
  no separate MRP tier in this source (inventing one would be
  fabrication). A row whose RATE doesn't match the expected format is
  flagged `needs_pricing` and imported at ₹0 rather than held back.
* Column positions are discovered from the header row text, not hardcoded
  letters/indices — tolerant of the source file being re-exported with
  columns in a different order on a future refresh.
"""
from __future__ import annotations
import io
import re

from ..base import MISSING, BrandAdapter, ExtractionReport, ProductRow, dedupe_iter
from ..image_extractor import ExtractedImage, extract_images_from_xlsx_ex

BRAND = "Qutone"
CATEGORY = "Tiles"

# Explicit lookup, not regex guessing — built from the 6 distinct raw
# "FINISHES" cell values actually present in the real source file
# (verified by direct inspection before writing this table).
FINISH_LOOKUP: dict[str, tuple[str, str]] = {
    "MATT": ("Matt", "MT"),
    "GLOSSY": ("Glossy", "GL"),
    "CHIFFON": ("Chiffon", "CH"),
    "DOVE": ("Dove", "DV"),
    "SILK": ("Silk", "SK"),
    "STRUCTURE-MATT": ("Structure Matt", "SM"),
}

_RATE_RE = re.compile(r"^\s*([\d,]+(?:\.\d+)?)\s*PER\s*SQFT\s*$", re.I)


def normalize_finish(raw: str) -> tuple[str | None, str | None, str | None]:
    """Returns (finish_label, finish_code, note). `note` is set only when the
    value is unrecognized (needs manual review) — callers should surface it
    rather than silently swallowing it."""
    s = re.sub(r"\s+", " ", str(raw or "").replace("\xa0", " ")).strip().upper()
    hit = FINISH_LOOKUP.get(s)
    if hit:
        return hit[0], hit[1], None
    return None, None, f"unrecognized finish {raw!r} — needs manual review"


def parse_rate_per_sqft(raw) -> tuple[float | None, str | None]:
    """Returns (price_per_sqft, note). `note` is set when the source RATE
    cell doesn't match the expected "<number> PER SQFT" format — the row
    still imports (at ₹0, flagged needs_pricing), never dropped."""
    s = str(raw or "").strip()
    m = _RATE_RE.match(s)
    if not m:
        return None, f"unrecognized RATE format {raw!r} — expected '<number> PER SQFT'"
    try:
        return float(m.group(1).replace(",", "")), None
    except ValueError:
        return None, f"unrecognized RATE format {raw!r} — expected '<number> PER SQFT'"


def _slug(s: str) -> str:
    return re.sub(r"[^a-z0-9]+", "-", (s or "").lower()).strip("-")


def _compact(s: str) -> str:
    return re.sub(r"[^A-Z0-9]", "", (s or "").upper())


def family_key_for(series: str, name: str) -> str:
    return f"qutone:{_slug(series)}:{_slug(name)}"


def sku_for(series: str, name: str, size: str, finish_code: str) -> str:
    series_c = _compact(series)[:24] or "SERIES"
    name_c = _compact(name)[:32] or "PRODUCT"
    size_c = _compact(size)[:16] or "SIZE"
    return f"QUTONE-{series_c}-{name_c}-{size_c}-{finish_code}"


def _fmt_pcs(v) -> str | None:
    if v == MISSING or v is None:
        return None
    try:
        f = float(v)
        return str(int(f)) if f.is_integer() else str(f)
    except (TypeError, ValueError):
        return str(v)


def _find_header_row(all_rows: list[list]) -> int:
    for i, r in enumerate(all_rows[:5]):
        cells = [str(c or "").lower().strip() for c in r]
        if any("product name" in c for c in cells):
            return i
    return 0


def _build_col_map(header_row: list) -> dict[str, int]:
    col_map: dict[str, int] = {}
    for c, val in enumerate(header_row):
        key = str(val or "").lower().strip()
        if not key:
            continue
        if "company" in key:
            col_map["company"] = c
        elif "product name" in key:
            col_map["name"] = c
        elif key == "image" or "image" in key:
            col_map["image"] = c
        elif "size" in key:
            col_map["size"] = c
        elif "series" in key:
            col_map["series"] = c
        elif "finish" in key:
            col_map["finish"] = c
        elif "box" in key and ("pc" in key or "pis" in key):
            col_map["pcs_per_box"] = c
        elif "box" in key and "sqft" in key:
            col_map["sqft_per_box"] = c
        elif "rate" in key:
            col_map["rate"] = c
    return col_map


class QutoneAdapter(BrandAdapter):
    brand = BRAND
    supported_extensions = (".xlsx", ".xls")

    def extract(self, data: bytes, filename: str) -> tuple[list[ProductRow], ExtractionReport]:
        report = ExtractionReport(brand=self.brand, filename=filename, source_type="excel")
        rows: list[ProductRow] = []
        try:
            from openpyxl import load_workbook
        except Exception as e:  # pragma: no cover
            report.warnings.append(f"openpyxl missing: {e}")
            return rows, report

        try:
            wb = load_workbook(io.BytesIO(data), data_only=True)
        except Exception as e:
            report.warnings.append(f"xlsx open failed: {e}")
            return rows, report

        sku_counts: dict[str, int] = {}
        qrank = {"excellent": 4, "good": 3, "acceptable": 2, "poor": 1}

        for ws in wb.worksheets:
            all_rows = [list(r) for r in ws.iter_rows(values_only=True)]
            if not all_rows:
                continue
            hdr_idx = _find_header_row(all_rows)
            col_map = _build_col_map(all_rows[hdr_idx])
            if not {"name", "size", "finish", "rate"} <= col_map.keys():
                report.warnings.append(
                    f"Sheet {ws.title!r} in {filename}: could not locate name/size/finish/rate columns"
                )
                continue

            # Images extracted with optimize=False — this brand's explicit
            # requirement is original quality, no resize/recompress.
            image_col = col_map.get("image")
            by_row: dict[int, ExtractedImage] = {}
            for sheet_name, row_idx, col_idx, img in extract_images_from_xlsx_ex(data, optimize=False):
                if sheet_name != ws.title:
                    continue
                if image_col is not None and col_idx != image_col:
                    continue
                prev = by_row.get(row_idx)
                if prev is None or (qrank.get(img.quality, 0), img.longest_edge) > (qrank.get(prev.quality, 0), prev.longest_edge):
                    by_row[row_idx] = img
            report.images_found += len(by_row)

            for r_idx, row in enumerate(all_rows[hdr_idx + 1:], start=hdr_idx + 2):
                def cell(key, _row=row):
                    idx = col_map.get(key)
                    return _row[idx] if idx is not None and idx < len(_row) else None

                name_cell, size_cell, finish_cell = cell("name"), cell("size"), cell("finish")
                if not name_cell or not size_cell or not finish_cell:
                    continue  # blank / section / trailer row

                company = str(cell("company") or "").strip()
                name = str(name_cell).strip()
                size = str(size_cell).strip()
                series = str(cell("series") or "").strip()
                finish_label, finish_code, finish_note = normalize_finish(finish_cell)
                display_finish = finish_label or str(finish_cell).strip()

                pcs_per_box = self.to_number(cell("pcs_per_box"))
                sqft_per_box = self.to_number(cell("sqft_per_box"))

                rate, rate_note = parse_rate_per_sqft(cell("rate"))
                needs_pricing = rate is None
                price = rate if rate is not None else 0.0

                family_key = family_key_for(series, name)
                base_sku = sku_for(series, name, size, finish_code or "UNK")
                occurrence = sku_counts.get(base_sku, 0) + 1
                sku_counts[base_sku] = occurrence
                sku = base_sku if occurrence == 1 else f"{base_sku}-{occurrence}"

                img = by_row.get(r_idx)
                image_urls = [img.data_url] if img else []
                image_meta = [img.to_dict()] if img else []
                image_quality = img.quality if img else "missing"
                if img:
                    report.images_mapped += 1

                pr = ProductRow(
                    brand=self.brand,
                    sku=sku,
                    name=f"{name} - {display_finish} ({size})",
                    category=CATEGORY,
                    subcategory=MISSING,
                    series=series or MISSING,
                    family_key=family_key,
                    variant=f"{size} · {display_finish}",
                    finish=finish_label or display_finish,
                    finish_code=finish_code or MISSING,
                    colour=MISSING,
                    material=MISSING,
                    dimensions=MISSING,
                    size=size,
                    description=MISSING,
                    mrp=price,
                    dealer_price=price,
                    warranty=MISSING,
                    collection=MISSING,
                    images=image_urls,
                    image_meta=image_meta,
                    image_quality=image_quality,
                    image_page=None,
                    specs={
                        "company_name": company or None,
                        "pcs_per_box": _fmt_pcs(pcs_per_box),
                        "sqft_per_box": None if sqft_per_box == MISSING else sqft_per_box,
                        "source_file": filename,
                        **({"needs_pricing": True} if needs_pricing else {}),
                        **({"duplicate_listing": True} if occurrence > 1 else {}),
                    },
                    tags=dedupe_iter([
                        CATEGORY.lower(), self.brand.lower(), (series or "").lower(), (finish_label or "").lower(),
                        *(["needs-pricing"] if needs_pricing else []),
                    ]),
                    # Confidence reflects data-quality signals only (finish
                    # recognized + price parsed) — NOT image presence, so
                    # rows with no supplier photo still auto-import instead
                    # of being held back for manual review.
                    confidence=0.95 if (finish_label and not needs_pricing) else 0.6,
                )
                if not finish_label:
                    pr.issues.append(finish_note or f"Unrecognized finish {finish_cell!r} — needs manual review")
                if needs_pricing:
                    pr.issues.append(rate_note or "Missing/unrecognized RATE in source — imported at ₹0, needs manual pricing")
                if not img:
                    pr.issues.append("No image mapped from supplier file")
                if occurrence > 1:
                    pr.issues.append(f"Duplicate listing in source file (occurrence {occurrence}) - SKU suffixed to keep as a separate product")
                rows.append(pr)

        report.parsed_rows = len(rows)
        return rows, report
