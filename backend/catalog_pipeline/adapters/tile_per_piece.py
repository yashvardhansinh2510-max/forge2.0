"""Adapters for simple ground-floor tile price lists priced per piece.

Milagro and Kenzo each use ``NAME / IMAGE / SIZE / RATE PER PCS`` spreadsheets.
The embedded supplier art is reframed onto a 4:3 landscape canvas before it
reaches Supabase.  This is deliberate: quotation cards are landscape and a
portrait source asset must not become a clipped vertical image there.  The
original artwork is never cropped or rotated; a neutral background provides
the extra horizontal space.
"""
from __future__ import annotations

import base64
import io
import re

from PIL import Image, ImageFilter

from ..base import MISSING, BrandAdapter, ExtractionReport, ProductRow, dedupe_iter
from ..image_extractor import ExtractedImage, classify_quality, extract_images_from_xlsx_ex

CATEGORY = "Tiles"
_RATE_RE = re.compile(r"^\s*([\d,]+(?:\.\d+)?)\s*PER\s*PCS\s*$", re.I)
_RATE_ANY_RE = re.compile(r"([\d,]+(?:\.\d+)?)\s*PER\s*PCS", re.I)


def _slug(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "-", value.lower()).strip("-")


def _compact(value: str) -> str:
    return re.sub(r"[^A-Z0-9]+", "", value.upper())


def _parse_rate(value: object) -> float | None:
    match = _RATE_RE.match(str(value or ""))
    if match:
        return float(match.group(1).replace(",", ""))
    # Two Milagro sheets rows specify PLAN and HL prices in a single cell.
    # The first (PLAN) rate is the only defensible default for a one-SKU row;
    # retain the full source cell below so the alternate HL price is visible.
    alternate = _RATE_ANY_RE.search(str(value or ""))
    return float(alternate.group(1).replace(",", "")) if alternate else None


def _landscape_image(img: ExtractedImage) -> ExtractedImage:
    """Return a clear, full-bleed-free landscape rendition for quotations."""
    try:
        _, encoded = img.data_url.split(",", 1)
        source = Image.open(io.BytesIO(base64.b64decode(encoded))).convert("RGBA")
        # A 1024×768 final asset comfortably supports the quotation UI.  The
        # 6% margin ensures no supplier artwork is clipped at the canvas edge.
        canvas_w, canvas_h, margin = 1024, 768, 46
        scale = min((canvas_w - 2 * margin) / source.width, (canvas_h - 2 * margin) / source.height)
        rendered = source.resize(
            (max(1, round(source.width * scale)), max(1, round(source.height * scale))),
            Image.Resampling.LANCZOS,
        )
        # A restrained sharpening pass offsets the soft scaling in the small
        # embedded workbook previews without inventing or removing detail.
        rendered = rendered.filter(ImageFilter.UnsharpMask(radius=1.1, percent=115, threshold=3))
        canvas = Image.new("RGBA", (canvas_w, canvas_h), (250, 250, 250, 255))
        x, y = (canvas_w - rendered.width) // 2, (canvas_h - rendered.height) // 2
        canvas.alpha_composite(rendered, (x, y))
        output = io.BytesIO()
        canvas.convert("RGB").save(output, format="JPEG", quality=92, optimize=True, progressive=True)
        data = output.getvalue()
        sha1 = __import__("hashlib").sha1(data).hexdigest()[:16]
        return ExtractedImage(
            data_url="data:image/jpeg;base64," + base64.b64encode(data).decode("ascii"),
            sha1=sha1, mime="image/jpeg", width=canvas_w, height=canvas_h,
            quality=classify_quality(max(canvas_w, canvas_h), "jpeg"),
            source_format="jpeg", bytes_len=len(data),
        )
    except Exception:
        # Preserve a source image if an unusual but still supported image type
        # cannot be rendered; the import remains lossless rather than failing.
        return img


class PerPieceTileAdapter(BrandAdapter):
    brand = ""

    def extract(self, data: bytes, filename: str) -> tuple[list[ProductRow], ExtractionReport]:
        report = ExtractionReport(brand=self.brand, filename=filename, source_type="excel")
        try:
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(data), data_only=True)
        except Exception as exc:
            report.warnings.append(f"xlsx open failed: {exc}")
            return [], report

        rows: list[ProductRow] = []
        sku_counts: dict[str, int] = {}
        quality_rank = {"excellent": 4, "good": 3, "acceptable": 2, "poor": 1}
        for ws in wb.worksheets:
            values = [list(row) for row in ws.iter_rows(values_only=True)]
            if not values:
                continue
            headers = {str(v or "").strip().lower(): i for i, v in enumerate(values[0])}
            name_col = headers.get("name")
            size_col = headers.get("size")
            rate_col = next((i for key, i in headers.items() if "rate" in key), None)
            image_col = headers.get("image")
            if None in (name_col, size_col, rate_col):
                report.warnings.append(f"Sheet {ws.title!r}: expected NAME, SIZE and RATE PER PCS columns")
                continue

            images_by_row: dict[int, ExtractedImage] = {}
            for sheet, row, col, image in extract_images_from_xlsx_ex(data, optimize=False):
                if sheet != ws.title or (image_col is not None and col != image_col):
                    continue
                previous = images_by_row.get(row)
                if previous is None or quality_rank.get(image.quality, 0) > quality_rank.get(previous.quality, 0):
                    images_by_row[row] = image
            report.images_found += len(images_by_row)

            for row_no, row in enumerate(values[1:], start=2):
                name = str(row[name_col] or "").strip()
                size = str(row[size_col] or "").strip()
                if not name or not size:
                    continue
                price = _parse_rate(row[rate_col])
                base_sku = f"{_compact(self.brand)}-{_compact(name)[:36] or 'PRODUCT'}-{_compact(size)[:16] or 'SIZE'}"
                occurrence = sku_counts.get(base_sku, 0) + 1
                sku_counts[base_sku] = occurrence
                sku = base_sku if occurrence == 1 else f"{base_sku}-{occurrence}"
                source_image = images_by_row.get(row_no)
                image = _landscape_image(source_image) if source_image else None
                if image:
                    report.images_mapped += 1
                product = ProductRow(
                    brand=self.brand, sku=sku, name=f"{name} ({size})", category=CATEGORY,
                    subcategory=MISSING, series=MISSING, family_key=f"{_slug(self.brand)}:{_slug(name)}",
                    variant=size, finish=MISSING, finish_code=MISSING, colour=MISSING,
                    material=MISSING, dimensions=MISSING, size=size, description=MISSING,
                    mrp=price if price is not None else 0.0, dealer_price=price if price is not None else 0.0,
                    warranty=MISSING, collection=MISSING,
                    images=[image.data_url] if image else [], image_meta=[image.to_dict()] if image else [],
                    image_quality=image.quality if image else "missing", image_page=None,
                    specs={"source_file": filename, "price_unit": "per piece", "source_rate": str(row[rate_col] or "").strip(), "image_layout": "landscape_4_3"},
                    tags=dedupe_iter(["tiles", self.brand.lower(), "per-piece"]),
                    confidence=0.95 if price is not None else 0.6,
                )
                if price is None:
                    product.issues.append(f"Unrecognized RATE PER PCS value {row[rate_col]!r}; imported at ₹0 for review")
                if not image:
                    product.issues.append("No image mapped from supplier file")
                if occurrence > 1:
                    product.issues.append(f"Duplicate listing occurrence {occurrence}; SKU suffixed to retain it")
                rows.append(product)
        report.parsed_rows = len(rows)
        return rows, report


class MilagroAdapter(PerPieceTileAdapter):
    brand = "Milagro"


class KenzoAdapter(PerPieceTileAdapter):
    brand = "Kenzo"
