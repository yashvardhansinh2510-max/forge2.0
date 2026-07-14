"""VITRA XLSX adapter.

The 2026 Price Table stores each product family on ONE row, then repeats
Code / IMAGE / MRP for every finish (WHITE 003/403, MATT WHITE 401, MATT TAUPE 420,
MATT STONE GREY 476, MATT BLACK 483). We fan those out into one ProductRow per
non-empty finish variant.

Row layout observed in the real 2026 file:
  Row 1 : finish headers (spanning across 4 columns each)
  Row 2 : per-finish sub-headers: Code, IMAGE, MRP, (blank)
  Row 3+: product rows. Col A = Family (Design). Col B = Detail. Then 4-col groups.

Iteration 1 changes:
* Emits `image_meta` (per-image {width,height,quality,source_format}) and an
  aggregate `image_quality` bucket per row.
* Produces `subcategory` from the sheet name (kept as-is, title-cased) so the
  catalog can drill down Brand → Category → Subcategory → Series → Family →
  Variants without extra parsing later.
* Series is the Design column (e.g. "Metropole", "Sento", "Zentrum").
* Family key is stable across finishes: {sheet_slug}:{design_slug}:{detail_slug}.
* Finish code is the numeric suffix in the finish header (003, 401, 420, 476, 483).
"""
from __future__ import annotations
import io
import re
from collections import defaultdict

from ..base import MISSING, BrandAdapter, ExtractionReport, ProductRow, dedupe_iter
from ..image_extractor import ExtractedImage, extract_images_from_xlsx_ex

SKU_RE = re.compile(r"[A-Z]?\d{3,6}B[A-Z0-9]+H\d{2,6}", re.IGNORECASE)
FINISH_CODE_RE = re.compile(r"\b(\d{3})\b")

SHEET_TO_CATEGORY = {
    "csw": "Water Closets",     # Ceramic Sanitary Ware
    "wc": "Water Closets", "toilet": "Water Closets", "closet": "Water Closets",
    "urinal": "Urinals", "bidet": "Bidets",
    "basin": "Basins", "washbasin": "Basins", "counter": "Basins", "vanity": "Basins",
    "mixer": "Faucets", "faucet": "Faucets", "tap": "Faucets",
    "shower": "Showers",
    "cistern": "Concealed Cisterns", "flush": "Flush Plates",
    "bathtub": "Bathtubs", "tub": "Bathtubs",
    "accessor": "Accessories",
}

# Recognised subcategory keywords found in the "Detail" column
SUBCATEGORY_KEYWORDS = [
    "Wall Hung WC", "Back to Wall WC", "Rimless WC", "Rim-Ex WC",
    "Wall Hung Bidet", "Console Basin", "Vanity Basin", "Countertop Basin",
    "Under Counter Basin", "Wall Hung Basin", "Semi-Recessed Basin",
    "Concealed Cistern", "Exposed Cistern", "Actuator Plate",
    "Wall Hung Urinal", "Floor Urinal",
    "Single Lever Basin Mixer", "Basin Mixer", "Wall Mixer",
    "Bath Mixer", "Shower Mixer", "Rain Shower", "Hand Shower",
    "Freestanding Bathtub", "Built-in Bathtub", "Whirlpool",
]


def _classify(sheet_name: str, detail: str) -> str:
    """Determine category. Prefer the more-specific `detail` field over the
    generic sheet name (Vitra ships every ceramic product on one 'CSW' sheet)."""
    detail_low = (detail or "").lower()
    # Order: most-specific first (bidet, urinal, shower, bathtub, cistern, mixer, basin, WC, accessory)
    detail_priority = [
        ("bidet",   "Bidets"),
        ("urinal",  "Urinals"),
        ("shower",  "Showers"),
        ("bathtub", "Bathtubs"),
        ("tub",     "Bathtubs"),
        ("cistern", "Concealed Cisterns"),
        ("flush",   "Flush Plates"),
        ("mixer",   "Faucets"),
        ("faucet",  "Faucets"),
        ("tap",     "Faucets"),
        ("basin",   "Basins"),
        ("counter", "Basins"),
        ("vanity",  "Basins"),
        ("washbasin", "Basins"),
        ("wc",      "Water Closets"),
        ("toilet",  "Water Closets"),
        ("closet",  "Water Closets"),
        ("accessor","Accessories"),
    ]
    for kw, cat in detail_priority:
        if kw in detail_low:
            return cat
    # Fallback: sheet-name heuristics (retains previous coarse mapping)
    sheet_low = (sheet_name or "").lower()
    for k, v in SHEET_TO_CATEGORY.items():
        if k in sheet_low:
            return v
    return MISSING


def _extract_subcategory(sheet_name: str, detail: str, category: str) -> str:
    """Best-effort subcategory.

    Returns a stable, meaningful label — not free-text like '600Mm & 400Mm'.
    Priority:
      1. Keyword match from the controlled SUBCATEGORY_KEYWORDS list.
      2. A cleaner form of `detail` if it looks like a product-form label
         (all letters, ≤ 30 chars, ≥ 3 chars).
      3. The category itself (so hierarchy stays coherent).
    """
    hay = f"{sheet_name} {detail}"
    for kw in SUBCATEGORY_KEYWORDS:
        if kw.lower() in hay.lower():
            return kw
    d = (detail or "").strip()
    # Reject noisy details containing numbers or dimension markers
    letters = sum(1 for c in d if c.isalpha())
    if 3 <= len(d) <= 30 and letters >= 3 and not any(c.isdigit() for c in d):
        return d.title()
    if category and category != MISSING:
        return category
    return MISSING


def _finish_code_from_header(finish_header: str) -> str:
    """Vitra headers like 'MATT BLACK 483' embed the 3-digit finish code."""
    m = FINISH_CODE_RE.search(finish_header or "")
    return m.group(1) if m else MISSING


def _slug(s: str) -> str:
    return re.sub(r"[^a-z0-9]+", "-", (s or "").lower()).strip("-")


class VitraAdapter(BrandAdapter):
    brand = "Vitra"
    supported_extensions = (".xlsx", ".xls")

    def extract(self, data: bytes, filename: str) -> tuple[list[ProductRow], ExtractionReport]:
        report = ExtractionReport(brand=self.brand, filename=filename, source_type="excel")
        rows: list[ProductRow] = []
        try:
            from openpyxl import load_workbook  # type: ignore
        except Exception as e:  # pragma: no cover
            report.warnings.append(f"openpyxl missing: {e}")
            return rows, report
        try:
            wb = load_workbook(io.BytesIO(data), data_only=True)
        except Exception as e:
            report.warnings.append(f"xlsx open failed: {e}")
            return rows, report

        last_design = MISSING

        # Extract every image with quality metadata, keyed by its EXACT
        # (sheet, row, column) anchor position — never collapsed to
        # (sheet, row) alone. The Vitra sheet places one finish's photo per
        # finish-group, side-by-side in the SAME row (WHITE / MATT WHITE /
        # MATT TAUPE / ... each own their own "IMAGE" sub-column). Keying on
        # row alone would keep only the single highest-quality blip per row
        # and silently hand every other finish that same picture — this is
        # the exact bug that made every Vitra finish variant show an
        # identical photo. Keeping the column means each finish keeps its
        # own picture.
        by_sheet_row_col: dict[tuple[str, int, int], ExtractedImage] = {}
        qrank_pre = {"excellent": 4, "good": 3, "acceptable": 2, "poor": 1}
        for sheet, row_idx, col_idx, img in extract_images_from_xlsx_ex(data):
            key = (sheet, row_idx, col_idx)
            prev = by_sheet_row_col.get(key)
            if prev is None or (qrank_pre.get(img.quality, 0), img.longest_edge) > (qrank_pre.get(prev.quality, 0), prev.longest_edge):
                by_sheet_row_col[key] = img
        report.images_found = len(by_sheet_row_col)

        # Group image anchor (row, col) pairs by sheet for closest-match
        # lookup, restricted per-column below so different finishes never
        # borrow each other's anchor.
        anchors_by_sheet: dict[str, list[tuple[int, int]]] = {}
        for (sheet, row_idx, col_idx) in by_sheet_row_col:
            anchors_by_sheet.setdefault(sheet, []).append((row_idx, col_idx))
        for s in anchors_by_sheet:
            anchors_by_sheet[s].sort()

        for ws in wb.worksheets:
            all_rows: list[list] = [list(r) for r in ws.iter_rows(values_only=True)]
            if len(all_rows) < 3:
                continue

            finish_header = all_rows[0]
            sub_header = all_rows[1]

            raw_groups: list[tuple[str, int]] = []
            for col, val in enumerate(finish_header):
                if val and str(val).strip():
                    raw_groups.append((str(val).strip(), col))
            group_columns: list[tuple[str, int, int, int | None]] = []
            for gi, (finish, start_col) in enumerate(raw_groups):
                next_col = raw_groups[gi + 1][1] if gi + 1 < len(raw_groups) else len(sub_header)
                code_col: int | None = None
                mrp_col: int | None = None
                image_col: int | None = None
                for c in range(start_col, min(next_col, len(sub_header))):
                    label = str(sub_header[c] or "").strip().lower()
                    if label == "code" and code_col is None:
                        code_col = c
                    elif label == "mrp" and mrp_col is None:
                        mrp_col = c
                    elif "image" in label and image_col is None:
                        image_col = c
                if code_col is not None and mrp_col is not None:
                    if image_col is None:
                        # Sub-header didn't literally spell "IMAGE" — infer it
                        # from the documented layout (Code, IMAGE, MRP) as the
                        # column immediately after Code, but only when that
                        # column actually falls before MRP (never guess past
                        # the group's own boundary into the next finish).
                        candidate = code_col + 1
                        image_col = candidate if candidate < mrp_col else None
                    group_columns.append((finish, code_col, mrp_col, image_col))

            if not group_columns:
                report.warnings.append(f"Sheet '{ws.title}' has no recognizable finish groups; fell back to narrow parser")

            for r_idx, row in enumerate(all_rows[2:], start=3):
                design_raw = str(row[0] or "").strip()
                detail_raw = str(row[1] or "").strip() if len(row) > 1 else ""
                # The supplier's own sheet repeats the header block itself
                # partway down (verified in the 2026 file at rows 45 & 69) —
                # without this guard those rows get parsed as bogus products
                # whose "SKU" is literally the finish-header text (e.g.
                # "WHITE 003/403"). Never treat a repeated header as data.
                if design_raw.upper() == "DESIGN" and detail_raw.upper() == "DETAIL":
                    continue
                # Series column can contain embedded newlines ("OPTIONS \nWITH TAP HOLE").
                # Collapse them so the hierarchy label stays clean.
                design = re.sub(r"\s+", " ", design_raw) if design_raw else last_design
                detail = re.sub(r"\s+", " ", detail_raw) if detail_raw else ""
                if design:
                    last_design = design

                if group_columns:
                    for finish, code_col, mrp_col, image_col in group_columns:
                        code_cell = str(row[code_col]).strip() if code_col < len(row) and row[code_col] not in (None, "") else ""
                        if not code_cell:
                            continue
                        primary_sku = code_cell.splitlines()[0].strip()
                        extras = [s.strip() for s in code_cell.splitlines()[1:] if s.strip()]
                        mrp_raw = row[mrp_col] if mrp_col < len(row) else None
                        mrp = self.to_number(mrp_raw)
                        cat = _classify(ws.title, detail)
                        subcat = _extract_subcategory(ws.title, detail, cat)

                        # Family key is stable across finishes of the same
                        # design+detail combination.
                        family_key = f"vitra:{_slug(ws.title)}:{_slug(design)}:{_slug(detail)}"

                        # Locate the closest image anchor within a narrow row
                        # window (±2), constrained to THIS finish's own image
                        # column (±1, to tolerate the anchor landing on the
                        # neighbouring cell boundary within the same 4-column
                        # group). Never search outside this finish's own
                        # column range — that would borrow a sibling finish's
                        # photo, which is the bug this fix removes. If this
                        # finish truly has no photo of its own, it stays
                        # unmapped (no fabrication, no borrowing).
                        best: ExtractedImage | None = None
                        best_score = (99, 99)
                        if image_col is not None:
                            for (a_row, a_col) in anchors_by_sheet.get(ws.title, []):
                                d_col = abs(a_col - image_col)
                                if d_col > 1:
                                    continue
                                d_row = abs(a_row - r_idx)
                                if d_row > 2:
                                    continue
                                score = (d_row, d_col)
                                if score < best_score:
                                    candidate = by_sheet_row_col.get((ws.title, a_row, a_col))
                                    if candidate is not None:
                                        best = candidate
                                        best_score = score
                                        if score == (0, 0):
                                            break
                        image_urls = [best.data_url] if best else []
                        image_meta = [best.to_dict()] if best else []
                        image_quality = best.quality if best else "missing"
                        if best:
                            report.images_mapped += 1

                        finish_label = finish.title()
                        finish_code = _finish_code_from_header(finish)
                        # Colour = finish header stripped of the numeric code
                        # AND anything after a "/" (Vitra sometimes writes
                        # "WHITE 003/403" meaning code 003 with 403 as alt);
                        # trailing punctuation is also removed.
                        colour = re.sub(r"\s*\d{3}\s*", " ", finish_label)
                        colour = re.sub(r"[/\s]+$", "", colour)
                        colour = re.sub(r"^[/\s]+", "", colour)
                        colour = re.sub(r"\s+", " ", colour).strip()
                        if not colour:
                            colour = finish_label

                        family_name = f"{design} · {detail}".strip(" ·") or design or MISSING
                        product_name = f"{design} {detail} · {colour}".strip(" ·") or family_name

                        specs = {}
                        if detail:
                            specs["form"] = detail
                        if extras:
                            specs["related_codes"] = dedupe_iter(extras)

                        pr = ProductRow(
                            brand=self.brand, sku=primary_sku,
                            name=product_name,
                            category=cat,
                            subcategory=subcat,
                            series=design or MISSING,
                            family_key=family_key,
                            variant=colour or finish_label,
                            finish=colour or finish_label,
                            finish_code=finish_code,
                            colour=colour or finish_label,
                            dimensions=detail if detail else MISSING,
                            description=family_name,
                            mrp=mrp, dealer_price=MISSING,
                            images=image_urls, image_meta=image_meta,
                            image_quality=image_quality,
                            image_page=None,
                            accessories=extras,
                            specs=specs,
                            tags=dedupe_iter([
                                cat if cat != MISSING else "",
                                self.brand.lower(),
                                (colour or "").lower(),
                                design.lower() if design else "",
                            ]),
                            confidence=0.96 if mrp != MISSING and cat != MISSING else 0.7,
                        )
                        if mrp == MISSING:
                            pr.issues.append("Missing MRP for this finish")
                        if cat == MISSING:
                            pr.issues.append("Category not derivable from sheet + detail")
                        if not best:
                            pr.issues.append("No image mapped from supplier file")
                        elif best.quality == "poor":
                            pr.issues.append(f"Only thumbnail-grade image available ({best.longest_edge}px longest edge)")
                        rows.append(pr)
                else:
                    joined = " | ".join(str(c) for c in row if c)
                    m = SKU_RE.search(joined)
                    if m:
                        cat = _classify(ws.title, detail)
                        rows.append(ProductRow(
                            brand=self.brand, sku=m.group(0),
                            name=(design + " · " + detail).strip(" ·") or MISSING,
                            category=cat,
                            subcategory=_extract_subcategory(ws.title, detail, cat),
                            series=design or MISSING,
                            family_key=f"vitra:{_slug(ws.title)}:{_slug(design)}",
                            image_quality="missing",
                            confidence=0.55,
                        ))

        # Same-SKU-different-payload conflict guard: the supplier file can
        # (rarely) list the identical code under two different finish
        # columns with different colours/prices — a genuine source-file
        # error, not something this pipeline should silently resolve by
        # picking one and discarding the other. Flag both, drop confidence
        # below the certifier's auto-accept threshold, and surface it in the
        # import report so a human decides which (if either) is correct.
        by_sku: dict[str, list[ProductRow]] = defaultdict(list)
        for r in rows:
            by_sku[r.sku].append(r)
        for sku, group in by_sku.items():
            if len(group) < 2:
                continue
            distinct_payloads = {(g.colour, g.mrp) for g in group}
            if len(distinct_payloads) > 1:
                for g in group:
                    g.confidence = min(g.confidence, 0.4)
                    g.issues.append(
                        f"Conflict: SKU {sku!r} appears {len(group)}x in the supplier file "
                        f"with different colour/price combinations {sorted(distinct_payloads)} — needs manual review"
                    )

        report.parsed_rows = len(rows)
        return rows, report
