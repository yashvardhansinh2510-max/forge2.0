"""Generic rows->CSV/XLSX export, shared by every Phase 2 table. CSV and XLSX
only — see the plan's Global Constraints for why PDF is explicitly out of
scope this phase."""
from __future__ import annotations

import csv
import io

import openpyxl

from services.export import export_response, rows_to_csv, rows_to_xlsx

COLUMNS = [("name", "Name"), ("revenue", "Revenue")]
ROWS = [{"name": "ABC Architects", "revenue": 1200000.0}, {"name": "XYZ Interiors", "revenue": 300000.0}]


def test_csv_header_row_uses_the_declared_labels():
    text = rows_to_csv(ROWS, COLUMNS).decode("utf-8-sig")
    reader = csv.reader(io.StringIO(text))
    header = next(reader)
    assert header == ["Name", "Revenue"]


def test_csv_data_rows_follow_column_order():
    text = rows_to_csv(ROWS, COLUMNS).decode("utf-8-sig")
    reader = list(csv.reader(io.StringIO(text)))
    assert reader[1] == ["ABC Architects", "1200000.0"]


def test_csv_handles_a_row_missing_a_declared_key():
    text = rows_to_csv([{"name": "No Revenue Field"}], COLUMNS).decode("utf-8-sig")
    reader = list(csv.reader(io.StringIO(text)))
    assert reader[1] == ["No Revenue Field", ""]


def test_csv_is_safe_with_the_rupee_symbol():
    text = rows_to_csv([{"name": "₹ Test", "revenue": 1.0}], COLUMNS).decode("utf-8-sig")
    assert "₹ Test" in text


def test_xlsx_roundtrips_through_openpyxl():
    data = rows_to_xlsx(ROWS, COLUMNS, sheet_title="Referrers")
    wb = openpyxl.load_workbook(io.BytesIO(data))
    ws = wb["Referrers"]
    assert [c.value for c in ws[1]] == ["Name", "Revenue"]
    assert [c.value for c in ws[2]] == ["ABC Architects", 1200000.0]


def test_export_response_rejects_an_unknown_format():
    import pytest
    with pytest.raises(ValueError):
        export_response(ROWS, COLUMNS, "referrers", "pdf")


def test_export_response_sets_a_content_disposition_filename():
    response = export_response(ROWS, COLUMNS, "referrers", "csv")
    assert "referrers.csv" in response.headers["content-disposition"]
