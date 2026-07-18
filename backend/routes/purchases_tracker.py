"""Purchases — Material Tracker.

A per-LINE-ITEM lifecycle tracker built on top of the PO document store.

Each PO line item flows through 6 stages:
    order_in_company → company_billing → in_box → dispatched → in_transit → delivered

Endpoints:
    GET  /purchases/items                     — flat cross-PO items list
    GET  /purchases/brands                    — brand facets w/ counts
    GET  /purchases/customers                 — customer facets w/ counts
    GET  /purchases/stages                    — stage catalog + counts
    GET  /purchases/dispatch-record           — history for dispatched+ stages
    GET  /purchases/settings                  — {sla_days}
    POST /purchases/settings                  — update SLA
    POST /purchases/items/{item_id}/move      — advance a single item
    POST /purchases/items/bulk-move           — advance many items in one call
    POST /purchases/items/{item_id}/transfer  — reduce qty on source, create
                                                new PO for destination customer
    GET  /purchases/export.xlsx               — filtered .xlsx export

The tracker does NOT concern itself with taxes — Forge uses final prices only.
"""
from __future__ import annotations
import asyncio
import io
from datetime import datetime, timezone
from typing import Optional
from uuid import uuid4

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from pydantic import BaseModel, Field

from auth import floor_scope_ids, get_current_user, require_min_role
from db import db
from models import (
    PurchaseOrder, PurchaseOrderItem, PurchaseShortage, PurchaseStageEvent, PurchaseStatusEvent,
    PURCHASE_STAGES, PurchaseStage, Quotation, QuotationLineItem, UserPublic, now_iso,
)
from routes.purchase_routes import ALLOWED_TRANSITIONS, STATUS_LABELS
from routes.quotation_routes import _next_number as _next_quotation_number
from services.activity_log import log_event, timeline_for
from services.followup_engine import reconcile_followups
from services.transfer_workflow import execute_transfer, transfer_history

router = APIRouter(prefix="/purchases", tags=["purchases"])


# =============================================================================
# Stage catalog — the ONE source of truth used by both backend + frontend.
# =============================================================================
STAGE_LABELS: dict[str, str] = {
    "order_in_company": "Order in Company",
    "company_billing":  "Company Billing",
    "in_box":           "In Box",
    "dispatched":       "Dispatched",
    "in_transit":       "In Transit",
    "delivered":        "Delivered",
}

# UI tones aligned with the reference (light-mode badges).
STAGE_TONES: dict[str, dict[str, str]] = {
    "order_in_company": {"bg": "#FFEDD5", "fg": "#9A3412"},   # amber
    "company_billing":  {"bg": "#FEF3C7", "fg": "#92400E"},   # yellow
    "in_box":           {"bg": "#DBEAFE", "fg": "#1E40AF"},   # blue
    "dispatched":       {"bg": "#DCFCE7", "fg": "#166534"},   # green
    "in_transit":       {"bg": "#E9D5FF", "fg": "#6B21A8"},   # purple
    "delivered":        {"bg": "#D1FAE5", "fg": "#065F46"},   # emerald
}

EARLY_STAGES = ("order_in_company", "company_billing", "in_box")
DISPATCH_STAGES = ("dispatched", "in_transit", "delivered")

DEFAULT_SLA_DAYS = 7
SETTINGS_KEY = "purchases_tracker"


# =============================================================================
# Settings — data-driven Blocked SLA (default 7d, configurable)
# =============================================================================
class TrackerSettings(BaseModel):
    sla_days: int = Field(default=DEFAULT_SLA_DAYS, ge=1, le=365)


async def _load_settings() -> TrackerSettings:
    doc = await db.settings.find_one({"key": SETTINGS_KEY}, {"_id": 0})
    if not doc:
        return TrackerSettings()
    return TrackerSettings(**{k: v for k, v in doc.items() if k in TrackerSettings.__fields__})


@router.get("/settings", response_model=TrackerSettings)
async def get_settings(_: UserPublic = Depends(get_current_user)):
    return await _load_settings()


@router.post("/settings", response_model=TrackerSettings)
async def update_settings(
    body: TrackerSettings,
    user: UserPublic = Depends(require_min_role("manager")),
):
    await db.settings.update_one(
        {"key": SETTINGS_KEY},
        {"$set": {
            "key": SETTINGS_KEY,
            "sla_days": int(body.sla_days),
            "updated_at": now_iso(),
            "updated_by": user.id,
            "updated_by_name": user.full_name,
        }},
        upsert=True,
    )
    await log_event(
        event_type="purchase.settings.updated",
        entity_type="settings", entity_id=SETTINGS_KEY, actor=user,
        summary=f"Purchases SLA set to {body.sla_days} days",
        payload={"sla_days": body.sla_days},
    )
    return body


# =============================================================================
# Helpers
# =============================================================================
def _iso(ts: datetime) -> str:
    if ts.tzinfo is None:
        ts = ts.replace(tzinfo=timezone.utc)
    return ts.isoformat()


def _parse_iso(s: Optional[str]) -> Optional[datetime]:
    if not s:
        return None
    try:
        # Support trailing Z
        return datetime.fromisoformat(s.replace("Z", "+00:00"))
    except Exception:
        return None


def _age_days(iso_ts: Optional[str]) -> Optional[int]:
    dt = _parse_iso(iso_ts)
    if not dt:
        return None
    now = datetime.now(timezone.utc)
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return max(0, (now - dt).days)


def _flatten_item(po: dict, it: dict, sla_days: int) -> dict:
    """Materialize a single tracker row from a PO doc + its item sub-doc."""
    stage = it.get("stage") or "order_in_company"
    last_moved_at = it.get("last_moved_at") or po.get("created_at")
    age = _age_days(last_moved_at) or 0
    blocked = (stage in EARLY_STAGES) and (age >= sla_days)
    return {
        "item_id": it["id"],
        "po_id": po["id"],
        "po_number": po.get("number"),
        "po_status": po.get("status"),
        "quotation_id": po.get("quotation_id"),
        "quotation_number": po.get("quotation_number"),
        "quotation_line_id": it.get("quotation_line_id"),
        "product_id": it.get("product_id"),
        "sku": it.get("sku"),
        "name": it.get("name"),
        "image": it.get("image"),
        "finish": it.get("finish"),
        "colour": it.get("colour"),
        "customer_id": it.get("customer_id") or po.get("customer_id"),
        "customer_name": it.get("customer_name") or po.get("customer_name"),
        "brand_id": it.get("brand_id") or po.get("brand_id"),
        "brand_name": it.get("brand_name") or po.get("brand_name"),
        "supplier_id": po.get("supplier_id"),
        "supplier_name": po.get("supplier_name"),
        "stage": stage,
        "stage_label": STAGE_LABELS.get(stage, stage),
        "stage_tone": STAGE_TONES.get(stage, {"bg": "#F4F4F5", "fg": "#3F3F46"}),
        "qty": float(it.get("qty") or 0),
        "unit_cost": float(it.get("unit_cost") or 0),
        "room": it.get("room"),
        "expected_delivery_at": po.get("expected_delivery_at"),
        "last_moved_at": last_moved_at,
        "last_moved_by_name": it.get("last_moved_by_name") or po.get("created_by_name"),
        "age_days": age,
        "blocked": blocked,
        "sla_days": sla_days,
        "split_from_item_id": it.get("split_from_item_id"),
        "transferred_from_item_id": it.get("transferred_from_item_id"),
        "transferred_from_customer_id": it.get("transferred_from_customer_id"),
    }


async def _iter_items(
    view: str,
    brand: Optional[str],
    customer: Optional[str],
    stage: Optional[str],
    q: Optional[str],
    sla_days: int,
    limit: int = 2000,
    product_id: Optional[str] = None,
    floor_ids: Optional[list[str]] = None,
) -> list[dict]:
    """Return a flat list of tracker rows across all POs, filtered."""
    match: dict = {"status": {"$ne": "cancelled"}}
    if floor_ids is not None:
        match["floor_id"] = {"$in": floor_ids}
    if q:
        term = {"$regex": q, "$options": "i"}
        match["$or"] = [
            {"number": term},
            {"customer_name": term},
            {"items.sku": term},
            {"items.name": term},
        ]
    if brand and brand.lower() != "all":
        match["brand_id"] = brand
    if customer:
        match["customer_id"] = customer
    if product_id:
        match["items.product_id"] = product_id

    pipeline: list[dict] = [
        {"$match": match},
        {"$unwind": "$items"},
        {"$project": {
            "_id": 0,
            "id": 1, "number": 1, "customer_id": 1, "customer_name": 1,
            "brand_id": 1, "brand_name": 1, "quotation_id": 1, "quotation_number": 1,
            "created_at": 1, "created_by_name": 1, "status": 1,
            "supplier_id": 1, "supplier_name": 1, "expected_delivery_at": 1,
            "items": 1,
        }},
    ]

    if product_id:
        pipeline.append({"$match": {"items.product_id": product_id}})
    if view == "dispatch_record":
        pipeline.append({"$match": {"items.stage": {"$in": list(DISPATCH_STAGES)}}})
    elif stage:
        if stage not in PURCHASE_STAGES:
            raise HTTPException(status_code=400, detail=f"Unknown stage '{stage}'")
        pipeline.append({"$match": {"items.stage": stage}})

    docs = await db.purchase_orders.aggregate(pipeline).to_list(limit * 3)

    rows: list[dict] = []
    for d in docs:
        po = {k: v for k, v in d.items() if k != "items"}
        po["id"] = d.get("id")
        it = d.get("items") or {}
        rows.append(_flatten_item(po, it, sla_days))

    # Item-level search (for the sub-doc match already done, but also allow
    # customer_name text search post-unwind).
    if q:
        term = q.lower()
        rows = [r for r in rows if any(term in str(r.get(k) or "").lower() for k in ("sku", "name", "customer_name", "po_number", "brand_name"))]

    if view == "today":
        # Blocked/aging early-stage items first, newest activity next.
        rows = [r for r in rows if r["stage"] not in DISPATCH_STAGES or r["stage"] == "dispatched"]
        rows.sort(key=lambda r: (not r["blocked"], -(r.get("age_days") or 0)))
    elif view == "stock":
        rows.sort(key=lambda r: (r["stage"] == "delivered", -(r.get("age_days") or 0)))
    elif view == "customers":
        rows.sort(key=lambda r: ((r.get("customer_name") or "").lower(), r.get("po_number") or ""))
    elif view == "dispatch_record":
        rows.sort(key=lambda r: r.get("last_moved_at") or "", reverse=True)
    else:
        rows.sort(key=lambda r: r.get("last_moved_at") or "", reverse=True)

    return rows[:limit]


# =============================================================================
# Read endpoints
# =============================================================================
@router.get("/stages")
async def stage_catalog(_: UserPublic = Depends(get_current_user)):
    """Stage list with counts across ALL non-cancelled items."""
    pipeline = [
        {"$match": {"status": {"$ne": "cancelled"}}},
        {"$unwind": "$items"},
        {"$group": {"_id": {"$ifNull": ["$items.stage", "order_in_company"]}, "count": {"$sum": 1}}},
    ]
    rows = await db.purchase_orders.aggregate(pipeline).to_list(20)
    counts = {r["_id"]: r["count"] for r in rows}
    return [
        {"key": k, "label": STAGE_LABELS[k], "count": counts.get(k, 0), "tone": STAGE_TONES[k]}
        for k in PURCHASE_STAGES
    ]


@router.get("/brands")
async def brand_facets(_: UserPublic = Depends(get_current_user)):
    """Brand list with counts of tracked items."""
    pipeline = [
        {"$match": {"status": {"$ne": "cancelled"}}},
        {"$unwind": "$items"},
        {"$group": {
            "_id": {"id": "$brand_id", "name": "$brand_name"},
            "count": {"$sum": 1},
        }},
        {"$sort": {"count": -1}},
    ]
    rows = await db.purchase_orders.aggregate(pipeline).to_list(50)
    total = sum(r["count"] for r in rows)
    facets = [
        {"id": r["_id"].get("id"), "name": r["_id"].get("name") or "Unbranded", "count": r["count"]}
        for r in rows if r["_id"].get("id")
    ]
    return {"all": total, "brands": facets}


@router.get("/customers")
async def customer_facets(_: UserPublic = Depends(get_current_user)):
    """Customers with open (non-delivered) tracked items."""
    pipeline = [
        {"$match": {"status": {"$ne": "cancelled"}}},
        {"$unwind": "$items"},
        {"$group": {
            "_id": {"id": "$customer_id", "name": "$customer_name"},
            "count": {"$sum": 1},
            "open": {"$sum": {"$cond": [{"$in": ["$items.stage", ["delivered"]]}, 0, 1]}},
        }},
        {"$sort": {"open": -1, "count": -1}},
    ]
    rows = await db.purchase_orders.aggregate(pipeline).to_list(500)
    return [
        {
            "id": r["_id"].get("id"),
            "name": r["_id"].get("name"),
            "count": r["count"],
            "open": r["open"],
        }
        for r in rows if r["_id"].get("id")
    ]


@router.get("/customers/{customer_id}/workspace")
async def customer_workspace(customer_id: str, _: UserPublic = Depends(get_current_user)):
    """One-call aggregate powering the Customer Purchase Workspace: summary,
    products ordered, brand/stage breakdowns, POs, outstanding items, recent
    activity, and expected delivery. Everything here is derived live from the
    same PO/item documents the rest of Purchases uses — no separate cache to
    go stale.
    """
    customer = await db.customers.find_one({"id": customer_id}, {"_id": 0})
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")

    settings = await _load_settings()
    rows = await _iter_items("stock", None, customer_id, None, None, settings.sla_days, limit=2000)
    pos = await db.purchase_orders.find({"customer_id": customer_id}, {"_id": 0}).sort("created_at", -1).to_list(200)

    total_value = round(sum(r["qty"] * r["unit_cost"] for r in rows), 2)
    outstanding_rows = [r for r in rows if r["stage"] != "delivered"]
    outstanding_value = round(sum(r["qty"] * r["unit_cost"] for r in outstanding_rows), 2)
    blocked_rows = [r for r in rows if r["blocked"]]
    delivered_rows = [r for r in rows if r["stage"] == "delivered"]

    brand_map: dict = {}
    for r in rows:
        key = r.get("brand_id") or "unbranded"
        b = brand_map.setdefault(key, {"id": r.get("brand_id"), "name": r.get("brand_name") or "Unbranded", "count": 0})
        b["count"] += 1
    stage_counts = {k: 0 for k in PURCHASE_STAGES}
    for r in rows:
        stage_counts[r["stage"]] = stage_counts.get(r["stage"], 0) + 1
    stages = [
        {"key": k, "label": STAGE_LABELS[k], "count": stage_counts.get(k, 0), "tone": STAGE_TONES[k]}
        for k in PURCHASE_STAGES
    ]

    activity = await timeline_for(customer_id=customer_id, limit=15)

    shortages = await db.purchase_shortages.find(
        {"customer_id": customer_id, "status": "awaiting_reorder"}, {"_id": 0},
    ).sort("created_at", -1).to_list(100)

    order_quotes = await db.quotations.find(
        {"customer_id": customer_id, "status": {"$in": ["ordered", "won"]}},
        {"_id": 0, "id": 1, "grand_total": 1},
    ).to_list(500)
    from routes.payment_routes import _paid_by_quotation
    paid_map = await _paid_by_quotation([quote["id"] for quote in order_quotes])
    outstanding_balance = round(sum(
        max(0.0, float(quote.get("grand_total") or 0) - paid_map.get(quote["id"], 0.0))
        for quote in order_quotes
    ), 2)
    payments = await db.payments.find({"customer_id": customer_id}, {"_id": 0}).sort("created_at", -1).to_list(100)
    followups = await db.followups.find(
        {"customer_id": customer_id, "status": {"$in": ["open", "snoozed"]}}, {"_id": 0},
    ).sort("due_at", 1).to_list(100)

    open_pos = [p for p in pos if p.get("status") != "cancelled"]
    expected = sorted(
        [
            {"po_id": p["id"], "po_number": p.get("number"), "expected_delivery_at": p.get("expected_delivery_at")}
            for p in open_pos if p.get("expected_delivery_at")
        ],
        key=lambda x: x["expected_delivery_at"],
    )

    return {
        "customer": {
            "id": customer["id"], "name": customer.get("name"), "company": customer.get("company"),
            "tier": customer.get("tier"), "email": customer.get("email"), "phone": customer.get("phone"),
        },
        "summary": {
            "total_items": len(rows),
            "total_value": total_value,
            "outstanding_value": outstanding_value,
            "outstanding_count": len(outstanding_rows),
            "open_pos": len(open_pos),
            "blocked_count": len(blocked_rows),
            "delivered_count": len(delivered_rows),
            "shortage_count": len(shortages),
            "outstanding_balance": outstanding_balance,
            "open_followup_count": len(followups),
        },
        "shortages": shortages,
        "payments": payments,
        "followups": followups,
        "products": rows,
        "brands": sorted(brand_map.values(), key=lambda x: -x["count"]),
        "stages": stages,
        "purchase_orders": [
            {
                "id": p["id"], "number": p.get("number"), "status": p.get("status"),
                "brand_name": p.get("brand_name"), "supplier_name": p.get("supplier_name"),
                "grand_total": p.get("grand_total"), "created_at": p.get("created_at"),
                "expected_delivery_at": p.get("expected_delivery_at"),
                "item_count": len(p.get("items") or []),
            }
            for p in pos
        ],
        "outstanding_items": outstanding_rows,
        "recent_activity": activity,
        "expected_delivery": {
            "next_at": expected[0]["expected_delivery_at"] if expected else None,
            "purchase_orders": expected[:5],
        },
    }


@router.get("/items")
async def list_items(
    view: str = Query("stock", regex="^(today|stock|customers|dispatch_record)$"),
    brand: Optional[str] = None,
    customer: Optional[str] = None,
    stage: Optional[str] = None,
    q: Optional[str] = None,
    product_id: Optional[str] = None,
    limit: int = Query(500, ge=1, le=2000),
    user: UserPublic = Depends(get_current_user),
):
    """Flat tracker rows filtered by view/brand/customer/stage/q/product_id."""
    settings = await _load_settings()
    rows = await _iter_items(
        view, brand, customer, stage, q, settings.sla_days, limit, product_id,
        floor_ids=floor_scope_ids(user),
    )

    blocked_count = sum(1 for r in rows if r["blocked"])
    return {
        "sla_days": settings.sla_days,
        "count": len(rows),
        "blocked_count": blocked_count,
        "items": rows,
    }


@router.get("/dispatch-record")
async def dispatch_record(
    brand: Optional[str] = None,
    customer: Optional[str] = None,
    q: Optional[str] = None,
    limit: int = Query(500, ge=1, le=2000),
    _: UserPublic = Depends(get_current_user),
):
    settings = await _load_settings()
    rows = await _iter_items("dispatch_record", brand, customer, None, q, settings.sla_days, limit)
    return {"count": len(rows), "items": rows}


@router.get("/items/{item_id}")
async def get_item(item_id: str, _: UserPublic = Depends(get_current_user)):
    po = await db.purchase_orders.find_one({"items.id": item_id}, {"_id": 0})
    if not po:
        raise HTTPException(status_code=404, detail="Item not found")
    it = next((i for i in po.get("items", []) if i.get("id") == item_id), None)
    if not it:
        raise HTTPException(status_code=404, detail="Item not found")
    settings = await _load_settings()
    row = _flatten_item(po, it, settings.sla_days)
    row["stage_history"] = it.get("stage_history") or []
    row["po_status"] = po.get("status")
    return row


# =============================================================================
# Mutations — move, bulk move, transfer
# =============================================================================
class MoveBody(BaseModel):
    stage: PurchaseStage
    note: Optional[str] = None
    qty: Optional[float] = Field(default=None, gt=0)  # partial move — e.g. "3 of 20"


class BulkMoveBody(BaseModel):
    item_ids: list[str]
    stage: PurchaseStage
    note: Optional[str] = None


class TransferBody(BaseModel):
    destination_customer_id: Optional[str] = None
    new_customer_id: Optional[str] = None  # legacy request compatibility
    new_customer: Optional[dict] = None
    qty: float = Field(..., gt=0)
    reason: Optional[str] = None
    idempotency_key: Optional[str] = None


def _derive_po_status_from_stages(items: list[dict], current_status: str) -> Optional[str]:
    """Reconcile the PO-level `status` (purchase_routes.py state machine) with
    the per-item `stage` (Material Tracker). These are two views of the same
    physical reality — incoming-from-supplier receiving progress — and MUST
    never silently diverge (e.g. Kanban shows "Draft" while every line is
    physically "Delivered"). Returns the target status, or None if no forward
    sync is warranted (PO is cancelled, or already at/past the target).

    Mapping (supplier → our warehouse):
        order_in_company / company_billing / in_box  → still "ordered" once
            procurement has actually started moving it (was draft/awaiting_review)
        dispatched / in_transit                       → "awaiting_supplier"
        delivered (some, not all)                     → "partial_received"
        delivered (all)                                → "fully_received"

    We only ever move status FORWARD relative to ALLOWED_TRANSITIONS reachability
    from the current status, and we never touch a PO that is "cancelled" or has
    already progressed to "packed"/"ready_for_dispatch" (post-receiving, dispatch
    to the customer is a distinct, later concern the Material Tracker doesn't
    own).
    """
    if current_status in ("cancelled", "packed", "ready_for_dispatch"):
        return None
    if not items:
        return None
    stages = [i.get("stage") or "order_in_company" for i in items]

    if all(s == "delivered" for s in stages):
        target = "fully_received"
    elif any(s == "delivered" for s in stages):
        target = "partial_received"
    elif any(s in ("dispatched", "in_transit") for s in stages):
        target = "awaiting_supplier"
    elif any(s in ("company_billing", "in_box") for s in stages):
        target = "ordered"
    else:
        return None  # every item still order_in_company — nothing to sync yet

    if target == current_status:
        return None

    # Rank statuses so we never move backwards (e.g. a late single-item
    # re-move to an earlier stage shouldn't downgrade a PO that's otherwise
    # fully_received from other items already reconciled).
    rank = ["draft", "awaiting_review", "ordered", "awaiting_supplier",
            "partial_received", "fully_received"]
    try:
        if rank.index(target) <= rank.index(current_status):
            return None
    except ValueError:
        pass
    return target


async def _sync_po_status_with_stages(po_id: str, user: UserPublic) -> Optional[str]:
    """Fetch the fresh PO, derive the correct status from item stages, and
    persist it (+ status_history + qty_received bookkeeping + activity event)
    if a forward sync is warranted. Returns the new status, or None."""
    fresh = await db.purchase_orders.find_one({"id": po_id}, {"_id": 0})
    if not fresh:
        return None
    target = _derive_po_status_from_stages(fresh.get("items", []), fresh.get("status", "draft"))
    if not target:
        return None

    now = now_iso()
    ev = PurchaseStatusEvent(
        from_status=fresh.get("status"), to_status=target,
        by_user_id=user.id, by_user_name=user.full_name,
        note="Auto-synced from Material Tracker stage change",
    ).dict()
    set_fields: dict = {"status": target, "updated_at": now}
    await db.purchase_orders.update_one(
        {"id": po_id},
        {"$set": set_fields, "$push": {"status_history": ev}},
    )
    # Keep qty_received consistent with /receive bookkeeping for every item
    # that has physically reached "delivered" — otherwise the receive-progress
    # math (partial vs full) on the PO-lifecycle side would still read 0.
    for it2 in fresh.get("items", []):
        if it2.get("stage") == "delivered" and float(it2.get("qty_received") or 0) < float(it2.get("qty") or 0):
            await db.purchase_orders.update_one(
                {"id": po_id, "items.id": it2["id"]},
                {"$set": {"items.$.qty_received": it2.get("qty")}},
            )

    await log_event(
        event_type="purchase.status_changed",
        entity_type="purchase", entity_id=po_id, actor=user,
        customer_id=fresh.get("customer_id"),
        summary=f"Status auto-synced to {STATUS_LABELS.get(target, target)} (Material Tracker)",
        payload={"from": fresh.get("status"), "to": target, "source": "material_tracker_sync"},
    )
    return target


async def _apply_stage_change(
    item_id: str, to_stage: str, user: UserPublic, note: Optional[str],
    qty: Optional[float] = None,
) -> dict:
    """Atomic update of a single item's stage — writes stage_history entry.

    If `qty` is given and is LESS than the item's current quantity, this is a
    PARTIAL move ("3 of 20") — the line is split: a brand-new tracked item is
    created at `to_stage` carrying `qty` units, and the original item's
    quantity is reduced by that amount and stays at its current stage. Full
    lineage is recorded both ways (split_out on the original, split_in on the
    new piece) so the Movement History for either piece traces back to the
    other.
    """
    po = await db.purchase_orders.find_one({"items.id": item_id}, {"_id": 0})
    if not po:
        raise HTTPException(status_code=404, detail="Item not found")
    it = next((i for i in po.get("items", []) if i.get("id") == item_id), None)
    if not it:
        raise HTTPException(status_code=404, detail="Item not found")
    from_stage = it.get("stage") or "order_in_company"
    full_qty = float(it.get("qty") or 0)
    move_qty = float(qty) if qty is not None else full_qty
    if move_qty <= 0:
        raise HTTPException(status_code=400, detail="Move quantity must be greater than 0")
    if move_qty > full_qty + 1e-6:
        raise HTTPException(status_code=400, detail=f"Only {full_qty} available to move")

    is_partial = move_qty < full_qty - 1e-6

    if not is_partial and from_stage == to_stage:
        return {"po_id": po["id"], "item_id": item_id, "no_change": True}

    now = now_iso()

    # -------------------------------------------------------------------
    # PARTIAL MOVE — split the line into two tracked pieces.
    # -------------------------------------------------------------------
    if is_partial:
        remaining = round(full_qty - move_qty, 4)
        new_item_id = str(uuid4())

        move_ev = PurchaseStageEvent(
            from_stage=from_stage, to_stage=to_stage,
            by_user_id=user.id, by_user_name=user.full_name,
            note=note, action="split_in",
            ref_item_id=item_id, ref_po_id=po["id"], qty=move_qty,
        ).dict()
        new_item = dict(it)
        new_item["id"] = new_item_id
        new_item["qty"] = move_qty
        new_item["qty_received"] = 0
        new_item["stage"] = to_stage
        new_item["last_moved_at"] = now
        new_item["last_moved_by"] = user.id
        new_item["last_moved_by_name"] = user.full_name
        new_item["split_from_item_id"] = item_id
        new_item["split_into_item_id"] = None
        new_item["stage_history"] = list(it.get("stage_history") or []) + [move_ev]

        src_ev = PurchaseStageEvent(
            from_stage=from_stage, to_stage=from_stage,
            by_user_id=user.id, by_user_name=user.full_name,
            note=note or f"Split {move_qty} of {full_qty} → {STAGE_LABELS.get(to_stage, to_stage)}",
            action="split_out",
            ref_item_id=new_item_id, ref_po_id=po["id"], qty=move_qty,
        ).dict()

        await db.purchase_orders.update_one(
            {"id": po["id"], "items.id": item_id},
            {
                "$set": {
                    "items.$.qty": remaining,
                    "items.$.split_into_item_id": new_item_id,
                    "items.$.last_moved_at": now,
                    "items.$.last_moved_by": user.id,
                    "items.$.last_moved_by_name": user.full_name,
                    "updated_at": now,
                },
                "$push": {"items.$.stage_history": src_ev},
            },
        )
        await db.purchase_orders.update_one(
            {"id": po["id"]},
            {"$push": {"items": new_item}},
        )

        await log_event(
            event_type="purchase.stage_split",
            entity_type="purchase", entity_id=po["id"], actor=user,
            customer_id=po.get("customer_id"),
            summary=(
                f"{it.get('name')} · split {move_qty} of {full_qty} · "
                f"{STAGE_LABELS.get(from_stage, from_stage)} → {STAGE_LABELS.get(to_stage, to_stage)}"
                + (f" · {note}" if note else "")
            ),
            payload={
                "item_id": item_id, "new_item_id": new_item_id, "po_number": po.get("number"),
                "from_stage": from_stage, "to_stage": to_stage,
                "sku": it.get("sku"), "qty_moved": move_qty, "qty_remaining": remaining,
            },
        )

        new_status = await _sync_po_status_with_stages(po["id"], user)
        if to_stage in ("dispatched", "in_transit", "delivered") or new_status:
            asyncio.create_task(reconcile_followups())

        return {
            "po_id": po["id"], "item_id": item_id, "split": True,
            "new_item_id": new_item_id,
            "from_stage": from_stage, "to_stage": to_stage,
            "qty_moved": move_qty, "qty_remaining": remaining,
            "po_status_synced_to": new_status,
        }

    # -------------------------------------------------------------------
    # FULL MOVE — existing behaviour, whole line advances together.
    # -------------------------------------------------------------------
    ev = PurchaseStageEvent(
        from_stage=from_stage, to_stage=to_stage,
        by_user_id=user.id, by_user_name=user.full_name,
        note=note, action="move", qty=move_qty,
    ).dict()

    await db.purchase_orders.update_one(
        {"id": po["id"], "items.id": item_id},
        {
            "$set": {
                "items.$.stage": to_stage,
                "items.$.last_moved_at": now,
                "items.$.last_moved_by": user.id,
                "items.$.last_moved_by_name": user.full_name,
                "updated_at": now,
            },
            "$push": {"items.$.stage_history": ev},
        },
    )

    await log_event(
        event_type="purchase.stage_moved",
        entity_type="purchase", entity_id=po["id"], actor=user,
        customer_id=po.get("customer_id"),
        summary=(
            f"{it.get('name')} · {STAGE_LABELS.get(from_stage, from_stage)} → "
            f"{STAGE_LABELS.get(to_stage, to_stage)}"
            + (f" · {note}" if note else "")
        ),
        payload={
            "item_id": item_id, "po_number": po.get("number"),
            "from_stage": from_stage, "to_stage": to_stage,
            "sku": it.get("sku"), "qty": it.get("qty"),
        },
    )

    # Reconcile the PO-level status with the new item-stage reality — this is
    # THE fix for the two-systems divergence (Kanban `status` vs Tracker
    # `stage` silently disagreeing). Runs on every move, not just dispatch+.
    new_status = await _sync_po_status_with_stages(po["id"], user)

    if to_stage in ("dispatched", "in_transit", "delivered") or new_status:
        # Dispatch/delivery reminders derive from item stage/PO status —
        # event-triggered refresh right here, not a cron job.
        asyncio.create_task(reconcile_followups())
    return {
        "po_id": po["id"], "item_id": item_id,
        "from_stage": from_stage, "to_stage": to_stage,
        "po_status_synced_to": new_status,
    }


@router.post("/items/{item_id}/move")
async def move_item(
    item_id: str,
    body: MoveBody,
    user: UserPublic = Depends(require_min_role("sales")),
):
    if body.stage not in PURCHASE_STAGES:
        raise HTTPException(status_code=400, detail=f"Unknown stage '{body.stage}'")
    return await _apply_stage_change(item_id, body.stage, user, body.note, body.qty)


@router.post("/items/bulk-move")
async def bulk_move(
    body: BulkMoveBody,
    user: UserPublic = Depends(require_min_role("sales")),
):
    if body.stage not in PURCHASE_STAGES:
        raise HTTPException(status_code=400, detail=f"Unknown stage '{body.stage}'")
    if not body.item_ids:
        raise HTTPException(status_code=400, detail="No items selected")
    results: list[dict] = []
    for iid in body.item_ids:
        try:
            r = await _apply_stage_change(iid, body.stage, user, body.note)
            results.append({"item_id": iid, "ok": True, **r})
        except HTTPException as e:
            results.append({"item_id": iid, "ok": False, "error": e.detail})
    return {"count": len(results), "results": results}


async def _next_po_number() -> str:
    """Same simple counter used by quotation → PO creation."""
    year = datetime.now(timezone.utc).year
    prefix = f"FPO-{year}-"
    last = await db.purchase_orders.find(
        {"number": {"$regex": f"^{prefix}"}}, {"_id": 0, "number": 1}
    ).sort("number", -1).limit(1).to_list(1)
    n = 1
    if last:
        try:
            n = int(last[0]["number"].split("-")[-1]) + 1
        except Exception:
            n = 1
    return f"{prefix}{n:04d}"


async def _selling_price_for(item: dict, po: dict) -> float:
    """Best-effort customer-facing unit price for an auto-generated order.

    Preference order: (1) the ORIGINAL quotation line this item traces back
    to — the fairest number, since it's literally what the customer already
    agreed to pay for this exact SKU; (2) the product's current catalogue
    price; (3) the PO's unit_cost as an honest last resort (better than 0).
    """
    quotation_id = po.get("quotation_id")
    line_id = item.get("quotation_line_id")
    if quotation_id and line_id:
        q = await db.quotations.find_one({"id": quotation_id}, {"_id": 0, "items": 1})
        if q:
            line = next((l for l in q.get("items", []) if l.get("id") == line_id), None)
            if line and line.get("unit_price") is not None:
                return float(line["unit_price"])
    prod = await db.products.find_one({"id": item.get("product_id")}, {"_id": 0, "price": 1})
    if prod and prod.get("price"):
        return float(prod["price"])
    return float(item.get("unit_cost") or 0)


async def _reconcile_shortage_for_line(
    *, quotation_id: Optional[str], quotation_line_id: Optional[str],
    customer_id: str, customer_name: str, product_id: str, sku: str, name: str,
    image: Optional[str], dest_customer_id: str, dest_customer_name: str,
    transferred_qty: float, user: UserPublic,
) -> Optional[dict]:
    """Recompute the shortage for ONE customer's ONE original commitment after
    a transfer moved units away from them. Purely additive — never blocks the
    transfer, only opens/updates/auto-resolves a `purchase_shortages` record.
    """
    if not quotation_id or not quotation_line_id:
        return None  # this item was never tied to a customer commitment — nothing to protect

    q = await db.quotations.find_one({"id": quotation_id}, {"_id": 0})
    if not q:
        return None
    line = next((l for l in q.get("items", []) if l.get("id") == quotation_line_id), None)
    if not line:
        return None
    committed_qty = float(line.get("qty") or 0)
    if committed_qty <= 0:
        return None

    # Allocated = everything still sitting under this customer's name across
    # every PO that traces back to this exact quotation line.
    pipeline = [
        {"$match": {"items.quotation_line_id": quotation_line_id}},
        {"$unwind": "$items"},
        {"$match": {"items.quotation_line_id": quotation_line_id, "items.customer_id": customer_id}},
        {"$group": {"_id": None, "total": {"$sum": "$items.qty"}}},
    ]
    rows = await db.purchase_orders.aggregate(pipeline).to_list(1)
    allocated_qty = float(rows[0]["total"]) if rows else 0.0
    shortage_qty = round(committed_qty - allocated_qty, 4)

    existing = await db.purchase_shortages.find_one(
        {"quotation_line_id": quotation_line_id, "customer_id": customer_id, "status": "awaiting_reorder"},
        {"_id": 0},
    )
    now = now_iso()

    if shortage_qty > 1e-6:
        reason = (
            f"{transferred_qty:g} unit(s) of {name} transferred to {dest_customer_name} — "
            f"{shortage_qty:g} of the original {committed_qty:g} still need to be reordered."
        )
        fields = {
            "customer_id": customer_id, "customer_name": customer_name,
            "quotation_id": quotation_id, "quotation_number": q.get("number"),
            "quotation_line_id": quotation_line_id,
            "product_id": product_id, "sku": sku, "name": name, "image": image,
            "committed_qty": committed_qty, "allocated_qty": allocated_qty, "shortage_qty": shortage_qty,
            "status": "awaiting_reorder", "reason": reason,
            "transferred_to_customer_id": dest_customer_id, "transferred_to_customer_name": dest_customer_name,
            "updated_at": now,
        }
        if existing:
            await db.purchase_shortages.update_one({"id": existing["id"]}, {"$set": fields})
            shortage_id = existing["id"]
        else:
            doc = PurchaseShortage(**fields)
            await db.purchase_shortages.insert_one(doc.dict())
            shortage_id = doc.id
        await log_event(
            event_type="purchase.shortage_flagged",
            entity_type="purchase", entity_id=shortage_id, actor=user,
            customer_id=customer_id, quotation_id=quotation_id,
            summary=f"⚠ Awaiting reorder — {shortage_qty:g} × {name}",
            payload={**fields, "id": shortage_id},
        )
        return {"id": shortage_id, **fields}

    # No shortage (or resolved itself) — close out any previously-open record.
    if existing:
        await db.purchase_shortages.update_one(
            {"id": existing["id"]},
            {"$set": {
                "status": "resolved", "resolved_at": now,
                "resolved_by": user.id, "resolved_by_name": user.full_name,
                "allocated_qty": allocated_qty, "shortage_qty": 0, "updated_at": now,
            }},
        )
    return None

@router.post("/items/{item_id}/transfer")
async def transfer_item_command(
    item_id: str,
    body: TransferBody,
    user: UserPublic = Depends(require_min_role("sales")),
):
    """Transactional, idempotent transfer command for existing or new customers."""
    return await execute_transfer(
        item_id=item_id,
        destination_customer_id=body.destination_customer_id,
        new_customer=body.new_customer,
        qty=float(body.qty),
        reason=body.reason,
        idempotency_key=body.idempotency_key,
        user=user,
    )


@router.get("/items/{item_id}/transfer-history")
async def item_transfer_history(item_id: str, _: UserPublic = Depends(get_current_user)):
    return {"item_id": item_id, "transfers": await transfer_history(item_id)}




@router.post("/legacy/items/{item_id}/transfer")
async def transfer_item(
    item_id: str,
    body: TransferBody,
    user: UserPublic = Depends(require_min_role("sales")),
):
    """Move `qty` units of an item to another customer.

    Steps (atomic per document):
      1. Load source PO and the item.
      2. Validate the requested qty is ≤ current qty.
      3. Reduce qty on source item (deletes the item if it hits 0).
      4. Create a NEW draft PO for the destination customer with a single item
         carrying the transferred qty, at the same stage as the source.
      5. Emit `purchase.transferred_out` on the source and `purchase.transferred_in`
         on the destination — both entries also live in each item's stage_history.
    """
    if body.qty <= 0:
        raise HTTPException(status_code=400, detail="Transfer qty must be > 0")

    po = await db.purchase_orders.find_one({"items.id": item_id}, {"_id": 0})
    if not po:
        raise HTTPException(status_code=404, detail="Source item not found")
    it = next((i for i in po.get("items", []) if i.get("id") == item_id), None)
    if not it:
        raise HTTPException(status_code=404, detail="Source item not found")

    cur_qty = float(it.get("qty") or 0)
    if body.qty > cur_qty + 1e-6:
        raise HTTPException(status_code=400, detail=f"Only {cur_qty} available for transfer")

    src_customer_id = it.get("customer_id") or po.get("customer_id")
    if body.new_customer_id == src_customer_id:
        raise HTTPException(status_code=400, detail="Destination customer must differ from source")

    new_cust = await db.customers.find_one({"id": body.new_customer_id}, {"_id": 0})
    if not new_cust:
        raise HTTPException(status_code=404, detail="Destination customer not found")

    now = now_iso()
    stage = it.get("stage") or "order_in_company"

    # Auto-generate the "order" this transfer represents for the destination
    # customer BEFORE building the PO — Payments/Follow-ups only ever look at
    # Quotations with status ordered/won, so this one insert is what makes
    # "Customer B appears in Payments automatically" true with ZERO changes
    # to the Payments module itself.
    unit_price = await _selling_price_for(it, po)
    auto_line = QuotationLineItem(
        product_id=it["product_id"], sku=it["sku"], name=it["name"],
        image=it.get("image"), category_id=it.get("category_id"), room=it.get("room"),
        qty=float(body.qty), unit_price=unit_price,
    )
    auto_quotation_number = await _next_quotation_number()
    auto_quotation = Quotation(
        number=auto_quotation_number,
        customer_id=body.new_customer_id,
        customer_name=new_cust.get("company") or new_cust.get("name"),
        status="ordered",
        items=[auto_line],
        subtotal=round(auto_line.net, 2),
        grand_total=round(auto_line.net, 2),
        notes=(
            f"Auto-created by transfer — {body.qty:g} × {it['name']} from "
            f"{po.get('customer_name')} ({po.get('number')})"
            + (f" — {body.reason}" if body.reason else "")
        ),
        created_by=user.id, created_by_name=user.full_name,
        source="transfer",
    )

    # Build the destination item — same shape, fresh id, transfer bookkeeping.
    dest_item_id = str(uuid4())
    dest_item = PurchaseOrderItem(
        id=dest_item_id,
        product_id=it["product_id"], sku=it["sku"], name=it["name"],
        image=it.get("image"), category_id=it.get("category_id"), room=it.get("room"),
        qty=float(body.qty), unit_cost=float(it.get("unit_cost") or 0),
        quotation_line_id=auto_line.id,     # now owned by the auto-generated order, not the old one
        stage=stage,
        customer_id=body.new_customer_id,
        customer_name=new_cust.get("company") or new_cust.get("name"),
        brand_id=it.get("brand_id") or po.get("brand_id"),
        brand_name=it.get("brand_name") or po.get("brand_name"),
        last_moved_at=now, last_moved_by=user.id, last_moved_by_name=user.full_name,
        transferred_from_item_id=item_id,
        transferred_from_po_id=po["id"],
        transferred_from_customer_id=src_customer_id,
        stage_history=[
            PurchaseStageEvent(
                from_stage=None, to_stage=stage,
                by_user_id=user.id, by_user_name=user.full_name,
                note=body.reason or f"Transferred from {po.get('customer_name')}",
                action="transfer_in",
                ref_item_id=item_id, ref_po_id=po["id"],
            )
        ],
    )

    # Create destination PO — one-item, draft.
    number = await _next_po_number()
    dest_po = PurchaseOrder(
        number=number,
        quotation_id=auto_quotation.id,
        quotation_number=auto_quotation.number,
        customer_id=body.new_customer_id,
        customer_name=new_cust.get("company") or new_cust.get("name"),
        brand_id=it.get("brand_id") or po.get("brand_id"),
        brand_name=it.get("brand_name") or po.get("brand_name"),
        supplier_id=po.get("supplier_id"),
        supplier_name=po.get("supplier_name"),
        status="draft",
        items=[dest_item],
        internal_notes=(
            f"Transferred from {po.get('number')} · "
            f"{po.get('customer_name')}" + (f" — {body.reason}" if body.reason else "")
        ),
        subtotal=round(dest_item.qty * dest_item.unit_cost, 2),
        grand_total=round(dest_item.qty * dest_item.unit_cost, 2),
        created_by=user.id,
        created_by_name=user.full_name,
        status_history=[
            PurchaseStatusEvent(
                from_status=None, to_status="draft",
                by_user_id=user.id, by_user_name=user.full_name,
                note=f"Customer transfer from {po.get('number')}",
            ).dict()
        ],
    )
    await db.purchase_orders.insert_one(dest_po.dict())
    auto_quotation.source_purchase_order_id = dest_po.id
    auto_quotation.source_item_id = dest_item_id
    await db.quotations.insert_one(auto_quotation.dict())
    await log_event(
        event_type="quotation.auto_created_from_transfer",
        entity_type="quotation", entity_id=auto_quotation.id, actor=user,
        customer_id=body.new_customer_id, quotation_id=auto_quotation.id,
        summary=(
            f"Order {auto_quotation.number} auto-created — ₹{auto_quotation.grand_total:,.0f} due "
            f"from a {body.qty:g}-unit transfer"
        ),
        payload={
            "quotation_id": auto_quotation.id, "quotation_number": auto_quotation.number,
            "grand_total": auto_quotation.grand_total, "dest_po_id": dest_po.id,
            "dest_po_number": dest_po.number, "src_po_number": po.get("number"),
        },
    )

    # Update source — reduce qty or drop the item if 0.
    remaining = round(cur_qty - float(body.qty), 4)
    src_event = PurchaseStageEvent(
        from_stage=stage, to_stage=stage,
        by_user_id=user.id, by_user_name=user.full_name,
        note=body.reason or f"Transferred {body.qty} → {new_cust.get('name')}",
        action="transfer_out",
        ref_item_id=dest_item_id, ref_po_id=dest_po.id,
    ).dict()

    if remaining <= 1e-6:
        # Push an audit event onto the item, then pull it from the array.
        await db.purchase_orders.update_one(
            {"id": po["id"], "items.id": item_id},
            {"$push": {"items.$.stage_history": src_event},
             "$set": {"updated_at": now}},
        )
        await db.purchase_orders.update_one(
            {"id": po["id"]},
            {"$pull": {"items": {"id": item_id}}},
        )
    else:
        await db.purchase_orders.update_one(
            {"id": po["id"], "items.id": item_id},
            {
                "$set": {
                    "items.$.qty": remaining,
                    "items.$.last_moved_at": now,
                    "items.$.last_moved_by": user.id,
                    "items.$.last_moved_by_name": user.full_name,
                    "updated_at": now,
                },
                "$push": {"items.$.stage_history": src_event},
            },
        )

    # Activity events on BOTH ends — customer timelines pick these up.
    await log_event(
        event_type="purchase.transferred_out",
        entity_type="purchase", entity_id=po["id"], actor=user,
        customer_id=src_customer_id,
        summary=(
            f"Transferred {body.qty} × {it.get('name')} to "
            f"{new_cust.get('company') or new_cust.get('name')}"
            + (f" — {body.reason}" if body.reason else "")
        ),
        payload={
            "src_po_id": po["id"], "src_po_number": po.get("number"),
            "dest_po_id": dest_po.id, "dest_po_number": dest_po.number,
            "dest_customer_id": body.new_customer_id,
            "item_id": item_id, "new_item_id": dest_item_id,
            "sku": it.get("sku"), "qty": body.qty, "reason": body.reason,
        },
    )
    await log_event(
        event_type="purchase.transferred_in",
        entity_type="purchase", entity_id=dest_po.id, actor=user,
        customer_id=body.new_customer_id,
        summary=(
            f"Received {body.qty} × {it.get('name')} from "
            f"{po.get('customer_name')}"
            + (f" — {body.reason}" if body.reason else "")
        ),
        payload={
            "src_po_id": po["id"], "src_po_number": po.get("number"),
            "dest_po_id": dest_po.id, "dest_po_number": dest_po.number,
            "src_customer_id": src_customer_id,
            "item_id": item_id, "new_item_id": dest_item_id,
            "sku": it.get("sku"), "qty": body.qty, "reason": body.reason,
        },
    )

    # ---- Shortage tracking for the ORIGINAL customer -------------------------
    # If this item traced back to a real quotation commitment, check whether
    # that commitment is now under-fulfilled and raise/refresh the alert.
    shortage = await _reconcile_shortage_for_line(
        quotation_id=po.get("quotation_id"),
        quotation_line_id=it.get("quotation_line_id"),
        customer_id=src_customer_id, customer_name=po.get("customer_name"),
        product_id=it["product_id"], sku=it["sku"], name=it["name"], image=it.get("image"),
        dest_customer_id=body.new_customer_id,
        dest_customer_name=new_cust.get("company") or new_cust.get("name"),
        transferred_qty=float(body.qty), user=user,
    )

    # Payment/Follow-up automation must reflect this transfer immediately —
    # the new order needs its own reminder timeline, and a shortage (if any)
    # needs to surface as a "recommend reorder" card. Event-triggered, no cron.
    asyncio.create_task(reconcile_followups())

    return {
        "source": {
            "po_id": po["id"], "po_number": po.get("number"),
            "item_id": item_id, "remaining_qty": max(0, remaining),
            "removed": remaining <= 1e-6,
        },
        "destination": {
            "po_id": dest_po.id, "po_number": dest_po.number,
            "item_id": dest_item_id, "qty": float(body.qty),
            "customer_id": body.new_customer_id,
            "customer_name": new_cust.get("company") or new_cust.get("name"),
            "order": {
                "quotation_id": auto_quotation.id, "quotation_number": auto_quotation.number,
                "grand_total": auto_quotation.grand_total, "unit_price": unit_price,
            },
        },
        "shortage": shortage,
    }


# =============================================================================
# Shortages — "Awaiting Reorder" alerts opened by the transfer workflow above.
# =============================================================================
class ShortageDismissBody(BaseModel):
    note: Optional[str] = None


@router.get("/shortages")
async def list_shortages(
    customer_id: Optional[str] = None,
    status_filter: str = Query("awaiting_reorder", alias="status"),
    limit: int = Query(200, ge=1, le=1000),
    _: UserPublic = Depends(get_current_user),
):
    query: dict = {}
    if status_filter and status_filter != "all":
        query["status"] = status_filter
    if customer_id:
        query["customer_id"] = customer_id
    docs = await db.purchase_shortages.find(query, {"_id": 0}).sort("created_at", -1).to_list(limit)
    return {"count": len(docs), "items": docs}


@router.post("/shortages/{shortage_id}/create-po")
async def create_po_for_shortage(
    shortage_id: str,
    user: UserPublic = Depends(require_min_role("purchase")),
):
    """One click on the "Create Purchase Order" recommendation — opens a new
    draft PO for exactly the missing quantity, for the same customer, at the
    first stage. Never runs automatically; a human always presses this."""
    s = await db.purchase_shortages.find_one({"id": shortage_id}, {"_id": 0})
    if not s:
        raise HTTPException(status_code=404, detail="Shortage not found")
    if s.get("status") != "awaiting_reorder":
        raise HTTPException(status_code=400, detail="Shortage is not open")

    now = now_iso()
    item = PurchaseOrderItem(
        product_id=s["product_id"], sku=s["sku"], name=s["name"], image=s.get("image"),
        qty=float(s["shortage_qty"]), unit_cost=0,
        quotation_line_id=s.get("quotation_line_id"),
        stage="order_in_company",
        customer_id=s["customer_id"], customer_name=s["customer_name"],
        last_moved_at=now, last_moved_by=user.id, last_moved_by_name=user.full_name,
        stage_history=[
            PurchaseStageEvent(
                from_stage=None, to_stage="order_in_company",
                by_user_id=user.id, by_user_name=user.full_name,
                note="Reorder for shortage created by a customer transfer", action="create",
            )
        ],
    )
    number = await _next_po_number()
    new_po = PurchaseOrder(
        number=number,
        quotation_id=s.get("quotation_id"), quotation_number=s.get("quotation_number"),
        customer_id=s["customer_id"], customer_name=s["customer_name"],
        status="draft", items=[item],
        internal_notes=f"Reorder — {s.get('reason')}",
        subtotal=0, grand_total=0,
        created_by=user.id, created_by_name=user.full_name,
        status_history=[
            PurchaseStatusEvent(
                from_status=None, to_status="draft",
                by_user_id=user.id, by_user_name=user.full_name,
                note="Created from a shortage recommendation",
            ).dict()
        ],
    )
    await db.purchase_orders.insert_one(new_po.dict())
    await db.purchase_shortages.update_one(
        {"id": shortage_id},
        {"$set": {
            "status": "reordered", "resolved_po_id": new_po.id, "resolved_po_number": new_po.number,
            "resolved_at": now, "resolved_by": user.id, "resolved_by_name": user.full_name,
            "updated_at": now,
        }},
    )
    await log_event(
        event_type="purchase.shortage_reordered",
        entity_type="purchase", entity_id=new_po.id, actor=user,
        customer_id=s["customer_id"], quotation_id=s.get("quotation_id"),
        summary=f"Reorder PO {new_po.number} created for {s['shortage_qty']:g} × {s['name']}",
        payload={"shortage_id": shortage_id, "po_id": new_po.id, "po_number": new_po.number},
    )
    asyncio.create_task(reconcile_followups())
    return {"po_id": new_po.id, "po_number": new_po.number, "shortage_id": shortage_id}


@router.post("/shortages/{shortage_id}/dismiss")
async def dismiss_shortage(
    shortage_id: str,
    body: ShortageDismissBody,
    user: UserPublic = Depends(require_min_role("purchase")),
):
    s = await db.purchase_shortages.find_one({"id": shortage_id}, {"_id": 0})
    if not s:
        raise HTTPException(status_code=404, detail="Shortage not found")
    now = now_iso()
    await db.purchase_shortages.update_one(
        {"id": shortage_id},
        {"$set": {
            "status": "dismissed", "resolved_at": now,
            "resolved_by": user.id, "resolved_by_name": user.full_name,
            "reason": body.note or s.get("reason"), "updated_at": now,
        }},
    )
    asyncio.create_task(reconcile_followups())
    return {"ok": True}


# =============================================================================
# Excel export (real .xlsx via openpyxl)
# =============================================================================
@router.get("/export.xlsx")
async def export_xlsx(
    view: str = Query("stock", regex="^(today|stock|customers|dispatch_record)$"),
    brand: Optional[str] = None,
    customer: Optional[str] = None,
    stage: Optional[str] = None,
    q: Optional[str] = None,
    _: UserPublic = Depends(get_current_user),
):
    settings = await _load_settings()
    rows = await _iter_items(view, brand, customer, stage, q, settings.sla_days, limit=2000)

    wb = Workbook()
    ws = wb.active
    ws.title = "Purchases"

    # Header block
    ws["A1"] = "Forge — Purchases Tracker Export"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:H1")
    stamp = datetime.now(timezone.utc).strftime("%d %b %Y · %H:%M UTC")
    filter_bits = [f"View: {view}"]
    if brand:
        filter_bits.append(f"Brand: {brand}")
    if customer:
        filter_bits.append(f"Customer: {customer}")
    if stage:
        filter_bits.append(f"Stage: {STAGE_LABELS.get(stage, stage)}")
    if q:
        filter_bits.append(f"Search: {q}")
    ws["A2"] = " · ".join(filter_bits) + f" · Exported {stamp}"
    ws["A2"].font = Font(color="6B7280", size=10)
    ws.merge_cells("A2:H2")

    headers = [
        "PO Number", "SKU", "Product", "Customer", "Brand", "Stage",
        "Qty", "Last Move (UTC)", "Moved By", "Age (days)", "Blocked",
    ]
    header_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = Font(bold=True, color="374151", size=11)
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center")

    for i, r in enumerate(rows, start=5):
        ws.cell(row=i, column=1, value=r.get("po_number"))
        ws.cell(row=i, column=2, value=r.get("sku"))
        ws.cell(row=i, column=3, value=r.get("name"))
        ws.cell(row=i, column=4, value=r.get("customer_name"))
        ws.cell(row=i, column=5, value=r.get("brand_name"))
        ws.cell(row=i, column=6, value=r.get("stage_label"))
        ws.cell(row=i, column=7, value=r.get("qty"))
        ws.cell(row=i, column=8, value=r.get("last_moved_at"))
        ws.cell(row=i, column=9, value=r.get("last_moved_by_name"))
        ws.cell(row=i, column=10, value=r.get("age_days"))
        ws.cell(row=i, column=11, value="Yes" if r.get("blocked") else "")

    # Column widths — heuristic
    widths = [16, 14, 40, 26, 14, 20, 8, 24, 20, 12, 10]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A5"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"forge-purchases-{view}-{datetime.now(timezone.utc).strftime('%Y%m%d-%H%M')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
