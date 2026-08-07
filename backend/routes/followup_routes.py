"""Follow-ups · Sales Command Center — the API surface.

Reuses existing collections/helpers everywhere possible:
  * payment aggregation  -> routes.payment_routes._paid_by_quotation / ORDER_STATUSES / _clean_phone
  * timeline             -> services.activity_log.timeline_for
  * scoring / reconcile   -> services.followup_engine

No new business logic is duplicated — this module is orchestration + reads.
"""
from __future__ import annotations

import asyncio
import base64
import csv
import io
import json
import re
from collections import Counter
from datetime import datetime, timedelta, timezone
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from pydantic import BaseModel
from pymongo.errors import DuplicateKeyError

from auth import floor_for_write, floor_inherit, floor_query, get_current_user, get_floor_scoped_or_404, require_floor_access, require_min_role
from db import db
from models import (
    AutomationRuleUpdate, Followup, FollowupCallOutcomePayload, FollowupCompletePayload,
    FollowupContactPayload, FollowupCreate, FollowupSavedView,
    FollowupSavedViewCreate, FollowupSnoozePayload, FollowupUpdate,
    NotebookCellPatchPayload, NotebookConversionPayload, NotebookFollowupCreatePayload,
    UserPublic, now_iso,
)
from services import automation_rules, workflow_transitions
from services.activity_log import log_event, timeline_for
from services.followup_engine import (
    RULE_DEFINITIONS, _followup_sort_key, age_days,
    compute_bucket, ist_day_bounds_utc, money_short, parse_iso,
    reason_factors_for, reconcile_followups, score_followup,
)
from services.followup_notebook import (
    NotebookConflictError, NotebookValidationError, convert_notebook_row,
    notebook_query, notebook_search_query, normalize_mobile, patch_notebook_row,
    serialize_notebook_row, validate_notebook_patch, resolve_or_create_customer,
)

router = APIRouter(prefix="/followups", tags=["followups"])
NOTEBOOK_FLOOR_IDS = frozenset({"second-floor", "third-floor"})


# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────
def require_notebook_floor(floor_id: str, user: UserPublic) -> None:
    """Limit the notebook surface to its two declared floor modules."""
    require_floor_access(floor_id, user)
    if floor_id not in NOTEBOOK_FLOOR_IDS:
        raise HTTPException(status_code=404, detail="Notebook is available only on Kitchen or Furniture Floor")


async def _wake_snoozed() -> None:
    """Self-healing: any snooze whose timer has elapsed flips back to open so
    it resurfaces in the inbox without needing a background job."""
    await db.followups.update_many(
        {"status": "snoozed", "snoozed_until": {"$lte": now_iso()}},
        {"$set": {"status": "open", "updated_at": now_iso()}},
    )


async def _all_with_bucket(user: UserPublic | None = None) -> list[dict]:
    docs = await db.followups.find(floor_query(user, {}) if user else {}, {"_id": 0}).to_list(10000)
    for d in docs:
        d["bucket"] = compute_bucket(d)
        d["effective_priority_level"] = d.get("manual_priority_override") or d.get("priority_level")
    return docs


async def _rule_counts(user: UserPublic) -> dict[str, int]:
    pipeline = [
        {"$match": floor_query(user, {"status": {"$in": ["open", "snoozed"]}})},
        {"$group": {"_id": "$rule_type", "count": {"$sum": 1}}},
    ]
    rows = await db.followups.aggregate(pipeline).to_list(30)
    return {r["_id"]: r["count"] for r in rows}


def _get(d: dict, key: str, default=None):
    return d.get(key, default)


def _encode_notebook_cursor(updated_at: str, row_id: str) -> str:
    raw = json.dumps({"updated_at": updated_at, "id": row_id}, separators=(",", ":")).encode()
    return base64.urlsafe_b64encode(raw).decode()


def _decode_notebook_cursor(cursor: str) -> dict:
    try:
        value = json.loads(base64.urlsafe_b64decode(cursor.encode()).decode())
        if not value.get("updated_at") or not value.get("id"):
            raise ValueError
        return value
    except (ValueError, TypeError, json.JSONDecodeError):
        raise HTTPException(status_code=400, detail="Invalid notebook cursor")


def _notebook_conflict(error: NotebookConflictError) -> HTTPException:
    return HTTPException(status_code=409, detail={
        "message": str(error), "row": error.row, "changed_fields": error.changed_fields,
    })


# ─────────────────────────────────────────────────────────────────────────────
# Automation
# ─────────────────────────────────────────────────────────────────────────────
@router.post("/reconcile")
async def reconcile(_: UserPublic = Depends(get_current_user)):
    return await reconcile_followups()


@router.get("/config/rules")
async def rules_config(user: UserPublic = Depends(get_current_user)):
    counts = await _rule_counts(user)
    return [{**r, "active_count": counts.get(r["rule_type"], 0)} for r in RULE_DEFINITIONS]


@router.get("/config/automation-rules")
async def get_automation_rules(user: UserPublic = Depends(get_current_user)):
    """Configurable reminder cadences (day thresholds) per workspace category
    — see services/automation_rules.py. Editable by manager+ only."""
    return await automation_rules.list_rules()


@router.put("/config/automation-rules/{category}")
async def put_automation_rule(category: str, body: AutomationRuleUpdate, user: UserPublic = Depends(require_min_role("manager"))):
    patch = body.dict(exclude_unset=True)
    updated = await automation_rules.update_rule(category, patch, user_id=user.id, user_name=user.full_name)
    asyncio.create_task(reconcile_followups())
    return updated


@router.get("/config/assignees")
async def assignees(_: UserPublic = Depends(get_current_user)):
    return await db.users.find(
        {"active": True}, {"_id": 0, "id": 1, "full_name": 1, "role": 1},
    ).sort("full_name", 1).to_list(100)


class WorkflowTransitionUpdate(BaseModel):
    title: Optional[str] = None
    message_template: Optional[str] = None
    priority: Optional[str] = None
    is_active: Optional[bool] = None


@router.get("/config/workflow-transitions")
async def get_workflow_transitions(_: UserPublic = Depends(get_current_user)):
    """Configurable business-event → operational-follow-up transitions (2026-08
    CRM foundation) — e.g. Quotation → Order Confirmed. See
    services/workflow_transitions.py / followup_engine._operational_followup_producer.
    Editable by manager+ only."""
    return await workflow_transitions.list_transitions()


@router.put("/config/workflow-transitions/{key}")
async def put_workflow_transition(key: str, body: WorkflowTransitionUpdate, user: UserPublic = Depends(require_min_role("manager"))):
    patch = body.dict(exclude_unset=True)
    updated = await workflow_transitions.update_transition(key, patch, user_id=user.id, user_name=user.full_name)
    asyncio.create_task(reconcile_followups())
    return updated


# ─────────────────────────────────────────────────────────────────────────────
# KPIs / Today's Mission / Insights — all literal paths, MUST precede /{id}
# ─────────────────────────────────────────────────────────────────────────────
@router.get("/stats")
async def stats(user: UserPublic = Depends(get_current_user)):
    await _wake_snoozed()
    docs = await _all_with_bucket(user)

    counts = {b: 0 for b in ("overdue", "today", "tomorrow", "this_week", "later", "completed", "snoozed")}
    overdue_critical = 0
    for d in docs:
        counts[d["bucket"]] = counts.get(d["bucket"], 0) + 1
        if d["bucket"] == "overdue" and d["effective_priority_level"] in ("critical", "high"):
            overdue_critical += 1

    today_critical = sum(1 for d in docs if d["bucket"] == "today" and d["effective_priority_level"] == "critical")
    waiting_for_customer = sum(
        1 for d in docs if d.get("status") == "open" and d.get("rule_type") == "payment_partial"
    )

    # Split "overdue" into payment-specific vs generic — fixes the "which
    # payment is overdue" 5-second-scan gap identified in the UX audit.
    overdue_payments_count = sum(1 for d in docs if d.get("status") == "open" and d.get("rule_type") == "payment_overdue")
    overdue_payments_amount = sum(
        d.get("value", 0) for d in docs if d.get("status") == "open" and d.get("rule_type") == "payment_overdue"
    )
    expiring_quotations_count = sum(
        1 for d in docs if d.get("status") == "open" and d.get("rule_type") == "quotation_expiring"
    )

    start, _end = ist_day_bounds_utc(0)
    y_start, y_end = ist_day_bounds_utc(-1)
    completed_today = sum(1 for d in docs if d.get("completed_at") and d["completed_at"] >= start.isoformat())
    completed_yesterday = sum(
        1 for d in docs if d.get("completed_at") and y_start.isoformat() <= d["completed_at"] < y_end.isoformat()
    )

    rules = await _rule_counts(user)
    workspace_counts = {
        "selection": rules.get("selection_waiting", 0),
        "quotation_tiles": rules.get("quotation_tiles_waiting", 0),
        "payment": rules.get("payment_overdue", 0) + rules.get("payment_partial", 0),
        "walk_in": 0,  # Phase 4 — Walk-ins module not yet built
    }
    return {
        "today_tasks": counts["today"], "today_critical": today_critical,
        "overdue": counts["overdue"], "overdue_critical": overdue_critical,
        "overdue_payments_count": overdue_payments_count,
        "overdue_payments_amount": round(overdue_payments_amount, 2),
        "overdue_payments_amount_short": money_short(overdue_payments_amount),
        "expiring_quotations_count": expiring_quotations_count,
        "tomorrow": counts["tomorrow"],
        "this_week": counts["this_week"],
        "waiting_for_customer": waiting_for_customer,
        "completed_today": completed_today,
        "completed_trend": completed_today - completed_yesterday,
        "snoozed": counts["snoozed"],
        "later": counts["later"],
        "rules": [{**r, "active_count": rules.get(r["rule_type"], 0)} for r in RULE_DEFINITIONS],
        "workspace_counts": workspace_counts,
    }


@router.get("/mission")
async def mission(user: UserPublic = Depends(get_current_user)):
    await _wake_snoozed()
    docs = await _all_with_bucket(user)
    actionable = [d for d in docs if d["bucket"] in ("overdue", "today")]

    revenue_at_risk = sum(d.get("value", 0) for d in actionable)
    overdue_payments = sum(1 for d in actionable if d["rule_type"] == "payment_overdue")
    expiring_today = sum(1 for d in docs if d["rule_type"] == "quotation_expiring" and d["bucket"] == "today")
    critical_count = sum(1 for d in actionable if d["effective_priority_level"] == "critical")

    minutes = 0
    for d in actionable:
        minutes += 6 if d["suggested_channel"] == "call" else 2 if d["suggested_channel"] == "whatsapp" else 3

    top = sorted(actionable, key=lambda d: -(d.get("priority_score") or 0))[:3]
    return {
        "due_count": len(actionable),
        "revenue_at_risk": round(revenue_at_risk, 2),
        "revenue_at_risk_short": money_short(revenue_at_risk),
        "overdue_payments": overdue_payments,
        "quotations_expiring_today": expiring_today,
        "critical_count": critical_count,
        "estimated_minutes": minutes,
        "top_priorities": [
            {"id": d["id"], "customer_name": d["customer_name"], "reason": d["reason"], "priority_score": d["priority_score"]}
            for d in top
        ],
        "greeting_name": (user.full_name or "").split()[0] if user.full_name else "there",
    }


@router.get("/insights")
async def insights(user: UserPublic = Depends(get_current_user)):
    start, end = ist_day_bounds_utc(0)
    rng = {"$gte": start.isoformat(), "$lt": end.isoformat()}

    calls = await db.activity_events.count_documents(
        floor_query(user, {"event_type": "followup.call_logged", "created_at": rng})
    )
    whatsapps = await db.activity_events.count_documents(floor_query(user, {
        "event_type": "followup.contacted", "payload.channel": "whatsapp", "created_at": rng,
    }))
    pay_docs = await db.payments.find(
        floor_query(user, {"paid_at": rng}), {"_id": 0, "amount": 1},
    ).to_list(1000)
    payments_collected = sum(p.get("amount", 0) for p in pay_docs)
    quotations_approved = await db.quotations.count_documents(floor_query(user, {
        "status": {"$in": ["approved", "won"]}, "updated_at": rng,
    }))
    completed_today = await db.followups.count_documents(floor_query(user, {"completed_at": rng}))
    still_open = await db.followups.count_documents(
        floor_query(user, {"status": {"$in": ["open", "snoozed"]}})
    )
    response_rate = round(100 * completed_today / max(1, completed_today + still_open))

    return {
        "calls_completed": calls,
        "whatsapps_sent": whatsapps,
        "payments_collected": round(payments_collected, 2),
        "quotations_approved": quotations_approved,
        "response_rate": min(100, response_rate),
    }


# ─────────────────────────────────────────────────────────────────────────────
# Export — CSV / Excel of the current filtered list. Saved Views — persisted
# filter configurations per user. Both literal-path routes; MUST precede /{id}.
# ─────────────────────────────────────────────────────────────────────────────
@router.get("/export")
async def export_followups(
    format: str = Query("xlsx", regex="^(xlsx|csv)$"),
    bucket: Optional[str] = None,
    priority: Optional[str] = None,
    category: Optional[str] = None,
    customer_tier: Optional[str] = None,
    assigned_to: Optional[str] = None,
    q: Optional[str] = None,
    user: UserPublic = Depends(get_current_user),
):
    rows = await list_followups(
        bucket=bucket, priority=priority, category=category, channel=None,
        customer_tier=customer_tier, assigned_to=assigned_to, q=q, limit=3000, user=user,
    )
    stamp = datetime.now(timezone.utc).strftime("%Y%m%d-%H%M")

    if format == "csv":
        buf = io.StringIO()
        writer = csv.writer(buf)
        writer.writerow(["Customer", "Phone", "Type", "Reason", "Next Action", "Value", "Priority", "Score", "Due", "Status", "Assigned To"])
        for d in rows:
            writer.writerow([
                d.get("customer_name"), d.get("customer_phone"), d.get("category"), d.get("reason"),
                d.get("next_action"), d.get("value"), d.get("effective_priority_level") or d.get("priority_level"),
                d.get("priority_score"), d.get("due_at"), d.get("status"), d.get("assigned_to_name"),
            ])
        mem = io.BytesIO(buf.getvalue().encode("utf-8"))
        return StreamingResponse(mem, media_type="text/csv", headers={
            "Content-Disposition": f'attachment; filename="followups-{stamp}.csv"',
        })

    wb = Workbook()
    ws = wb.active
    ws.title = "Follow-ups"
    ws["A1"] = "BuildCon House — Follow-ups Export"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:K1")
    ws["A2"] = f"{len(rows)} follow-ups · Exported {datetime.now(timezone.utc).strftime('%d %b %Y · %H:%M UTC')}"
    ws["A2"].font = Font(color="6B7280", size=10)
    ws.merge_cells("A2:K2")

    headers = ["Customer", "Phone", "Type", "Reason", "Next Action", "Value (₹)", "Priority", "Score", "Due", "Status", "Assigned To"]
    header_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = Font(bold=True, color="374151", size=11)
        cell.fill = header_fill

    for i, d in enumerate(rows, start=5):
        ws.cell(row=i, column=1, value=d.get("customer_name"))
        ws.cell(row=i, column=2, value=d.get("customer_phone"))
        ws.cell(row=i, column=3, value=d.get("category"))
        ws.cell(row=i, column=4, value=d.get("reason"))
        ws.cell(row=i, column=5, value=d.get("next_action"))
        ws.cell(row=i, column=6, value=d.get("value"))
        ws.cell(row=i, column=7, value=(d.get("effective_priority_level") or d.get("priority_level")))
        ws.cell(row=i, column=8, value=d.get("priority_score"))
        ws.cell(row=i, column=9, value=d.get("due_at"))
        ws.cell(row=i, column=10, value=d.get("status"))
        ws.cell(row=i, column=11, value=d.get("assigned_to_name"))

    for i, w in enumerate([22, 16, 12, 44, 22, 12, 10, 8, 22, 12, 18], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A5"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={
        "Content-Disposition": f'attachment; filename="followups-{stamp}.xlsx"',
    })


@router.get("/saved-views")
async def list_saved_views(user: UserPublic = Depends(get_current_user)):
    return await db.followup_saved_views.find({"user_id": user.id}, {"_id": 0}).sort("created_at", -1).to_list(50)


@router.post("/saved-views", response_model=FollowupSavedView)
async def create_saved_view(body: FollowupSavedViewCreate, user: UserPublic = Depends(get_current_user)):
    v = FollowupSavedView(user_id=user.id, name=body.name, filters=body.filters)
    await db.followup_saved_views.insert_one(v.dict())
    return v


@router.delete("/saved-views/{view_id}")
async def delete_saved_view(view_id: str, user: UserPublic = Depends(get_current_user)):
    await db.followup_saved_views.delete_one({"id": view_id, "user_id": user.id})
    return {"ok": True}


# ─────────────────────────────────────────────────────────────────────────────
# List
# ─────────────────────────────────────────────────────────────────────────────
@router.get("")
async def list_followups(
    bucket: Optional[str] = None,
    priority: Optional[str] = None,
    category: Optional[str] = None,
    channel: Optional[str] = None,
    customer_tier: Optional[str] = None,
    assigned_to: Optional[str] = None,
    q: Optional[str] = None,
    limit: int = Query(1000, ge=1, le=3000),
    user: UserPublic = Depends(get_current_user),
):
    await _wake_snoozed()
    query: dict = {}
    if category:
        query["category"] = category
    if channel:
        query["suggested_channel"] = channel
    if customer_tier:
        query["customer_tier"] = customer_tier
    if assigned_to:
        query["assigned_to"] = assigned_to
    if q:
        term = {"$regex": re.escape(q), "$options": "i"}
        query["$or"] = [
            {"customer_name": term}, {"customer_phone": term}, {"quotation_number": term},
            {"purchase_number": term}, {"project_name": term}, {"reason": term}, {"tags": term},
        ]

    docs = await db.followups.find(floor_query(user, query), {"_id": 0}).to_list(limit * 3)
    for d in docs:
        d["bucket"] = compute_bucket(d)
        d["effective_priority_level"] = d.get("manual_priority_override") or d.get("priority_level")

    if bucket and bucket != "all":
        docs = [d for d in docs if d["bucket"] == bucket]
    if priority:
        docs = [d for d in docs if d["effective_priority_level"] == priority]

    docs.sort(key=lambda d: _followup_sort_key(d, user.id))
    return docs[:limit]


# ─────────────────────────────────────────────────────────────────────────────
# Kitchen/Furniture digital notebook — same followups collection, narrow DTO.
# Literal notebook paths precede /{followup_id} below.
# ─────────────────────────────────────────────────────────────────────────────
@router.get("/notebook/{floor_id}")
async def list_notebook(
    floor_id: str,
    view: str = Query("followups", pattern="^(followups|quotation)$"),
    status: Optional[str] = Query(None, pattern="^(all|new|pending|won|lost)$"),
    q: Optional[str] = None,
    cursor: Optional[str] = None,
    limit: int = Query(100, ge=1, le=250),
    user: UserPublic = Depends(get_current_user),
):
    require_notebook_floor(floor_id, user)
    query = notebook_query(user, floor_id, {"is_converted": view == "quotation"})
    if status and status != "all":
        query["notebook_status"] = status
    search = notebook_search_query(q or "")
    if search:
        query["$or"] = search["$or"]
    if cursor:
        after = _decode_notebook_cursor(cursor)
        query = {"$and": [query, {"$or": [
            {"updated_at": {"$lt": after["updated_at"]}},
            {"updated_at": after["updated_at"], "id": {"$lt": after["id"]}},
        ]}]}
    projection = {
        "_id": 0, **({
            "id": 1, "customer_name": 1, "customer_phone": 1, "address": 1,
            "kitchen_type": 1, "referred_by": 1, "architect_interior_designer": 1,
            "notebook_status": 1, "notes": 1, "is_converted": 1, "updated_at": 1,
            "quotation_price": 1, "estimated_value": 1, "quotation_date": 1,
        }),
    }
    rows = await db.followups.find(query, projection).sort([("updated_at", -1), ("id", -1)]).limit(limit + 1).to_list(limit + 1)
    has_more = len(rows) > limit
    rows = rows[:limit]
    return {
        "rows": [serialize_notebook_row(row) for row in rows],
        "next_cursor": _encode_notebook_cursor(rows[-1]["updated_at"], rows[-1]["id"]) if has_more and rows else None,
    }


@router.post("/notebook/{floor_id}")
async def create_notebook_row(
    floor_id: str,
    body: NotebookFollowupCreatePayload,
    user: UserPublic = Depends(get_current_user),
):
    require_notebook_floor(floor_id, user)
    patch = body.dict(exclude_none=True)
    patch["notebook_status"] = "new"
    try:
        clean = validate_notebook_patch(
            patch, converted=False, current={}, creating=True, floor_id=floor_id,
        )
    except NotebookValidationError as error:
        raise HTTPException(status_code=422, detail=str(error))
    customer = await resolve_or_create_customer(
        db, user=user, floor_id=floor_id, name=clean["customer_name"],
        phone=clean["customer_phone"], address=clean.get("address"),
    )
    # A customer has one notebook row per floor. Repeated creation requests
    # must resolve to that row instead of making the register appear to ignore
    # edits or creating a second customer conversation.
    notebook_key = f"{floor_id}:{customer['id']}"
    existing = await db.followups.find_one(
        notebook_query(user, floor_id, {"notebook_key": notebook_key}), {"_id": 0},
    )
    if existing:
        return serialize_notebook_row(existing)
    now = now_iso()
    row = Followup(
        floor_id=floor_id, notebook_key=notebook_key, source_key=f"notebook:{notebook_key}",
        rule_type="manual", category="sales", customer_id=customer["id"],
        customer_name=clean["customer_name"], customer_phone=clean["customer_phone"],
        customer_tier=customer.get("tier", "retail"), reason="Notebook follow-up",
        next_action="Call customer", next_action_reason="Notebook customer follow-up",
        suggested_channel="call", priority_score=0, priority_level="medium", due_at=now,
        is_automated=False, is_converted=False, notebook_status="new",
        address=clean.get("address"), kitchen_type=clean.get("kitchen_type"),
        referred_by=clean.get("referred_by"),
        architect_interior_designer=clean.get("architect_interior_designer"),
        notes=clean.get("notes"),
    )
    try:
        await db.followups.insert_one(row.dict())
    except DuplicateKeyError:
        # The unique notebook_key index wins races between two staff members
        # creating the same floor/customer record at once.
        existing = await db.followups.find_one(
            notebook_query(user, floor_id, {"notebook_key": notebook_key}), {"_id": 0},
        )
        if existing:
            return serialize_notebook_row(existing)
        raise
    await log_event(
        event_type="project_followup.created", entity_type="followup", entity_id=row.id,
        actor=user, customer_id=row.customer_id, floor_id=floor_id,
        summary="Notebook follow-up created",
    )
    return serialize_notebook_row(row.dict())


@router.get("/notebook/{floor_id}/{row_id}")
async def get_notebook_row(floor_id: str, row_id: str, user: UserPublic = Depends(get_current_user)):
    require_notebook_floor(floor_id, user)
    row = await db.followups.find_one(notebook_query(user, floor_id, {"id": row_id}), {"_id": 0})
    if not row:
        raise HTTPException(status_code=404, detail="Notebook row not found")
    return serialize_notebook_row(row)


@router.patch("/notebook/{floor_id}/{row_id}")
async def patch_notebook(
    floor_id: str, row_id: str, body: NotebookCellPatchPayload,
    user: UserPublic = Depends(get_current_user),
):
    require_notebook_floor(floor_id, user)
    field = "notebook_status" if body.field == "status" else body.field
    if field == "customer_phone" and body.value is not None:
        value = normalize_mobile(str(body.value))
    else:
        value = body.value
    before = await db.followups.find_one(notebook_query(user, floor_id, {"id": row_id}), {"_id": 0})
    if not before:
        raise HTTPException(status_code=404, detail="Notebook row not found")
    try:
        row = await patch_notebook_row(
            db, user=user, floor_id=floor_id, row_id=row_id,
            patch={field: value}, expected_updated_at=body.updated_at,
        )
    except NotebookConflictError as error:
        raise _notebook_conflict(error)
    except NotebookValidationError as error:
        raise HTTPException(status_code=422, detail=str(error))
    event_type = "project_followup.status_changed" if field == "notebook_status" else "project_followup.edited"
    summary = f"{field.replace('_', ' ').title()} updated"
    if field == "notebook_status" and value == "won":
        event_type, summary = "project_followup.won", "Won"
    elif field == "notebook_status" and value == "lost":
        event_type, summary = "project_followup.lost", "Lost"
    await log_event(
        event_type=event_type, entity_type="followup", entity_id=row_id, actor=user,
        customer_id=before.get("customer_id"), floor_id=floor_id, payload={"field": field, "value": value}, summary=summary,
    )
    if field == "notebook_status" and value == "lost":
        await log_event(
            event_type="project_followup.lost_note", entity_type="followup", entity_id=row_id, actor=user,
            customer_id=before.get("customer_id"), floor_id=floor_id,
            payload={"note": row.get("notes")}, summary="Lost note recorded",
        )
    return row


@router.post("/notebook/{floor_id}/{row_id}/convert")
async def convert_notebook(
    floor_id: str, row_id: str, body: NotebookConversionPayload,
    user: UserPublic = Depends(get_current_user),
):
    require_notebook_floor(floor_id, user)
    before = await db.followups.find_one(notebook_query(user, floor_id, {"id": row_id}), {"_id": 0})
    if not before:
        raise HTTPException(status_code=404, detail="Notebook row not found")
    try:
        row = await convert_notebook_row(
            db, user=user, floor_id=floor_id, row_id=row_id,
            patch=body.dict(exclude={"updated_at"}, exclude_none=True), expected_updated_at=body.updated_at,
        )
    except NotebookConflictError as error:
        raise _notebook_conflict(error)
    except NotebookValidationError as error:
        raise HTTPException(status_code=422, detail=str(error))
    await log_event(
        event_type="project_followup.converted", entity_type="followup", entity_id=row_id, actor=user,
        customer_id=before.get("customer_id"), floor_id=floor_id,
        summary="Converted to quotation follow-up",
    )
    return row


@router.get("/notebook/{floor_id}/{row_id}/timeline")
async def notebook_timeline(floor_id: str, row_id: str, user: UserPublic = Depends(get_current_user)):
    require_notebook_floor(floor_id, user)
    row = await db.followups.find_one(notebook_query(user, floor_id, {"id": row_id}), {"_id": 0, "id": 1})
    if not row:
        raise HTTPException(status_code=404, detail="Notebook row not found")
    return await timeline_for(entity_type="followup", entity_id=row_id, limit=200, floor_ids=[floor_id])


_ASSIGNMENT_STATUS_RANK = {"open": 0, "snoozed": 1, "done": 2, "dismissed": 2}


def _assignment_row(f: dict) -> dict:
    return {
        "id": f["id"], "assigned_to": f.get("assigned_to"), "assigned_to_name": f.get("assigned_to_name"),
        "customer_name": f.get("customer_name"), "reason": f.get("reason"), "category": f.get("category"),
        "status": f.get("status"), "bucket": compute_bucket(f),
        "days_pending": age_days(parse_iso(f.get("created_at"))),
        "due_at": f.get("due_at"), "created_at": f.get("created_at"),
    }


def _assignment_sort_key(row: dict) -> tuple:
    return (_ASSIGNMENT_STATUS_RANK.get(row["status"], 3), -row["days_pending"])


@router.get("/assignments")
async def list_assignments(
    include_completed: bool = False,
    user: UserPublic = Depends(require_min_role("manager")),
):
    """Who has what assigned, how long it's been pending, and whether it's
    done — manager/admin/owner only. See
    docs/superpowers/specs/2026-07-27-followups-revamp-design.md."""
    status_filter = ["open", "snoozed"] + (["done", "dismissed"] if include_completed else [])
    docs = await db.followups.find(
        floor_query(user, {"assigned_to": {"$ne": None}, "status": {"$in": status_filter}}),
        {"_id": 0},
    ).to_list(5000)
    rows = [_assignment_row(f) for f in docs]
    rows.sort(key=_assignment_sort_key)
    return rows


# ─────────────────────────────────────────────────────────────────────────────
# Detail — powers the Customer Context Panel
# ─────────────────────────────────────────────────────────────────────────────
@router.get("/{followup_id}")
async def get_detail(followup_id: str, user: UserPublic = Depends(get_current_user)):
    from routes.payment_routes import ORDER_STATUSES, _paid_by_quotation

    f = await db.followups.find_one(floor_query(user, {"id": followup_id}), {"_id": 0})
    if not f:
        raise HTTPException(status_code=404, detail="Follow-up not found")
    f["bucket"] = compute_bucket(f)
    f["effective_priority_level"] = f.get("manual_priority_override") or f.get("priority_level")

    customer = await db.customers.find_one({"id": f["customer_id"]}, {"_id": 0, "password_hash": 0}) or {}
    all_q = await db.quotations.find({"customer_id": f["customer_id"]}, {"_id": 0}).sort("updated_at", -1).to_list(200)
    q_ids = [q["id"] for q in all_q]
    paid_map = await _paid_by_quotation(q_ids)

    lifetime_revenue = sum(q.get("grand_total", 0) for q in all_q if q.get("status") in ORDER_STATUSES)
    outstanding_total = sum(
        max(0.0, q.get("grand_total", 0) - paid_map.get(q["id"], 0.0)) for q in all_q if q.get("status") in ORDER_STATUSES
    )
    pending_quotations = [q for q in all_q if q.get("status") in ("draft", "sent", "pending_approval", "approved")]
    pending_orders = [
        q for q in all_q
        if q.get("status") in ORDER_STATUSES and (q.get("grand_total", 0) - paid_map.get(q["id"], 0.0)) > 1
    ]
    recent_payments = await db.payments.find({"customer_id": f["customer_id"]}, {"_id": 0}).sort("paid_at", -1).to_list(10)
    recent_purchases = await db.purchase_orders.find(
        {"customer_id": f["customer_id"]}, {"_id": 0},
    ).sort("updated_at", -1).to_list(10)
    timeline = await timeline_for(customer_id=f["customer_id"], limit=60)

    # ── Premium context additions (Follow-ups V2) — all derived from data
    # already loaded above, no new integration or LLM call needed. ──────────
    order_count = sum(1 for q in all_q if q.get("status") in ORDER_STATUSES)
    conversion_rate = round(100 * order_count / len(all_q)) if all_q else 0
    average_order_value = round(lifetime_revenue / order_count, 2) if order_count else 0.0
    creator_counts = Counter(q.get("created_by_name") for q in all_q if q.get("created_by_name"))
    preferred_salesperson = creator_counts.most_common(1)[0][0] if creator_counts else None

    last_touch_dt = None
    touches = [parse_iso(q.get("updated_at")) for q in all_q]
    touches = [t for t in touches if t]
    if touches:
        last_touch_dt = max(touches)
    days_silent = age_days(last_touch_dt) if last_touch_dt else 0
    has_overdue_payment = any(
        q.get("status") in ORDER_STATUSES and (q.get("grand_total", 0) - paid_map.get(q["id"], 0.0)) > 1
        and age_days(parse_iso(q.get("updated_at"))) >= 5
        for q in all_q
    )
    if has_overdue_payment:
        risk_level = "high"
    elif outstanding_total > 0 or days_silent >= 14:
        risk_level = "medium"
    else:
        risk_level = "low"

    return {
        "followup": f,
        "customer": customer,
        "stats": {
            "lifetime_revenue": round(lifetime_revenue, 2),
            "outstanding_total": round(outstanding_total, 2),
            "pending_quotations": len(pending_quotations),
            "pending_orders": len(pending_orders),
            "conversion_rate": conversion_rate,
            "average_order_value": average_order_value,
            "preferred_salesperson": preferred_salesperson,
            "risk_level": risk_level,
        },
        "quotations": [
            {"id": q["id"], "number": q["number"], "status": q["status"], "grand_total": q.get("grand_total", 0),
             "valid_until": q.get("valid_until"), "updated_at": q.get("updated_at")}
            for q in all_q[:12]
        ],
        "payments": recent_payments,
        "purchases": [
            {"id": p["id"], "number": p["number"], "status": p["status"], "grand_total": p.get("grand_total", 0),
             "updated_at": p.get("updated_at")}
            for p in recent_purchases
        ],
        "timeline": timeline,
    }


# ─────────────────────────────────────────────────────────────────────────────
# Manual create + edit
# ─────────────────────────────────────────────────────────────────────────────
@router.post("", response_model=Followup)
async def create_followup(body: FollowupCreate, user: UserPublic = Depends(get_current_user)):
    cust = await db.customers.find_one(floor_query(user, {"id": body.customer_id}), {"_id": 0})
    if not cust:
        raise HTTPException(status_code=404, detail="Customer not found")
    quotation = await db.quotations.find_one(floor_query(user, {"id": body.quotation_id}), {"_id": 0}) if body.quotation_id else None
    tier = cust.get("tier", "retail")
    value = float(quotation.get("grand_total", 0)) if quotation else 0.0
    score, level = score_followup(value, 0, 12, tier)

    assigned_to_name = user.full_name
    if body.assigned_to and body.assigned_to != user.id:
        au = await db.users.find_one({"id": body.assigned_to}, {"_id": 0, "full_name": 1})
        assigned_to_name = au["full_name"] if au else None

    action_label = {"call": "Call customer", "whatsapp": "Send WhatsApp", "email": "Send Email", "visit": "Schedule showroom visit"}
    f = Followup(
        source_key=None, rule_type="manual", category=body.category,
        customer_id=cust["id"], customer_name=cust.get("company") or cust.get("name"),
        customer_phone=cust.get("phone"), customer_tier=tier,
        quotation_id=quotation.get("id") if quotation else None,
        quotation_number=quotation.get("number") if quotation else None,
        purchase_id=body.purchase_id,
        value=value, reason=body.reason,
        reason_factors=reason_factors_for(value, 0, "Manual reminder", tier),
        next_action=action_label.get(body.channel, "Call customer"),
        next_action_reason=body.reason,
        suggested_channel=body.channel, priority_score=score,
        priority_level=body.priority_level or level,
        due_at=body.due_at or now_iso(), is_automated=False,
        assigned_to=body.assigned_to or user.id, assigned_to_name=assigned_to_name,
        notes=body.notes,
        floor_id=floor_for_write(user),
    )
    await db.followups.insert_one(f.dict())
    await log_event(
        event_type="followup.created", entity_type="followup", entity_id=f.id, actor=user,
        customer_id=f.customer_id, quotation_id=f.quotation_id,
        summary=f"Manual follow-up created — {f.reason}",
    )
    return f


@router.patch("/{followup_id}")
async def update_followup(followup_id: str, body: FollowupUpdate, user: UserPublic = Depends(get_current_user)):
    f = await db.followups.find_one(floor_query(user, {"id": followup_id}), {"_id": 0})
    if not f:
        raise HTTPException(status_code=404, detail="Follow-up not found")
    patch = body.dict(exclude_unset=True)
    if "assigned_to" in patch and patch["assigned_to"]:
        au = await db.users.find_one({"id": patch["assigned_to"]}, {"_id": 0, "full_name": 1})
        patch["assigned_to_name"] = au["full_name"] if au else None
        await log_event(
            event_type="followup.assigned", entity_type="followup", entity_id=followup_id, actor=user,
            customer_id=f.get("customer_id"),
            summary=f"Assigned to {patch.get('assigned_to_name') or '—'}",
        )
    if patch.get("status") == "dismissed":
        patch["completed_at"] = now_iso()
        patch["resolution_note"] = "Dismissed — not relevant"
        await log_event(
            event_type="followup.dismissed", entity_type="followup", entity_id=followup_id, actor=user,
            customer_id=f.get("customer_id"),
            summary=f"Follow-up dismissed — {f.get('reason')}",
        )
    if "notes" in patch and patch["notes"] and patch.get("status") != "dismissed" and "assigned_to" not in patch:
        await log_event(
            event_type="followup.note_added", entity_type="followup", entity_id=followup_id, actor=user,
            customer_id=f.get("customer_id"),
            summary=f"Note added: {patch['notes'][:120]}",
        )
    patch["updated_at"] = now_iso()
    await db.followups.update_one(floor_query(user, {"id": followup_id}), {"$set": patch})
    return await db.followups.find_one(floor_query(user, {"id": followup_id}), {"_id": 0})


@router.delete("/{followup_id}")
async def delete_followup(
    followup_id: str,
    user: UserPublic = Depends(require_min_role("manager")),
):
    """Permanently remove a follow-up during manager data cleanup."""
    f = await get_floor_scoped_or_404(
        db.followups, followup_id, user, not_found="Follow-up not found", projection={"_id": 0},
    )
    result = await db.followups.delete_one({"id": followup_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Follow-up not found")
    activity = await db.activity_events.delete_many({
        "$or": [
            {"entity_type": "followup", "entity_id": followup_id},
            {"followup_id": followup_id},
        ],
    })
    return {"ok": True, "followup_id": followup_id, "deleted": {"followups": result.deleted_count, "activity_events": activity.deleted_count}}


# ─────────────────────────────────────────────────────────────────────────────
# Actions — snooze / complete / contact / call outcome
# ─────────────────────────────────────────────────────────────────────────────
@router.post("/{followup_id}/snooze")
async def snooze_followup(followup_id: str, body: FollowupSnoozePayload, user: UserPublic = Depends(get_current_user)):
    f = await db.followups.find_one(floor_query(user, {"id": followup_id}), {"_id": 0})
    if not f:
        raise HTTPException(status_code=404, detail="Follow-up not found")
    now = datetime.now(timezone.utc)
    if body.until:
        until = parse_iso(body.until) or (now + timedelta(hours=1))
    elif body.minutes:
        until = now + timedelta(minutes=body.minutes)
    elif body.preset == "15m":
        until = now + timedelta(minutes=15)
    elif body.preset == "1h":
        until = now + timedelta(hours=1)
    elif body.preset == "tomorrow":
        until = (now + timedelta(days=1)).replace(hour=9, minute=0, second=0, microsecond=0)
    elif body.preset == "next_week":
        until = (now + timedelta(days=7)).replace(hour=9, minute=0, second=0, microsecond=0)
    else:
        until = now + timedelta(hours=1)

    await db.followups.update_one(floor_query(user, {"id": followup_id}), {"$set": {
        "status": "snoozed", "snoozed_until": until.isoformat(), "updated_at": now_iso(),
    }})
    await log_event(
        event_type="followup.snoozed", entity_type="followup", entity_id=followup_id, actor=user,
        customer_id=f.get("customer_id"),
        summary=f"Snoozed until {until.strftime('%d %b, %I:%M %p')}",
    )
    return await db.followups.find_one(floor_query(user, {"id": followup_id}), {"_id": 0})


@router.post("/{followup_id}/complete")
async def complete_followup(followup_id: str, body: FollowupCompletePayload, user: UserPublic = Depends(get_current_user)):
    f = await db.followups.find_one(floor_query(user, {"id": followup_id}), {"_id": 0})
    if not f:
        raise HTTPException(status_code=404, detail="Follow-up not found")
    await db.followups.update_one(floor_query(user, {"id": followup_id}), {"$set": {
        "status": "done", "completed_at": now_iso(),
        "notes": body.notes if body.notes is not None else f.get("notes"),
        "updated_at": now_iso(),
    }})
    await log_event(
        event_type="followup.completed", entity_type="followup", entity_id=followup_id, actor=user,
        customer_id=f.get("customer_id"), quotation_id=f.get("quotation_id"), purchase_id=f.get("purchase_id"),
        summary=f"Follow-up marked complete — {f.get('reason')}",
    )
    return await db.followups.find_one(floor_query(user, {"id": followup_id}), {"_id": 0})


@router.post("/{followup_id}/contact")
async def contact_followup(followup_id: str, body: FollowupContactPayload, user: UserPublic = Depends(get_current_user)):
    from routes.payment_routes import _clean_phone
    from services.email_service import build_email
    from services.messaging_service import build_message

    f = await db.followups.find_one(floor_query(user, {"id": followup_id}), {"_id": 0})
    if not f:
        raise HTTPException(status_code=404, detail="Follow-up not found")
    now = now_iso()
    await db.followups.update_one(floor_query(user, {"id": followup_id}), {"$set": {"last_contacted_at": now, "updated_at": now}})
    await log_event(
        event_type="followup.contacted", entity_type="followup", entity_id=followup_id, actor=user,
        customer_id=f.get("customer_id"), quotation_id=f.get("quotation_id"), purchase_id=f.get("purchase_id"),
        payload={"channel": body.channel},
        summary=f"{body.channel.title()} — {f.get('customer_name')}",
    )
    phone = _clean_phone(f.get("customer_phone"))
    result: dict = {"channel": body.channel, "phone": phone}
    # Category-aware message templates (services/messaging_service.py,
    # services/email_service.py) — provider-based, deep-link only in Phase 1.
    ctx = {
        "customer_name": (f.get("customer_name") or "there").split()[0],
        "quotation_number": f.get("quotation_number") or "",
        "outstanding_amount": money_short(f.get("value", 0)),
        "salesperson_name": f.get("assigned_to_name") or "",
        "reason": f.get("reason") or "Just checking in!",
    }
    if body.channel == "whatsapp":
        msg = build_message(f.get("category", "general"), phone, ctx)
        result["message"] = msg["message"]
        result["wa_url"] = msg["url"]
    elif body.channel == "email":
        cust = await db.customers.find_one({"id": f["customer_id"]}, {"_id": 0, "email": 1})
        result["email"] = cust.get("email") if cust else None
        template_key = "payment_reminder" if f.get("category") == "payment" else "quotation"
        email = build_email(template_key, result["email"], {**ctx, "invoice_number": f.get("quotation_number") or "", "order_number": f.get("quotation_number") or ""})
        result["subject"] = email["subject"]
        result["body"] = email["body"]
        result["mailto_url"] = email["mailto_url"]
    return result


@router.post("/{followup_id}/log-call")
async def log_call(followup_id: str, body: FollowupCallOutcomePayload, user: UserPublic = Depends(get_current_user)):
    f = await db.followups.find_one(floor_query(user, {"id": followup_id}), {"_id": 0})
    if not f:
        raise HTTPException(status_code=404, detail="Follow-up not found")
    now_dt = datetime.now(timezone.utc)
    patch: dict = {
        "last_contacted_at": now_dt.isoformat(), "updated_at": now_iso(),
        "notes": body.notes if body.notes is not None else f.get("notes"),
    }
    next_created = None

    if body.outcome == "lost" and not (body.notes or "").strip():
        raise HTTPException(status_code=422, detail="A reason is required when marking a client as lost")

    if body.outcome == "pending" and body.next_followup_at is None:
        raise HTTPException(status_code=422, detail="Choose the next follow-up date for a pending client")

    if body.outcome in ("interested", "call_back", "pending"):
        patch.update({"status": "done", "completed_at": now_iso(), "completed_outcome": body.outcome})
        if body.outcome == "pending":
            due = body.next_followup_at
            if due.tzinfo is None:
                due = due.replace(tzinfo=timezone.utc)
            else:
                due = due.astimezone(timezone.utc)
            if due <= now_dt:
                raise HTTPException(status_code=422, detail="The next follow-up date must be in the future")
        else:
            due = now_dt + timedelta(days=1 if body.outcome == "call_back" else 2)
        next_reason = {
            "call_back": "Call back requested",
            "interested": "Customer interested — follow up on their decision",
            "pending": "Customer asked to reconnect later",
        }[body.outcome]
        nf = Followup(
            rule_type="manual", category=f.get("category", "general"),
            customer_id=f["customer_id"], customer_name=f["customer_name"], customer_phone=f.get("customer_phone"),
            customer_tier=f.get("customer_tier", "retail"), quotation_id=f.get("quotation_id"),
            quotation_number=f.get("quotation_number"), purchase_id=f.get("purchase_id"),
            purchase_number=f.get("purchase_number"), value=f.get("value", 0),
            reason=next_reason,
            reason_factors=[f.get("reason", "")] if f.get("reason") else [],
            next_action="Call customer", next_action_reason="Scheduled automatically from the previous call outcome.",
            suggested_channel="call", priority_score=f.get("priority_score", 50), priority_level=f.get("priority_level", "medium"),
            due_at=due.isoformat(), is_automated=False,
            assigned_to=f.get("assigned_to") or user.id, assigned_to_name=f.get("assigned_to_name") or user.full_name,
            tags=f.get("tags", []), notes=body.notes,
            floor_id=floor_inherit(f),
        )
        await db.followups.insert_one(nf.dict())
        next_created = nf.id
    elif body.outcome == "no_answer":
        attempts = (f.get("contact_attempts") or 0) + 1
        patch["contact_attempts"] = attempts
        if attempts >= 2:
            # Escalate — stop same-day retries after the 2nd miss; push to
            # tomorrow morning and bump urgency so it doesn't get buried.
            next_due = (now_dt + timedelta(days=1)).replace(hour=9, minute=30, second=0, microsecond=0)
            bumped_score = min(100, (f.get("priority_score") or 0) + 10)
            bumped_level = "critical" if bumped_score >= 80 else "high" if bumped_score >= 60 else f.get("priority_level", "medium")
            patch["due_at"] = next_due.isoformat()
            patch["priority_score"] = bumped_score
            patch["priority_level"] = bumped_level
        else:
            patch["due_at"] = (now_dt + timedelta(hours=4)).isoformat()
    elif body.outcome == "rejected":
        patch.update({
            "status": "dismissed", "completed_at": now_iso(),
            "completed_outcome": "rejected", "resolution_note": "Customer rejected",
        })
    elif body.outcome == "converted":
        patch.update({
            "status": "done", "completed_at": now_iso(),
            "completed_outcome": "converted", "resolution_note": "Converted!",
        })
    elif body.outcome == "won":
        patch.update({
            "status": "done", "completed_at": now_iso(),
            "completed_outcome": "won", "resolution_note": "Client won",
        })
    elif body.outcome == "lost":
        patch.update({
            "status": "dismissed", "completed_at": now_iso(),
            "completed_outcome": "lost", "resolution_note": body.notes.strip(),
        })

    await db.followups.update_one(floor_query(user, {"id": followup_id}), {"$set": patch})
    await log_event(
        event_type="followup.call_logged", entity_type="followup", entity_id=followup_id, actor=user,
        customer_id=f.get("customer_id"), quotation_id=f.get("quotation_id"), purchase_id=f.get("purchase_id"),
        payload={"outcome": body.outcome, "next_followup_id": next_created},
        summary=f"Call logged — {body.outcome.replace('_', ' ').title()}",
    )
    return await db.followups.find_one(floor_query(user, {"id": followup_id}), {"_id": 0})
