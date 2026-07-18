"""Product / Brand / Category endpoints + AI-assisted catalog import stub."""
import asyncio
from datetime import datetime, timezone
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Query

from auth import floor_for_write, floor_scope_ids, get_current_user, require_min_role
from db import db, strip_ids
from models import Product, ProductCreate, ProductPatch, UserPublic
from services import catalog_service, media_service
from services.activity_log import log_event

router = APIRouter(tags=["catalog"])


# ---------- Brands ----------
@router.get("/brands")
async def list_brands(user: UserPublic = Depends(get_current_user)):
    """Return every brand + its active product count, scoped to the
    caller's floor(s). Counts drive the left-rail brand badges in the
    Quotation Builder V4."""
    return await catalog_service.list_brands_with_counts(floor_ids=floor_scope_ids(user))


@router.get("/categories")
async def list_categories(
    brand_id: Optional[str] = None,
    user: UserPublic = Depends(get_current_user),
):
    """Return categories + per-brand-scoped product counts, scoped to the
    caller's floor(s).

    When `brand_id` is passed, counts reflect ONLY that brand — this is what
    powers the left-rail "Categories under Hansgrohe" list.
    """
    return await catalog_service.list_categories_with_counts(brand_id, floor_ids=floor_scope_ids(user))




async def _usage_ranked_product_page(
    *,
    query: dict,
    ranked_ids: set[str],
    rank_key,
    skip: int,
    limit: int,
) -> list[dict]:
    """Return a deterministic page with sparse usage-ranked products first.

    Usage exists for only a small fraction of the catalog. The previous path
    downloaded all 2,966 matching {id,name} rows on every page and sorted them
    in Python. This path fetches only matching usage IDs, then lets Mongo page
    the non-usage remainder by the indexed/stable (name,id) order.
    """
    ranked_pool: list[dict] = []
    if ranked_ids:
        ranked_pool = await db.products.find(
            {**query, "id": {"$in": list(ranked_ids)}},
            {"_id": 0, "id": 1, "name": 1},
        ).to_list(len(ranked_ids))
        ranked_pool.sort(key=rank_key)

    ranked_count = len(ranked_pool)
    selected_ranked = ranked_pool[skip:min(skip + limit, ranked_count)] if skip < ranked_count else []
    regular_skip = max(0, skip - ranked_count)
    regular_needed = limit - len(selected_ranked)

    async def _ranked_full() -> list[dict]:
        ids = [d["id"] for d in selected_ranked]
        if not ids:
            return []
        rows = await db.products.find({"id": {"$in": ids}}, {"_id": 0}).to_list(len(ids))
        by_id = {row["id"]: row for row in rows}
        return [by_id[pid] for pid in ids if pid in by_id]

    async def _regular_full() -> list[dict]:
        if regular_needed <= 0:
            return []
        regular_query = query if not ranked_ids else {**query, "id": {"$nin": list(ranked_ids)}}
        return await db.products.find(regular_query, {"_id": 0}).sort([
            ("name", 1), ("id", 1),
        ]).skip(regular_skip).limit(regular_needed).to_list(regular_needed)

    ranked_full, regular_full = await asyncio.gather(_ranked_full(), _regular_full())
    return [*ranked_full, *regular_full]

# ---------- Products ----------
@router.get("/products")
async def list_products(
    q: Optional[str] = Query(None, description="Free text search on name/sku/description/series/family/finish/colour/tags"),
    brand_id: Optional[str] = None,
    category_id: Optional[str] = None,
    subcategory: Optional[str] = None,
    series: Optional[str] = None,
    family_key: Optional[str] = None,
    finish: Optional[str] = None,
    colour: Optional[str] = None,
    sort: str = Query("popular", description="popular | recent | price_asc | price_desc | name"),
    limit: int = 60,
    skip: int = 0,
    user: UserPublic = Depends(get_current_user),
):
    return await catalog_service.list_products_page(
        user_id=user.id,
        q=q,
        brand_id=brand_id,
        category_id=category_id,
        subcategory=subcategory,
        series=series,
        family_key=family_key,
        finish=finish,
        colour=colour,
        sort=sort,
        limit=limit,
        skip=skip,
        floor_ids=floor_scope_ids(user),
    )


# ---------- Hierarchy + family-grouped views ----------
@router.get("/catalog/hierarchy")
async def catalog_hierarchy(user: UserPublic = Depends(get_current_user)):
    """Return the full Brand → Category → Subcategory → Series → Family tree,
    scoped to the caller's floor(s). Only counts active products.
    """
    pipeline = [
        {"$match": {"active": True}},
        {"$group": {
            "_id": {
                "brand_id":    "$brand_id",
                "category_id": "$category_id",
                "subcategory": "$subcategory",
                "series":      "$series",
                "family_key":  "$family_key",
                "family_name": "$family_name",
            },
            "product_count": {"$sum": 1},
            "min_price":     {"$min": "$price"},
            "sample_image":  {"$first": {"$arrayElemAt": ["$images", 0]}},
            "image_quality": {"$first": "$image_quality"},
        }},
    ]
    rows, brands, cats = await catalog_service.hierarchy_rows(floor_ids=floor_scope_ids(user))

    # Fold into brand → category → subcategory → series → family tree
    tree: dict = {}
    for r in rows:
        k = r["_id"]
        bid = k.get("brand_id") or "_"
        cid = k.get("category_id") or "_"
        subcat = k.get("subcategory") or "General"
        series = k.get("series") or "Uncategorised"
        family_key = k.get("family_key") or f"{bid}:{cid}:{subcat}:{series}:_"
        family_name = k.get("family_name") or series

        brand = brands.get(bid, {"id": bid, "name": "Unknown"})
        category = cats.get(cid, {"id": cid, "name": "Uncategorised"})
        b = tree.setdefault(bid, {"brand": brand, "categories": {}})
        c = b["categories"].setdefault(cid, {"category": category, "subcategories": {}})
        s = c["subcategories"].setdefault(subcat, {"name": subcat, "series": {}})
        se = s["series"].setdefault(series, {"name": series, "families": {}})
        fam = se["families"].setdefault(family_key, {
            "family_key": family_key, "family_name": family_name,
            "product_count": 0, "min_price": 0.0,
            "sample_image": None, "image_quality": None,
        })
        fam["product_count"] += r["product_count"]
        if fam["min_price"] == 0 or r["min_price"] < fam["min_price"]:
            fam["min_price"] = r["min_price"]
        if not fam["sample_image"] and r.get("sample_image"):
            fam["sample_image"] = r["sample_image"]
        if not fam["image_quality"]:
            fam["image_quality"] = r.get("image_quality")

    # Flatten to arrays for JSON friendliness
    def _flatten(d: dict) -> list:
        out = []
        for bid, b in d.items():
            cats_out = []
            for cid, c in b["categories"].items():
                subs_out = []
                for subname, s in c["subcategories"].items():
                    ser_out = []
                    for sername, se in s["series"].items():
                        fams_out = list(se["families"].values())
                        fams_out.sort(key=lambda f: f["family_name"] or "")
                        ser_out.append({
                            "name": sername, "family_count": len(fams_out),
                            "product_count": sum(f["product_count"] for f in fams_out),
                            "families": fams_out,
                        })
                    ser_out.sort(key=lambda x: x["name"])
                    subs_out.append({
                        "name": subname, "series_count": len(ser_out),
                        "product_count": sum(sr["product_count"] for sr in ser_out),
                        "series": ser_out,
                    })
                subs_out.sort(key=lambda x: x["name"])
                cats_out.append({
                    "category": c["category"],
                    "subcategory_count": len(subs_out),
                    "product_count": sum(sub["product_count"] for sub in subs_out),
                    "subcategories": subs_out,
                })
            cats_out.sort(key=lambda x: x["category"].get("name", ""))
            out.append({
                "brand": b["brand"],
                "category_count": len(cats_out),
                "product_count": sum(c["product_count"] for c in cats_out),
                "categories": cats_out,
            })
        out.sort(key=lambda x: x["brand"].get("name", ""))
        return out

    return {"tree": _flatten(tree)}


@router.get("/products/families")
async def list_families(
    brand_id: Optional[str] = None,
    category_id: Optional[str] = None,
    subcategory: Optional[str] = None,
    series: Optional[str] = None,
    q: Optional[str] = None,
    limit: int = 60,
    skip: int = 0,
    user: UserPublic = Depends(get_current_user),
):
    """Return products grouped by family_key — one card per family, variants
    collapsed underneath. Ideal for the premium grouped catalog view.
    """
    return await catalog_service.list_family_groups(
        brand_id=brand_id,
        category_id=category_id,
        subcategory=subcategory,
        series=series,
        q=q,
        limit=limit,
        skip=skip,
        floor_ids=floor_scope_ids(user),
    )


# ---------- Ranked search (Iteration 2A) ----------
@router.get("/catalog/search")
async def catalog_search(
    q: str = Query("", description="Free-text query"),
    brand_id: Optional[str] = None,
    category_id: Optional[str] = None,
    subcategory: Optional[str] = None,
    series: Optional[str] = None,
    limit: int = 30,
    group: bool = Query(True, description="Group variants by family_key (Shopify-style)"),
    user: UserPublic = Depends(get_current_user),
):
    """Ranked catalog search, scoped to the caller's floor(s).

    Ranking priority (highest first):
      1. Exact SKU / SKU prefix
      2. Product name / family name (Mongo text score)
      3. Series / subcategory / finish / colour matches
      4. Fallback dimension / description matches

    Results are grouped by `family_key` by default so callers don't see 6
    duplicates of the same product with different finishes.
    """
    return await catalog_service.search_catalog(
        q=q,
        brand_id=brand_id,
        category_id=category_id,
        subcategory=subcategory,
        series=series,
        limit=limit,
        group=group,
        floor_ids=floor_scope_ids(user),
    )


@router.get("/catalog/facets")
async def catalog_facets(
    brand_id: Optional[str] = None,
    category_id: Optional[str] = None,
    subcategory: Optional[str] = None,
    series: Optional[str] = None,
    user: UserPublic = Depends(get_current_user),
):
    """Return the facet buckets (brands, categories, finishes, colours,
    price range) for the current selection, scoped to the caller's floor(s).
    """
    return await catalog_service.facet_buckets(
        brand_id=brand_id,
        category_id=category_id,
        subcategory=subcategory,
        series=series,
        floor_ids=floor_scope_ids(user),
    )


# ---------- Family-first canonical page (Iteration 2A shell — used by 2B) ----------
@router.get("/families/{family_key}")
async def get_family(family_key: str, _: UserPublic = Depends(get_current_user)):
    """Return everything the Shopify-style family page needs in a single
    call: family metadata + variants + gallery + specs union. Missing data
    is honestly reported (nulls / empty arrays), never fabricated.
    """
    doc = await catalog_service.family_detail(family_key)
    if not doc:
        raise HTTPException(status_code=404, detail="Family not found")
    return doc


@router.get("/products/recent")
async def recent_products(
    limit: int = 12,
    user: UserPublic = Depends(get_current_user),
):
    """Products this user has recently added to a quotation."""
    return await catalog_service.recent_or_frequent_products(
        user.id, limit=limit, recent=True,
    )


@router.get("/products/frequent")
async def frequent_products(
    limit: int = 12,
    user: UserPublic = Depends(get_current_user),
):
    """Products this user adds most often."""
    return await catalog_service.recent_or_frequent_products(
        user.id, limit=limit, recent=False,
    )


@router.get("/products/{product_id}")
async def get_product(product_id: str, _: UserPublic = Depends(get_current_user)):
    doc = await catalog_service.product_by_id(product_id)
    if not doc:
        raise HTTPException(status_code=404, detail="Product not found")
    return doc


@router.get("/products/{product_id}/alternates")
async def product_alternates(
    product_id: str,
    limit: int = 12,
    user: UserPublic = Depends(get_current_user),
):
    """Return alternate products a salesperson might swap in for `product_id`.

    Smart-mix ranking (closest matches first):
      Tier 1  same brand + same category + same name-prefix  (approximates family)
      Tier 2  same brand + same category
      Tier 3  same category (any brand)

    Within every tier we rank by (this user's usage count DESC, price ASC) so
    the salesperson sees products they actually reach for. The current product
    itself is always excluded.
    """
    result = await catalog_service.alternate_products(product_id, user.id, limit)
    if result is None:
        raise HTTPException(status_code=404, detail="Product not found")
    return result


@router.get("/products/{product_id}/complete-the-set")
async def complete_the_set(
    product_id: str,
    limit: int = 12,
    _: UserPublic = Depends(get_current_user),
):
    """"Complete the set" — same family (Series/Collection) but different
    category. E.g. viewing a Talis E basin mixer suggests the Talis E shower
    valve, spout, robe hook — the classic bathroom cross-sell.
    """
    result = await catalog_service.complete_set_products(product_id, limit)
    if result is None:
        raise HTTPException(status_code=404, detail="Product not found")
    return result


@router.post("/products/custom", response_model=Product)
async def create_custom_product(
    body: ProductCreate,
    user: UserPublic = Depends(require_min_role("sales")),
):
    """Create a custom / one-off product from the Quotation Builder.

    When body.is_custom=True, the SKU can collide with existing rows (we
    auto-suffix). If False, we behave like /products with duplicate-SKU
    rejection.
    """
    sku = body.sku or f"CUSTOM-{datetime.now(timezone.utc).strftime('%y%m%d%H%M%S')}"
    if body.is_custom:
        # Auto-uniquify — never fail because the user typed the same SKU twice.
        base = sku
        n = 1
        while await db.products.find_one({"sku": sku}):
            n += 1
            sku = f"{base}-{n}"
    elif await db.products.find_one({"sku": sku}):
        raise HTTPException(status_code=409, detail="SKU already exists")

    payload = body.dict()
    payload["sku"] = sku
    payload["tags"] = list(set([*(payload.get("tags") or []), "custom"]))
    payload["floor_id"] = floor_for_write(user)
    prod = Product(**payload)
    await db.products.insert_one(prod.dict())
    catalog_service.schedule_catalog_refresh()
    return prod


@router.post("/products", response_model=Product)
async def create_product(
    body: ProductCreate,
    user: UserPublic = Depends(require_min_role("purchase")),
):
    if await db.products.find_one({"sku": body.sku}):
        raise HTTPException(status_code=409, detail="SKU already exists")
    payload = body.dict()
    payload["floor_id"] = floor_for_write(user)
    prod = Product(**payload)
    await db.products.insert_one(prod.dict())
    catalog_service.schedule_catalog_refresh()
    return prod


@router.patch("/products/{product_id}", response_model=Product)
async def update_product(
    product_id: str,
    body: ProductPatch,
    user: UserPublic = Depends(require_min_role("purchase")),
):
    """The single write path for the "single source of truth" product
    editor — used identically from Catalog, the Quotation Builder's product
    sheet, and (per the same shared component) Purchases in future. Only
    fields present in the request are touched (exclude_unset); everything
    else on the product document is left exactly as-is.

    Quotations already snapshot `name`/`sku`/`unit_price`/`finish`/`colour`/
    `image` onto each line item at the moment a product is added (see
    QuotationLineItem + the /quotations/{id}/items endpoint) — so this
    endpoint intentionally never touches `db.quotations`. Editing a product
    here changes what FUTURE quotations copy in; every quotation already
    generated (and its PDF) is unaffected by construction, not by any special
    case here.
    """
    existing = await db.products.find_one({"id": product_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Product not found")

    updates = body.dict(exclude_unset=True)
    if not updates:
        return Product(**existing)

    new_sku = updates.get("sku")
    if new_sku and new_sku != existing.get("sku"):
        if await db.products.find_one({"sku": new_sku, "id": {"$ne": product_id}}):
            raise HTTPException(status_code=409, detail="SKU already exists")

    updates["updated_at"] = datetime.now(timezone.utc).isoformat()
    await db.products.update_one({"id": product_id}, {"$set": updates})

    catalog_service.patch_product_in_snapshot(product_id, updates)

    await log_event(
        event_type="product.updated",
        entity_type="product",
        entity_id=product_id,
        actor=user,
        payload={"fields": sorted(updates.keys())},
        summary=f"Updated {', '.join(sorted(k for k in updates if k != 'updated_at'))}",
    )

    updated = await catalog_service.product_by_id(product_id)
    return Product(**{k: v for k, v in updated.items() if k in Product.__fields__})


# ---------- Catalog import (AI-assisted scaffold) ----------
@router.get("/catalog/imports")
async def list_import_jobs(_: UserPublic = Depends(require_min_role("purchase"))):
    docs = await db.catalog_imports.find({}, {"_id": 0, "rows": 0}).sort("created_at", -1).to_list(200)
    return docs


# ---------- Catalog export (Settings > Catalog > Export) ----------
@router.get("/catalog/export.xlsx")
async def export_catalog_xlsx(_: UserPublic = Depends(get_current_user)):
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    products = await db.products.find({}, {"_id": 0}).sort("name", 1).to_list(10000)
    brands = {b["id"]: b["name"] for b in await db.brands.find({}, {"_id": 0, "id": 1, "name": 1}).to_list(500)}
    categories = {c["id"]: c["name"] for c in await db.categories.find({}, {"_id": 0, "id": 1, "name": 1}).to_list(500)}

    wb = Workbook()
    ws = wb.active
    ws.title = "Catalog"
    ws["A1"] = "Forge — Catalog Export"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:J1")
    stamp = datetime.now(timezone.utc).strftime("%d %b %Y · %H:%M UTC")
    ws["A2"] = f"{len(products)} products · Exported {stamp}"
    ws["A2"].font = Font(color="6B7280", size=10)
    ws.merge_cells("A2:J2")

    headers = ["SKU", "Name", "Brand", "Category", "Series", "Finish", "MRP", "Trade Price", "Stock", "Warranty"]
    header_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = Font(bold=True, color="374151", size=11)
        cell.fill = header_fill

    for i, p in enumerate(products, start=5):
        ws.cell(row=i, column=1, value=p.get("sku"))
        ws.cell(row=i, column=2, value=p.get("name"))
        ws.cell(row=i, column=3, value=brands.get(p.get("brand_id"), ""))
        ws.cell(row=i, column=4, value=categories.get(p.get("category_id"), ""))
        ws.cell(row=i, column=5, value=p.get("series"))
        ws.cell(row=i, column=6, value=p.get("finish"))
        ws.cell(row=i, column=7, value=p.get("mrp"))
        ws.cell(row=i, column=8, value=p.get("price"))
        ws.cell(row=i, column=9, value=p.get("stock"))
        ws.cell(row=i, column=10, value=p.get("warranty"))

    widths = [16, 40, 16, 18, 16, 16, 10, 12, 8, 16]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A5"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    from fastapi.responses import StreamingResponse
    filename = f"forge-catalog-{datetime.now(timezone.utc).strftime('%Y%m%d-%H%M')}.xlsx"
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
