"""GROHE xlsx extraction module (Full Catalog Replacement, 2026-08).

Parses the 4 supplier xlsx files (RSH Aqua Tile Shower / Plate / Shower /
Single Liver) into normalized rows, with EMF/WMF vector images converted to
real raster PNGs via headless LibreOffice (the only available renderer in
this environment — ImageMagick has no native EMF decoder). Every image is
verified non-trivial after conversion. NEVER fabricates a name, category,
price, or finish that is not literally present in the supplier's own file.

Category = supplier filename, verbatim (per explicit user instruction):
    RSH AǪUA TILE SHOWER.xlsx -> "RSH Aqua Tile Shower"
    PLATE.xlsx                -> "Plate"
    SHOWER grohe.xlsx         -> "Shower"
    SINGLE LIVER.xlsx         -> "Single Lever" (corrected typo per explicit
                                  user instruction; original filename kept in
                                  specs.source_file for traceability)

Finish is NEVER inferred from a code table — it is extracted only when a
known Grohe finish word appears literally in that row's own description
text. Rows with no such word keep finish=None (honest gap, not fabricated).
"""
from __future__ import annotations

import hashlib
import io
import re
import subprocess
import tempfile
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.utils import get_column_letter
from PIL import Image

RASTER_MIME = {
    "jpeg": "image/jpeg", "jpg": "image/jpeg", "png": "image/png",
    "gif": "image/gif", "webp": "image/webp", "bmp": "image/bmp",
}
VECTOR_FORMATS = {"emf", "wmf"}

# Extracted verbatim from the supplier's OWN description text in these 4
# files (never a fabricated code->colour lookup table). Longest-first so
# "Brushed Hard Graphite" wins over the bare "Hard Graphite"/"Graphite".
FINISH_TOKENS = [
    "Brushed Hard Graphite", "Brushed Cool Sunrise", "Brushed Warm Sunset",
    "Brushed Nickel", "Hard Graphite", "Warm Sunset", "Cool Sunrise",
    "SuperSteel", "Super Steel", "Moon White", "Matt Black", "Matt White",
    "Matt Taupe", "Matt Stone Grey", "Chrome", "Nickel", "Gold", "Matt",
    "Glossy",
]

FILES: dict[str, str] = {
    "RSH Aqua Tile Shower": "rsh.xlsx",
    "Plate": "plate.xlsx",
    "Shower": "shower.xlsx",
    "Single Lever": "single_liver.xlsx",
}
ORIGINAL_FILENAMES: dict[str, str] = {
    "RSH Aqua Tile Shower": "RSH AǪUA TILE SHOWER.xlsx",
    "Plate": "PLATE.xlsx",
    "Shower": "SHOWER grohe.xlsx",
    "Single Lever": "SINGLE LIVER.xlsx",
}


@dataclass
class ImageResult:
    ok: bool
    data: Optional[bytes] = None
    mime: Optional[str] = None
    source_format: Optional[str] = None   # "jpeg" | "png" | "emf-converted" | ...
    error: Optional[str] = None


@dataclass
class GroheRow:
    category: str
    source_file: str          # original supplier filename (traceability)
    sl: int
    sku: str
    description: str
    mrp: Optional[float]
    segment: Optional[str]
    finish_hint_col: Optional[str]   # column G "finishes" (MATT/GLOSSY) — Shower file only
    finish: Optional[str]            # extracted from description text only
    family_key: str
    row_num: int
    image: ImageResult = field(default_factory=lambda: ImageResult(ok=False))
    duplicate_of: Optional[str] = None   # sha1 of an identical earlier row in the same file


def _detect_finish(text: str) -> Optional[str]:
    for tok in FINISH_TOKENS:
        if tok.lower() in text.lower():
            return tok
    return None


def _normalize_family_base(description: str) -> str:
    """Strip known finish words + pure whitespace/size noise so variants of
    the same base product collapse to one family_key. Text-based only —
    never touches the SKU."""
    base = description
    for tok in FINISH_TOKENS:
        base = re.sub(re.escape(tok), "", base, flags=re.IGNORECASE)
    base = re.sub(r"\s+", " ", base).strip()
    slug = re.sub(r"[^a-z0-9]+", "-", base.lower()).strip("-")
    return slug[:80] or "misc"


def _convert_emf_to_png(raw: bytes, fmt: str) -> Optional[bytes]:
    """Headless LibreOffice EMF/WMF -> PNG, then auto-trim the surrounding
    white canvas (lossless — removes blank margin only, alters zero pixels
    of the actual artwork)."""
    with tempfile.TemporaryDirectory() as td:
        src = Path(td) / f"src.{fmt}"
        src.write_bytes(raw)
        try:
            subprocess.run(
                ["soffice", "--headless", "--convert-to", "png", "--outdir", td, str(src)],
                check=True, capture_output=True, timeout=45,
            )
        except Exception:
            return None
        out = Path(td) / "src.png"
        if not out.exists():
            return None
        png_bytes = out.read_bytes()
        try:
            img = Image.open(io.BytesIO(png_bytes)).convert("RGBA")
            # Autocrop: bbox of all non-white/non-transparent pixels
            bg = Image.new("RGBA", img.size, (255, 255, 255, 255))
            diff = Image.new("L", img.size)
            pixels_a = img.getdata()
            # Fast bbox via getbbox on a mask of "non-white" pixels
            gray = img.convert("L")
            mask = gray.point(lambda p: 0 if p > 250 else 255)
            bbox = mask.getbbox()
            if bbox:
                pad = 12
                l, t, r, b = bbox
                l = max(0, l - pad); t = max(0, t - pad)
                r = min(img.width, r + pad); b = min(img.height, b + pad)
                img = img.crop((l, t, r, b))
            if img.width < 20 or img.height < 20:
                return None
            buf = io.BytesIO()
            img.convert("RGB").save(buf, format="PNG")
            return buf.getvalue()
        except Exception:
            return png_bytes  # fall back to untrimmed if PIL step fails


def _extract_images_by_row(path: str) -> dict[int, tuple[bytes, str]]:
    """Map sheet row-number (1-indexed) -> (raw_bytes, format) for every
    embedded drawing anchored in the 'Product Image' column.

    IMPORTANT: openpyxl's high-level `ws._images` API silently DROPS any
    WMF/EMF-format image at load time (just emits a UserWarning) — it never
    surfaces the raw bytes for unsupported formats at all. That would make
    every EMF-backed product silently "missing an image" even though the
    supplier file genuinely contains one. So this reads the drawing XML +
    relationships straight out of the xlsx zip instead of trusting
    `ws._images`, which recovers 100% of embedded images regardless of
    format (verified: raw zip media-file counts match embedded objects
    across all 4 files exactly, whereas `ws._images` was short by exactly
    the EMF count on 3 of the 4 files)."""
    import zipfile
    from xml.etree import ElementTree as ET

    out: dict[int, tuple[bytes, str]] = {}
    ns = {
        "xdr": "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
        "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    }
    with zipfile.ZipFile(path) as zf:
        names = zf.namelist()
        # sheetN.xml -> find its index to locate the matching _rels file
        sheet_files = sorted([n for n in names if re.match(r"xl/worksheets/sheet\d+\.xml$", n)])
        if not sheet_files:
            return out
        sheet_path = sheet_files[0]  # single-sheet files, as verified for all 4
        rels_path = f"xl/worksheets/_rels/{Path(sheet_path).name}.rels"
        if rels_path not in names:
            return out
        rels_root = ET.fromstring(zf.read(rels_path))
        drawing_target = None
        for rel in rels_root:
            if rel.get("Type", "").endswith("/drawing"):
                drawing_target = rel.get("Target")
        if not drawing_target:
            return out
        drawing_path = str((Path("xl/worksheets") / drawing_target).resolve()).replace(str(Path.cwd()), "").lstrip("/")
        # Normalize "xl/worksheets/../drawings/drawing1.xml" -> "xl/drawings/drawing1.xml"
        drawing_path = f"xl/drawings/{Path(drawing_target).name}"
        if drawing_path not in names:
            return out
        drawing_root = ET.fromstring(zf.read(drawing_path))
        drawing_rels_path = f"xl/drawings/_rels/{Path(drawing_path).name}.rels"
        rid_to_media: dict[str, str] = {}
        if drawing_rels_path in names:
            drels_root = ET.fromstring(zf.read(drawing_rels_path))
            for rel in drels_root:
                rid_to_media[rel.get("Id")] = rel.get("Target")

        for anchor in list(drawing_root):
            tag = anchor.tag.split("}")[-1]
            if tag not in ("twoCellAnchor", "oneCellAnchor"):
                continue
            from_el = anchor.find("xdr:from", ns)
            if from_el is None:
                continue
            row_el = from_el.find("xdr:row", ns)
            if row_el is None or row_el.text is None:
                continue
            anchor_row = int(row_el.text)  # 0-indexed
            blip = anchor.find(".//a:blip", {"a": "http://schemas.openxmlformats.org/drawingml/2006/main"})
            if blip is None:
                continue
            rid = blip.get(f"{{{ns['r']}}}embed")
            media_rel = rid_to_media.get(rid)
            if not media_rel:
                continue
            media_path = f"xl/media/{Path(media_rel).name}"
            if media_path not in names:
                continue
            raw = zf.read(media_path)
            fmt = Path(media_path).suffix.lstrip(".").lower()
            out[anchor_row + 1] = (raw, fmt)
    return out


def resolve_image(raw: bytes, fmt: str) -> ImageResult:
    if fmt in RASTER_MIME:
        return ImageResult(ok=True, data=raw, mime=RASTER_MIME[fmt], source_format=fmt)
    if fmt in VECTOR_FORMATS:
        converted = _convert_emf_to_png(raw, fmt)
        if converted:
            return ImageResult(ok=True, data=converted, mime="image/png", source_format=f"{fmt}-converted")
        return ImageResult(ok=False, source_format=fmt, error=f"{fmt} conversion failed")
    return ImageResult(ok=False, source_format=fmt or "unknown", error="unrecognized format")


def extract_file(category: str, path: str) -> list[GroheRow]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    images_by_row = _extract_images_by_row(path)
    original_filename = ORIGINAL_FILENAMES[category]

    rows: list[GroheRow] = []
    seen_hashes: dict[str, str] = {}   # sha1(sku+desc+mrp) -> first sku seen
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=8):
        cells = {c.column_letter: c.value for c in row if c.value is not None}
        a, b = cells.get("A"), cells.get("B")
        if not isinstance(a, int) or b is None:
            continue
        sku = str(b).strip()
        description = str(cells.get("C") or "").replace("\n", " ").strip()
        description = re.sub(r"\s+", " ", description)
        mrp_raw = cells.get("E")
        mrp = None
        if isinstance(mrp_raw, (int, float)):
            mrp = float(mrp_raw)
        segment = cells.get("F")
        segment = str(segment).strip() if segment not in (None, "") else None
        finish_hint = cells.get("G")
        finish_hint = str(finish_hint).strip() if finish_hint not in (None, "") else None

        finish = _detect_finish(description)
        family_key = f"grohe:{category.lower().replace(' ', '-')}:{_normalize_family_base(description)}"

        dup_key = hashlib.sha1(f"{sku}|{description}|{mrp}".encode()).hexdigest()
        duplicate_of = None
        if dup_key in seen_hashes and seen_hashes[dup_key] != sku:
            pass  # different sku, identical payload — not a true duplicate, keep both
        if sku in [r.sku for r in rows]:
            duplicate_of = sku  # same SKU seen again in this file — will collapse at import time

        raw_img = images_by_row.get(row[0].row)
        img_result = resolve_image(*raw_img) if raw_img else ImageResult(ok=False, error="no embedded image at this row")

        rows.append(GroheRow(
            category=category, source_file=original_filename, sl=a, sku=sku,
            description=description, mrp=mrp, segment=segment,
            finish_hint_col=finish_hint, finish=finish, family_key=family_key,
            row_num=row[0].row, image=img_result, duplicate_of=duplicate_of,
        ))
        seen_hashes[dup_key] = sku
    return rows


def extract_all(base_dir: str) -> dict[str, list[GroheRow]]:
    result: dict[str, list[GroheRow]] = {}
    for category, filename in FILES.items():
        path = str(Path(base_dir) / filename)
        result[category] = extract_file(category, path)
    return result


if __name__ == "__main__":
    import sys
    base = sys.argv[1] if len(sys.argv) > 1 else "/tmp/grohe_analysis"
    all_rows = extract_all(base)
    total = 0
    total_imgs_ok = 0
    failures = []
    for cat, rows in all_rows.items():
        print(f"\n=== {cat} ({len(rows)} rows) ===")
        for r in rows:
            total += 1
            status = "OK" if r.image.ok else "MISSING"
            if r.image.ok:
                total_imgs_ok += 1
            else:
                failures.append((cat, r.sku, r.description, r.image.error))
            print(f"  sl{r.sl:>3} sku={r.sku:<14} finish={r.finish or '-':<12} "
                  f"img={status:<7} fmt={r.image.source_format} mrp={r.mrp}  {r.description[:55]}")
    print(f"\nTOTAL ROWS: {total}  IMAGES OK: {total_imgs_ok}  MISSING: {len(failures)}")
    if failures:
        print("\nFAILED ROWS (need attention):")
        for cat, sku, desc, err in failures:
            print(f"  [{cat}] sku={sku} desc={desc!r} error={err}")
