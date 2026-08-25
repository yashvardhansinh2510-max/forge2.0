"""Extract embedded workbook images and attach them to imported products."""
from __future__ import annotations

import asyncio
import sys
from pathlib import Path

from dotenv import load_dotenv
from openpyxl import load_workbook

BASE = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(BASE))
load_dotenv(BASE / ".env")

from db import db  # noqa: E402
from services import catalog_service  # noqa: E402
from services.media_service import upload_and_register  # noqa: E402

FLOOR_ID = "ground-floor"
WORKBOOKS = (
    ("Aurica", Path("/Users/yashvardhansinhjhala/Library/Containers/net.whatsapp.WhatsApp/Data/tmp/documents/8A812B22-6AB6-4FDF-908B-A96DDD460157/AURICA 2026.xlsx")),
    ("Casa Bath", Path("/Users/yashvardhansinhjhala/Library/Containers/net.whatsapp.WhatsApp/Data/tmp/documents/BEE84DD2-8F14-4149-A691-8A27E2F17E0B/CASA BATH 2026.xlsx")),
    ("Crystal Sanitation", Path("/Users/yashvardhansinhjhala/Library/Containers/net.whatsapp.WhatsApp/Data/tmp/documents/69000FAC-3422-4BFC-BC3E-5723A18876E2/CRYSTAL SANITATION 2026.xlsx")),
)


def extract_images(path: Path) -> dict[int, tuple[bytes, str]]:
    wb = load_workbook(path, data_only=True, read_only=False)
    ws = wb[wb.sheetnames[0]]
    result: dict[int, tuple[bytes, str]] = {}
    for image in ws._images:
        row = image.anchor._from.row + 1
        fmt = (image.format or "jpeg").lower()
        mime = "image/png" if fmt == "png" else "image/jpeg"
        result[row] = (image._data(), mime)
    return result


async def main() -> None:
    summary = {"attached": 0, "already_present": 0, "missing_workbook_images": [], "missing_products": []}
    for brand_name, workbook in WORKBOOKS:
        brand = await db.brands.find_one({"floor_id": FLOOR_ID, "name": brand_name}, {"_id": 0})
        if not brand:
            raise RuntimeError(f"Missing imported brand: {brand_name}")
        images = extract_images(workbook)
        wb = load_workbook(workbook, data_only=True, read_only=True)
        ws = wb[wb.sheetnames[0]]
        for source_row, values in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            serial, name, _image_cell, _size, _rate = (list(values) + [None] * 5)[:5]
            if name is None or str(name).strip() == "":
                continue
            sku = f"{brand_name.upper().replace(' ', '-').replace('_', '-')}-2026-{int(serial):03d}"
            product = await db.products.find_one(
                {"floor_id": FLOOR_ID, "brand_id": brand["id"], "sku": sku}, {"_id": 0}
            )
            if not product:
                summary["missing_products"].append({"brand": brand_name, "sku": sku})
                continue
            if source_row not in images:
                summary["missing_workbook_images"].append({"brand": brand_name, "source_row": source_row, "sku": sku})
                continue
            existing = await db.product_media.find_one(
                {"product_id": product["id"], "is_primary": True}, {"_id": 0, "id": 1}
            )
            if existing:
                summary["already_present"] += 1
                continue
            data, mime = images[source_row]
            await upload_and_register(
                data=data,
                mime=mime,
                brand_slug=brand["slug"],
                product_id=product["id"],
                family_key=product.get("family_key"),
                brand_id=brand["id"],
                floor_id=FLOOR_ID,
                source_type="supplier",
                role="hero",
                is_primary=True,
                sort_order=0,
                notes=f"embedded image from {workbook.name}, source row {source_row}",
            )
            await db.products.update_one(
                {"id": product["id"]},
                {"$set": {"specs.image_supplied": True}},
            )
            summary["attached"] += 1
        # A retried upload can pass the read-before-write dedupe check twice.
        # Keep one metadata row per product, and make exactly one row primary.
        products = await db.products.find({"floor_id": FLOOR_ID, "brand_id": brand["id"]}, {"_id": 0, "id": 1}).to_list(200)
        for product in products:
            media = await db.product_media.find({"product_id": product["id"]}, {"_id": 0}).sort("created_at", 1).to_list(20)
            if not media:
                continue
            keep = media[0]
            await db.product_media.update_many(
                {"product_id": product["id"]},
                {"$set": {"is_primary": False}},
            )
            await db.product_media.update_one({"id": keep["id"]}, {"$set": {"is_primary": True, "sort_order": 0}})
            duplicate_ids = [row["id"] for row in media[1:]]
            if duplicate_ids:
                # Duplicate rows share the same content-addressed storage key;
                # retain the object and remove only redundant metadata rows.
                await db.product_media.delete_many({"id": {"$in": duplicate_ids}})
            await db.products.update_one({"id": product["id"]}, {"$set": {"specs.image_supplied": True}})
    catalog_service.schedule_catalog_refresh()
    summary["products_with_primary_media"] = await db.product_media.count_documents(
        {"floor_id": FLOOR_ID, "brand_id": {"$in": [
            b["id"] for b in await db.brands.find({"floor_id": FLOOR_ID, "name": {"$in": ["Aurica", "Casa Bath", "Crystal Sanitation"]}}, {"_id": 0, "id": 1}).to_list(3)
        ]}, "is_primary": True}
    )
    print(summary)


if __name__ == "__main__":
    asyncio.run(main())
