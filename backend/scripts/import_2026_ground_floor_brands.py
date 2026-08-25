"""Import Aurica, Casa Bath, and Crystal Sanitation 2026 workbooks.

The workbooks contain product names, sizes, and one supplier rate per row;
they do not contain SKUs or image payloads.  The importer therefore creates a
stable source-row SKU and preserves the supplied rate text in ``specs`` rather
than inventing a different unit or image URL.  Re-running the script is safe:
existing rows are matched by (ground-floor, brand, source-row SKU).
"""
from __future__ import annotations

import asyncio
import re
import sys
from pathlib import Path
from uuid import uuid4

from dotenv import load_dotenv
from openpyxl import load_workbook

BASE = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(BASE))
load_dotenv(BASE / ".env")

from db import db  # noqa: E402
from models import Brand, Category, Product  # noqa: E402

FLOOR_ID = "ground-floor"
CATEGORY_NAME = "Tiles"
WORKBOOKS = (
    ("Aurica", Path("/Users/yashvardhansinhjhala/Library/Containers/net.whatsapp.WhatsApp/Data/tmp/documents/8A812B22-6AB6-4FDF-908B-A96DDD460157/AURICA 2026.xlsx")),
    ("Casa Bath", Path("/Users/yashvardhansinhjhala/Library/Containers/net.whatsapp.WhatsApp/Data/tmp/documents/BEE84DD2-8F14-4149-A691-8A27E2F17E0B/CASA BATH 2026.xlsx")),
    ("Crystal Sanitation", Path("/Users/yashvardhansinhjhala/Library/Containers/net.whatsapp.WhatsApp/Data/tmp/documents/69000FAC-3422-4BFC-BC3E-5723A18876E2/CRYSTAL SANITATION 2026.xlsx")),
)


def slug(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "-", value.strip().lower()).strip("-")


def parse_rate(raw: object) -> tuple[float, str]:
    text = str(raw or "").strip()
    match = re.search(r"[0-9]+(?:\.[0-9]+)?", text.replace(",", ""))
    if not match:
        raise ValueError(f"No numeric rate found in {raw!r}")
    return float(match.group()), text.upper()


def read_rows(path: Path) -> list[dict]:
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = []
    for row_number, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        serial, name, image, size, rate = (list(row) + [None] * 5)[:5]
        if name is None or str(name).strip() == "":
            continue
        price, rate_text = parse_rate(rate)
        rows.append({
            "source_row": row_number,
            "serial": serial,
            "name": str(name).strip(),
            "size": str(size).strip() if size is not None else None,
            "price": price,
            "rate_text": rate_text,
            "image_supplied": image is not None,
        })
    return rows


async def get_or_create(collection, query: dict, factory):
    existing = await collection.find_one(query, {"_id": 0})
    if existing:
        return existing, False
    doc = factory()
    await collection.insert_one(doc)
    return doc, True


async def main() -> None:
    category, category_created = await get_or_create(
        db.categories,
        {"floor_id": FLOOR_ID, "slug": "tiles"},
        lambda: Category(id=str(uuid4()), name=CATEGORY_NAME, slug="tiles", floor_id=FLOOR_ID).model_dump(),
    )
    summary = {"category_created": category_created, "brands": {}, "total_inserted": 0, "total_existing": 0}

    for brand_name, workbook in WORKBOOKS:
        brand_slug = slug(brand_name)
        brand, brand_created = await get_or_create(
            db.brands,
            {"floor_id": FLOOR_ID, "slug": brand_slug},
            lambda: Brand(id=str(uuid4()), name=brand_name, slug=brand_slug, floor_id=FLOOR_ID).model_dump(),
        )
        rows = read_rows(workbook)
        inserted = existing = 0
        for row in rows:
            sku = f"{brand_slug.upper()}-2026-{int(row['serial']):03d}"
            query = {"floor_id": FLOOR_ID, "brand_id": brand["id"], "sku": sku}
            if await db.products.find_one(query, {"_id": 1}):
                existing += 1
                continue
            family_key = slug(row["name"])
            product = Product(
                floor_id=FLOOR_ID,
                name=row["name"],
                sku=sku,
                brand_id=brand["id"],
                category_id=category["id"],
                family_key=family_key,
                family_name=row["name"],
                variant_label=row["name"],
                description=row["name"],
                size=row["size"],
                dimensions=row["size"],
                mrp=row["price"],
                price=row["price"],
                specs={
                    "source_file": workbook.name,
                    "source_row": row["source_row"],
                    "source_serial": row["serial"],
                    "supplier_rate_text": row["rate_text"],
                    "rate_unit": "sqft" if "SQFT" in row["rate_text"] else "pcs",
                    "image_supplied": row["image_supplied"],
                },
                tags=[brand_slug, "tiles", "ground-floor", "2026-import"],
                active=True,
            )
            await db.products.insert_one(product.model_dump())
            inserted += 1
        summary["brands"][brand_name] = {
            "brand_id": brand["id"],
            "brand_created": brand_created,
            "source_rows": len(rows),
            "inserted": inserted,
            "existing": existing,
        }
        summary["total_inserted"] += inserted
        summary["total_existing"] += existing

    # Verify exact floor/brand counts after the write.
    summary["ground_floor_product_count"] = await db.products.count_documents({"floor_id": FLOOR_ID})
    summary["ground_floor_brands"] = await db.brands.count_documents({"floor_id": FLOOR_ID})
    print(summary)


if __name__ == "__main__":
    asyncio.run(main())
